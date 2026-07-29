"""Gaji relawan, insentif, pengembalian & pengajuan dana mitra."""
from datetime import date
from typing import Any, Dict, List, Optional

from app.constants import (
    FEE_PAYROL_PER_ORANG,
    INSENTIF_MITRA_JUMLAH,
    PENGEMBALIAN_DANA_JUMLAH,
    TAGIHAN_CHARGES_AMOUNT,
)
from app.services.tagihan import get_all_tagihan
from app.utils.formatters import _parse_id_date, _parse_slash_date
from app.db import get_db

def _sort_key_no(item: Dict) -> int:
    try:
        raw = str(item.get("no") or "").strip()
        return int(raw) if raw else 999_999
    except (ValueError, TypeError):
        return 999_999


def get_gaji_relawan(filters: Dict = None) -> List[Dict]:
    f = dict(filters or {})
    f["kategori"] = "gaji_relawan"
    rows = get_all_tagihan(f)
    return sorted(rows, key=_sort_key_no)


def get_gaji_staff(filters: Dict = None) -> List[Dict]:
    f = dict(filters or {})
    f["kategori"] = "gaji_staff"
    rows = get_all_tagihan(f)
    return sorted(rows, key=_sort_key_no)


def calc_fee_payrol(item_count: int) -> int:
    return FEE_PAYROL_PER_ORANG * max(item_count, 0)


def calc_tagihan_charges(item: Dict) -> int:
    """Biaya transfer Rp6.500 per baris, kecuali bank Mandiri."""
    bank = (item.get("bank") or "").strip().lower()
    if "mandiri" in bank:
        return 0
    return TAGIHAN_CHARGES_AMOUNT


def format_tagihan_rekening_export(item: Dict) -> str:
    """REKENING untuk export: gabung nomor rekening + atas nama (bukan label kategori)."""
    nomor = str(item.get("nomor_rekening") or "").strip()
    atas = str(item.get("atas_nama") or "").strip()
    bank = str(item.get("bank") or "").strip()
    if nomor and atas:
        prefix = f"{bank} " if bank else ""
        return f"{prefix}{nomor} a.n. {atas}".strip()
    if nomor:
        return f"{bank} {nomor}".strip() if bank else nomor
    if atas:
        return atas
    return str(item.get("rekening") or "").strip()


def enrich_tagihan_item(item: Dict) -> Dict:
    row = dict(item)
    jumlah = row.get("jumlah") or 0
    charges = calc_tagihan_charges(row)
    row["charges"] = charges
    row["total"] = jumlah + charges
    row["rekening_export"] = format_tagihan_rekening_export(row)
    return row


def get_gaji_relawan_laporans(active_only: bool = False) -> List[Dict[str, Any]]:
    """Daftar laporan gaji relawan. active_only=True: hanya upload yang masih punya data."""
    conn = get_db()
    query = """
        SELECT g.*, u.record_count, u.created_at AS upload_at,
               (SELECT COUNT(*) FROM tagihan t
                WHERE t.upload_id = g.upload_id AND t.kategori = 'gaji_relawan') AS item_count
        FROM gaji_relawan_laporan g
        LEFT JOIN uploads u ON u.id = g.upload_id
    """
    if active_only:
        query += """
        WHERE EXISTS (
            SELECT 1 FROM tagihan t
            WHERE t.upload_id = g.upload_id AND t.kategori = 'gaji_relawan'
        )
        """
    query += " ORDER BY g.id DESC"
    rows = conn.execute(query).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def recalc_gaji_relawan_laporan(conn, upload_id: int):
    """Hitung ulang total laporan setelah hapus baris — metadata upload tetap ada."""
    if not upload_id:
        return
    rows = conn.execute(
        "SELECT jumlah FROM tagihan WHERE upload_id = ? AND kategori = 'gaji_relawan'",
        (upload_id,),
    ).fetchall()
    count = len(rows)
    total = sum(r[0] or 0 for r in rows)
    if count == 0:
        conn.execute("DELETE FROM gaji_relawan_laporan WHERE upload_id = ?", (upload_id,))
    else:
        conn.execute("""
            UPDATE gaji_relawan_laporan SET jumlah_penerima = ?, total_gaji = ?
            WHERE upload_id = ?
        """, (count, total, upload_id))
    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (count, upload_id))


def _strip_nama_gelar(nama: str) -> str:
    """Hilangkan awalan gelar (IBU, BPK, BPA, SDRI, SDR, dll) dari nama relawan."""
    import re
    if not nama:
        return nama
    return re.sub(
        r"^(IBU|BPK|BPA|SDRI|SDR|TN|NY|NN)\.?\s+",
        "",
        nama.strip(),
        flags=re.I,
    ).strip()


def _parse_pic_periode(filename: str) -> Optional[str]:
    import re
    m = re.search(r"PERIODE\s*(\d+)", filename, re.I)
    if m:
        return f"Periode {m.group(1)}"
    m = re.search(r"(\d{1,2}\s+\w+\s+\d{4})", filename.replace("_", " "), re.I)
    if m:
        return m.group(1)
    return None


def _parse_mandiri_csv_date(raw: str) -> Optional[str]:
    raw = (raw or "").strip()
    if len(raw) == 8 and raw.isdigit():
        return f"{raw[0:4]}-{raw[4:6]}-{raw[6:8]}"
    return _parse_id_date(raw)


def _pic_cell_str(val) -> str:
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y%m%d")
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    if isinstance(val, int):
        return str(val)
    return str(val).strip()


def _pic_parse_amount(val) -> int:
    if val is None or val == "":
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(str(val).strip().replace(".", "").replace(",", "") or 0)
    except ValueError:
        return 0


def _parse_pic_transfer_rows(raw_rows: List, filename: str = "", default_label: str = "Gaji Relawan") -> Dict[str, Any]:
    """Parse baris format transfer massal Mandiri (header P + data per kolom)."""
    meta = {
        "tanggal_pembayaran": None,
        "rekening_sumber": None,
        "jumlah_penerima": 0,
        "total_gaji": 0,
        "periode": _parse_pic_periode(filename),
        "bank": "MANDIRI",
        "kota": "Madiun",
        "filename": filename,
    }
    items: List[Dict[str, Any]] = []

    rows = [[_pic_cell_str(c) for c in row] for row in raw_rows if row is not None]
    if not rows:
        return {"meta": meta, "items": items}

    header = rows[0]
    if header and (header[0] or "").strip().upper() == "P":
        meta["tanggal_pembayaran"] = _parse_mandiri_csv_date(header[1] if len(header) > 1 else "")
        meta["rekening_sumber"] = (header[2] if len(header) > 2 else "").strip() or None
        try:
            meta["jumlah_penerima"] = int((header[3] if len(header) > 3 else "0").strip() or 0)
        except ValueError:
            meta["jumlah_penerima"] = 0
        meta["total_gaji"] = _pic_parse_amount(header[4] if len(header) > 4 else 0)
        data_rows = rows[1:]
    else:
        data_rows = rows

    for row in data_rows:
        if not row or not any(cell for cell in row):
            continue
        nomor_rek = (row[0] if len(row) > 0 else "").strip()
        nama = _strip_nama_gelar((row[1] if len(row) > 1 else "").strip())
        if not nama and not nomor_rek:
            continue
        if not nama:
            continue

        jumlah = _pic_parse_amount(row[6] if len(row) > 6 else 0)
        if jumlah <= 0:
            continue

        bank = (row[11] if len(row) > 11 else "MANDIRI").strip() or "MANDIRI"
        kota = (row[12] if len(row) > 12 else "").strip() or meta.get("kota")
        if kota:
            meta["kota"] = kota

        periode_label = meta.get("periode") or default_label
        items.append({
            "no": str(len(items) + 1),
            "pengajuan": f"{nama} — {periode_label}",
            "atas_nama": nama,
            "nomor_rekening": nomor_rek or None,
            "bank": bank,
            "jumlah": jumlah,
            "tanggal": meta.get("tanggal_pembayaran"),
            "rekening": "TRANSFER MASSAL",
            "status": "DIAJUKAN",
        })

    if items and not meta["total_gaji"]:
        meta["total_gaji"] = sum(i["jumlah"] for i in items)
    if items and not meta["jumlah_penerima"]:
        meta["jumlah_penerima"] = len(items)

    return {"meta": meta, "items": items}


def parse_pic_transfer_csv(file_path: str, filename: str = "", default_label: str = "Gaji Relawan") -> Dict[str, Any]:
    """Parse CSV format transfer massal Mandiri (PIC), contoh CSV PIC PERIODE10."""
    import csv

    with open(file_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    return _parse_pic_transfer_rows(rows, filename, default_label)


def parse_pic_transfer_xlsx(file_path: str, filename: str = "", default_label: str = "Gaji Relawan") -> Dict[str, Any]:
    """Parse Excel (.xlsx/.xls) dengan struktur kolom yang sama seperti CSV transfer massal Mandiri."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return _parse_pic_transfer_rows(rows, filename, default_label)


PIC_TRANSFER_COL_COUNT = 43


def _nama_from_gaji_item(item: Dict) -> str:
    atas = (item.get("atas_nama") or "").strip()
    if atas:
        return atas
    pengajuan = item.get("pengajuan") or ""
    if " — " in pengajuan:
        return pengajuan.split(" — ", 1)[0].strip()
    return pengajuan.strip()


def _export_filename_from_laporan(laporan: Dict, ext: str) -> str:
    import re

    raw = (laporan.get("filename") or laporan.get("periode") or "gaji_relawan").strip()
    raw = re.sub(r'[<>:"/\\|?*]', "_", raw)
    base = raw.rsplit(".", 1)[0] if "." in raw else raw
    if ext == "csv" and base.lower().endswith((".xlsx", ".xls")):
        base = base.rsplit(".", 1)[0]
    return f"{base}.{ext}"


def resolve_gaji_relawan_export(upload_id: int = 0) -> tuple:
    """Ambil laporan + baris untuk export sesuai batch upload."""
    laporans = get_gaji_relawan_laporans(active_only=True)
    laporan = None

    if upload_id:
        laporan = next((lap for lap in laporans if lap["upload_id"] == upload_id), None)
        if not laporan:
            conn = get_db()
            row = conn.execute(
                """
                SELECT g.*, u.filename AS upload_filename
                FROM gaji_relawan_laporan g
                LEFT JOIN uploads u ON u.id = g.upload_id
                WHERE g.upload_id = ?
                """,
                (upload_id,),
            ).fetchone()
            conn.close()
            if row:
                laporan = dict(row)
    elif laporans:
        laporan = laporans[0]
        upload_id = laporan["upload_id"]

    if not laporan or not upload_id:
        return None, [], 0

    items = get_gaji_relawan({"upload_id": upload_id})
    return laporan, items, upload_id


def build_pic_transfer_export_rows(laporan: Dict, items: List[Dict]) -> List[List]:
    """Bangun baris export format CSV transfer massal Mandiri (sama seperti file upload)."""
    tgl_raw = (laporan.get("tanggal_pembayaran") or "").strip()
    tgl_export = tgl_raw.replace("-", "") if len(tgl_raw) == 10 and tgl_raw[4] == "-" else tgl_raw.replace("-", "")

    total_gaji = sum(int(i.get("jumlah") or 0) for i in items)
    jumlah_penerima = len(items)

    header = [
        "P",
        tgl_export,
        str(laporan.get("rekening_sumber") or ""),
        str(jumlah_penerima),
        str(total_gaji),
    ]
    header += [""] * (PIC_TRANSFER_COL_COUNT - len(header))
    rows = [header[:PIC_TRANSFER_COL_COUNT]]

    bank = laporan.get("bank") or "MANDIRI"
    kota = laporan.get("kota") or "Madiun"

    for item in items:
        row = [""] * PIC_TRANSFER_COL_COUNT
        row[0] = str(item.get("nomor_rekening") or "")
        row[1] = _nama_from_gaji_item(item)
        row[6] = str(int(item.get("jumlah") or 0))
        row[9] = "IBU"
        row[11] = bank
        row[12] = kota
        row[16] = "N"
        row[37] = "OUR"
        row[38] = "1"
        row[39] = "E"
        rows.append(row)
    return rows


def build_pic_transfer_xlsx_bytes(laporan: Dict, items: List[Dict]) -> bytes:
    """Bangun file Excel format transfer massal Mandiri — struktur sama dengan file upload."""
    from openpyxl import Workbook
    from io import BytesIO

    rows = build_pic_transfer_export_rows(laporan, items)
    wb = Workbook()
    ws = wb.active
    ws.title = "Transfer Massal"

    numeric_cols = {4, 5, 7}
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell_val = val
            if c_idx in numeric_cols and str(val).isdigit():
                cell_val = int(val)
            cell = ws.cell(row=r_idx, column=c_idx, value=cell_val)
            if isinstance(cell_val, int):
                cell.number_format = "#,##0"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def parse_gaji_relawan_csv(file_path: str, filename: str = "") -> Dict[str, Any]:
    return parse_pic_transfer_csv(file_path, filename, "Gaji Relawan")


def parse_gaji_relawan_xlsx(file_path: str, filename: str = "") -> Dict[str, Any]:
    return parse_pic_transfer_xlsx(file_path, filename, "Gaji Relawan")


def parse_insentif_pic_csv(file_path: str, filename: str = "") -> Dict[str, Any]:
    return parse_pic_transfer_csv(file_path, filename, "Insentif PIC")


def parse_insentif_pic_xlsx(file_path: str, filename: str = "") -> Dict[str, Any]:
    return parse_pic_transfer_xlsx(file_path, filename, "Insentif PIC")


def resolve_insentif_pic_export(upload_id: int = 0) -> tuple:
    """Ambil laporan + baris untuk export sesuai batch upload Insentif PIC."""
    laporans = get_insentif_pic_laporans(active_only=True)
    laporan = None

    if upload_id:
        laporan = next((lap for lap in laporans if lap["upload_id"] == upload_id), None)
        if not laporan:
            conn = get_db()
            row = conn.execute(
                """
                SELECT g.*, u.filename AS upload_filename
                FROM insentif_pic_laporan g
                LEFT JOIN uploads u ON u.id = g.upload_id
                WHERE g.upload_id = ?
                """,
                (upload_id,),
            ).fetchone()
            conn.close()
            if row:
                laporan = dict(row)
    elif laporans:
        laporan = laporans[0]
        upload_id = laporan["upload_id"]

    if not laporan or not upload_id:
        return None, [], 0

    items = get_insentif_pic({"upload_id": upload_id})
    return laporan, items, upload_id


def get_insentif_pic(filters: Dict = None) -> List[Dict]:
    f = dict(filters or {})
    f["kategori"] = "insentif_pic"
    rows = get_all_tagihan(f)
    return sorted(rows, key=_sort_key_no)


def get_insentif_pic_laporans(active_only: bool = False) -> List[Dict[str, Any]]:
    conn = get_db()
    query = """
        SELECT g.*, u.record_count, u.created_at AS upload_at,
               (SELECT COUNT(*) FROM tagihan t
                WHERE t.upload_id = g.upload_id AND t.kategori = 'insentif_pic') AS item_count
        FROM insentif_pic_laporan g
        LEFT JOIN uploads u ON u.id = g.upload_id
    """
    if active_only:
        query += """
        WHERE EXISTS (
            SELECT 1 FROM tagihan t
            WHERE t.upload_id = g.upload_id AND t.kategori = 'insentif_pic'
        )
        """
    query += " ORDER BY g.id DESC"
    rows = conn.execute(query).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def recalc_insentif_pic_laporan(conn, upload_id: int):
    if not upload_id:
        return
    rows = conn.execute(
        "SELECT jumlah FROM tagihan WHERE upload_id = ? AND kategori = 'insentif_pic'",
        (upload_id,),
    ).fetchall()
    count = len(rows)
    total = sum(r[0] or 0 for r in rows)
    if count == 0:
        conn.execute("DELETE FROM insentif_pic_laporan WHERE upload_id = ?", (upload_id,))
    else:
        conn.execute("""
            UPDATE insentif_pic_laporan SET jumlah_penerima = ?, total_gaji = ?
            WHERE upload_id = ?
        """, (count, total, upload_id))
    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (count, upload_id))




def parse_insentif_mitra_csv(file_path: str, filename: str = "") -> Dict[str, Any]:
    return parse_pic_transfer_csv(file_path, filename, "Insentif Mitra")


def parse_insentif_mitra_xlsx(file_path: str, filename: str = "") -> Dict[str, Any]:
    return parse_pic_transfer_xlsx(file_path, filename, "Insentif Mitra")


def resolve_insentif_mitra_export(upload_id: int = 0) -> tuple:
    """Ambil laporan + baris untuk export sesuai batch upload Insentif Mitra."""
    laporans = get_insentif_mitra_laporans(active_only=True)
    laporan = None

    if upload_id:
        laporan = next((lap for lap in laporans if lap["upload_id"] == upload_id), None)
        if not laporan:
            conn = get_db()
            row = conn.execute(
                """
                SELECT g.*, u.filename AS upload_filename
                FROM insentif_mitra_laporan g
                LEFT JOIN uploads u ON u.id = g.upload_id
                WHERE g.upload_id = ?
                """,
                (upload_id,),
            ).fetchone()
            conn.close()
            if row:
                laporan = dict(row)
    elif laporans:
        laporan = laporans[0]
        upload_id = laporan["upload_id"]

    if not laporan or not upload_id:
        return None, [], 0

    items = get_insentif_mitra({"upload_id": upload_id})
    return laporan, items, upload_id


def get_insentif_mitra(filters: Dict = None) -> List[Dict]:
    f = dict(filters or {})
    f["kategori"] = "insentif_mitra"
    rows = get_all_tagihan(f)
    return sorted(rows, key=_sort_key_no)


def get_insentif_mitra_laporans(active_only: bool = False) -> List[Dict[str, Any]]:
    conn = get_db()
    query = """
        SELECT g.*, u.record_count, u.created_at AS upload_at,
               (SELECT COUNT(*) FROM tagihan t
                WHERE t.upload_id = g.upload_id AND t.kategori = 'insentif_mitra') AS item_count
        FROM insentif_mitra_laporan g
        LEFT JOIN uploads u ON u.id = g.upload_id
    """
    if active_only:
        query += """
        WHERE EXISTS (
            SELECT 1 FROM tagihan t
            WHERE t.upload_id = g.upload_id AND t.kategori = 'insentif_mitra'
        )
        """
    query += " ORDER BY g.id DESC"
    rows = conn.execute(query).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def recalc_insentif_mitra_laporan(conn, upload_id: int):
    if not upload_id:
        return
    rows = conn.execute(
        "SELECT jumlah FROM tagihan WHERE upload_id = ? AND kategori = 'insentif_mitra'",
        (upload_id,),
    ).fetchall()
    count = len(rows)
    total = sum(r[0] or 0 for r in rows)
    if count == 0:
        conn.execute("DELETE FROM insentif_mitra_laporan WHERE upload_id = ?", (upload_id,))
    else:
        conn.execute("""
            UPDATE insentif_mitra_laporan SET jumlah_penerima = ?, total_gaji = ?
            WHERE upload_id = ?
        """, (count, total, upload_id))
    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (count, upload_id))


def get_pengembalian_dana(filters: Dict = None) -> List[Dict]:
    f = dict(filters or {})
    f["kategori"] = "pengembalian_dana"
    rows = get_all_tagihan(f)
    return sorted(rows, key=_sort_key_no)


def get_pengembalian_dana_laporans(active_only: bool = False) -> List[Dict[str, Any]]:
    conn = get_db()
    query = """
        SELECT g.*, u.record_count, u.created_at AS upload_at,
               (SELECT COUNT(*) FROM tagihan t
                WHERE t.upload_id = g.upload_id AND t.kategori = 'pengembalian_dana') AS item_count
        FROM pengembalian_dana_laporan g
        LEFT JOIN uploads u ON u.id = g.upload_id
    """
    if active_only:
        query += """
        WHERE EXISTS (
            SELECT 1 FROM tagihan t
            WHERE t.upload_id = g.upload_id AND t.kategori = 'pengembalian_dana'
        )
        """
    query += " ORDER BY g.id DESC"
    rows = conn.execute(query).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def recalc_pengembalian_dana_laporan(conn, upload_id: int):
    if not upload_id:
        return
    rows = conn.execute(
        "SELECT jumlah FROM tagihan WHERE upload_id = ? AND kategori = 'pengembalian_dana'",
        (upload_id,),
    ).fetchall()
    count = len(rows)
    total = sum(r[0] or 0 for r in rows)
    if count == 0:
        conn.execute("DELETE FROM pengembalian_dana_laporan WHERE upload_id = ?", (upload_id,))
    else:
        conn.execute("""
            UPDATE pengembalian_dana_laporan SET jumlah_penerima = ?, total_gaji = ?
            WHERE upload_id = ?
        """, (count, total, upload_id))
    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (count, upload_id))


def get_pengajuan_dana_mitra_tanggal_options() -> List[str]:
    """Daftar tanggal pengajuan unik dari laporan PDF & entri manual."""
    conn = get_db()
    dates = set()
    for row in conn.execute("""
        SELECT DISTINCT tanggal_pembayaran AS t FROM pengajuan_dana_mitra_laporan
        WHERE tanggal_pembayaran IS NOT NULL AND tanggal_pembayaran != ''
        UNION
        SELECT DISTINCT tanggal AS t FROM tagihan
        WHERE kategori = 'pengajuan_dana_mitra' AND tanggal IS NOT NULL AND tanggal != ''
    """).fetchall():
        if row[0]:
            dates.add(row[0])
    conn.close()
    return sorted(dates, reverse=True)


def get_pengeluaran_mitra(filters: Dict = None) -> List[Dict]:
    f = dict(filters or {})
    f["kategori"] = "pengeluaran_mitra"
    rows = get_all_tagihan(f)
    return sorted(rows, key=_sort_key_no)


def get_pengajuan_dana_mitra(filters: Dict = None) -> List[Dict]:
    """Filter tanggal = Tanggal Pengajuan dari formulir PDF (header upload)."""
    f = dict(filters or {})
    tanggal_pengajuan = f.pop("tanggal", None)
    search = f.pop("search", None)
    status = f.pop("status", None)
    rekening = f.pop("rekening", None)
    upload_id = f.pop("upload_id", None)

    conn = get_db()
    query = """
        SELECT t.* FROM tagihan t
        LEFT JOIN pengajuan_dana_mitra_laporan g ON g.upload_id = t.upload_id
    """
    where = ["t.kategori = 'pengajuan_dana_mitra'"]
    params: List[Any] = []

    if search:
        where.append("(t.pengajuan LIKE ? OR t.atas_nama LIKE ? OR t.bank LIKE ?)")
        s = f"%{search}%"
        params.extend([s, s, s])
    if status:
        where.append("t.status = ?")
        params.append(status)
    if rekening:
        where.append("t.rekening = ?")
        params.append(rekening)
    if tanggal_pengajuan:
        where.append("(t.tanggal = ? OR g.tanggal_pembayaran = ?)")
        params.extend([tanggal_pengajuan, tanggal_pengajuan])
    if upload_id:
        where.append("t.upload_id = ?")
        params.append(upload_id)

    query += " WHERE " + " AND ".join(where)
    query += " ORDER BY COALESCE(t.tanggal, g.tanggal_pembayaran, '9999-12-31') DESC, t.id DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_pengajuan_dana_mitra_laporans(active_only: bool = False) -> List[Dict[str, Any]]:
    conn = get_db()
    query = """
        SELECT g.*, u.record_count, u.created_at AS upload_at,
               (SELECT COUNT(*) FROM tagihan t
                WHERE t.upload_id = g.upload_id AND t.kategori = 'pengajuan_dana_mitra') AS item_count
        FROM pengajuan_dana_mitra_laporan g
        LEFT JOIN uploads u ON u.id = g.upload_id
    """
    if active_only:
        query += """
        WHERE EXISTS (
            SELECT 1 FROM tagihan t
            WHERE t.upload_id = g.upload_id AND t.kategori = 'pengajuan_dana_mitra'
        )
        """
    query += " ORDER BY g.id DESC"
    rows = conn.execute(query).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def recalc_pengajuan_dana_mitra_laporan(conn, upload_id: int):
    if not upload_id:
        return
    rows = conn.execute(
        "SELECT jumlah FROM tagihan WHERE upload_id = ? AND kategori = 'pengajuan_dana_mitra'",
        (upload_id,),
    ).fetchall()
    count = len(rows)
    total = sum(r[0] or 0 for r in rows)
    if count == 0:
        conn.execute("DELETE FROM pengajuan_dana_mitra_laporan WHERE upload_id = ?", (upload_id,))
    else:
        conn.execute("""
            UPDATE pengajuan_dana_mitra_laporan SET jumlah_penerima = ?, total_gaji = ?
            WHERE upload_id = ?
        """, (count, total, upload_id))
    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (count, upload_id))
