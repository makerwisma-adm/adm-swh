"""File upload parsers (CSV, XLSX, PDF)."""
import csv
import os
import re
from typing import Any, Dict, List, Optional

from app.services.transfer_reports import (
    parse_gaji_relawan_csv,
    parse_gaji_relawan_xlsx,
    parse_insentif_mitra_csv,
    parse_insentif_mitra_xlsx,
    parse_insentif_pic_csv,
    parse_insentif_pic_xlsx,
    parse_pic_transfer_csv,
    parse_pic_transfer_xlsx,
)
from app.utils.formatters import _parse_id_date, _parse_rp_amount


def _extract_buku_besar_lines(pdf_path: str) -> List[str]:
    """Extract clean text lines from Accurate Buku Besar PDF."""
    import pdfplumber
    skip_exact = {
        "SPPG WISMA HAJI", "Rincian Buku Besar", "Filter berdasarkan : Kode Perkiraan",
        "Tanggal Tipe Transaksi Keterangan Debit Kredit Saldo Akhir", "11010101",
    }
    lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for raw in (page.extract_text() or "").split("\n"):
                line = raw.strip()
                if not line:
                    continue
                if line.startswith("ACCURATE") or "Halaman" in line or "Tercetak" in line:
                    continue
                if line in skip_exact:
                    continue
                lines.append(line)
    return lines


def parse_buku_besar_petty_cash(pdf_path: str, upload_id: int, filename: str = "") -> Dict[str, Any]:
    """Parse Accurate 'Rincian Buku Besar' PDF untuk akun PETTY CASH (11010101)."""
    import re

    meta = {
        "report_type": "buku_besar",
        "nama_karyawan": "11010101 — PETTY CASH",
        "divisi": "SPPG Wisma Haji Madiun",
        "saldo_awal": 0,
        "sisa_dana": 0,
        "saldo_akhir": 0,
        "total_digantikan": 0,
        "total_debit": 0,
        "total_kredit": 0,
        "payment_info": "Sumber: Accurate Accounting — Rincian Buku Besar",
        "bank": None,
        "nomor_rekening": None,
        "atas_nama": None,
        "periode": None,
        "filename": filename,
    }

    all_lines = _extract_buku_besar_lines(pdf_path)
    for line in all_lines:
        if line.startswith("Dari ") and "s/d" in line:
            meta["periode"] = line.replace("Dari ", "").strip()
            break
    if not meta["periode"]:
        periode_m = re.search(r"(\d{1,2}\s+\w+\s+\d{4})", filename.replace("_", " "), re.I)
        if periode_m:
            meta["periode"] = periode_m.group(1)

    tx_re = re.compile(
        r"^(\d{1,2}\s+\w{3}\s+\d{4})\s+(.+?)\s+(-?[\d.,]+)\s+(-?[\d.,]+)\s+(-?[\d.,]+)$"
    )

    items = []
    pending = None

    def _commit_pending():
        nonlocal pending
        if not pending or "debit" not in pending:
            pending = None
            return
        pending["no"] = str(len(items) + 1)
        pending["status"] = "DIAJUKAN"
        pending["rekening"] = "PETTY CASH"
        pending["jumlah"] = pending["kredit"] if pending["kredit"] > 0 else pending["debit"]
        items.append(pending)
        pending = None

    for line in all_lines:
        if line.startswith("11010101 -"):
            continue
        if re.match(r"^Dari \d", line):
            continue

        if re.match(r"^\d{1,2}\s+\w{3}\s+\d{4}\s+Saldo per", line):
            nums = re.findall(r"-?[\d.,]+", line)
            if nums:
                meta["saldo_awal"] = _parse_rp_amount(nums[-1])
            continue

        if re.match(r"^[\d.,]+\s+[\d.,]+$", line) and len(line.split()) == 2:
            parts = line.split()
            meta["total_debit"] = _parse_rp_amount(parts[0])
            meta["total_kredit"] = _parse_rp_amount(parts[1])
            continue

        m = tx_re.match(line)
        if m:
            _commit_pending()
            tanggal_raw = m.group(1)
            body = m.group(2).strip()
            debit = _parse_rp_amount(m.group(3))
            kredit = _parse_rp_amount(m.group(4))
            saldo = _parse_rp_amount(m.group(5))

            tipe = "Jurnal Umum"
            keterangan = body
            if body.startswith("Transfer Bank"):
                tipe = "Transfer Bank"
                keterangan = re.sub(r"^Transfer Bank\s*", "", body).strip()
            elif body.startswith("Jurnal Umum"):
                keterangan = body.replace("Jurnal Umum", "", 1).strip()

            pending = {
                "tanggal": _parse_id_date(tanggal_raw),
                "tanggal_display": tanggal_raw,
                "tipe_transaksi": tipe,
                "pengajuan": keterangan,
                "debit": debit,
                "kredit": kredit,
                "saldo_akhir": saldo,
            }
            continue

        if pending and "pengajuan" in pending:
            pending["pengajuan"] = (pending["pengajuan"] + " " + line).strip()

    _commit_pending()

    if items:
        meta["saldo_akhir"] = items[-1].get("saldo_akhir") or 0
        meta["sisa_dana"] = meta["saldo_akhir"]
    if not meta["total_kredit"]:
        meta["total_kredit"] = sum(i["kredit"] for i in items)
    if not meta["total_debit"]:
        meta["total_debit"] = sum(i["debit"] for i in items)
    meta["total_digantikan"] = meta["total_kredit"]

    return {"meta": meta, "items": items}


def parse_reimbursement_petty_cash(pdf_path: str, upload_id: int, filename: str = "") -> Dict[str, Any]:
    """Parse FORM REIMBURSEMENT / FORM PETTY CASH + lampiran nota per transaksi."""
    import re
    import pdfplumber

    meta = {
        "report_type": "reimbursement",
        "nama_karyawan": None,
        "divisi": None,
        "saldo_awal": 0,
        "sisa_dana": 0,
        "saldo_akhir": 0,
        "total_digantikan": 0,
        "total_debit": 0,
        "total_kredit": 0,
        "payment_info": None,
        "bank": None,
        "nomor_rekening": None,
        "atas_nama": None,
        "yang_menyetujui": None,
        "tanggal_ttd_pemohon": None,
        "tanggal_ttd_menyetujui": None,
        "periode": None,
        "filename": filename,
    }

    import re
    fname_norm = filename.replace("_", " ")
    periode_m = re.search(
        r"(?:FORM\s+PETTY\s+CASH|REIMBURSEMENT|FORM)\s+(\d{1,2}\s+\w+\s+\d{4})",
        fname_norm, re.IGNORECASE,
    )
    if not periode_m:
        periode_m = re.search(r"(\d{1,2}\s+\w+\s+\d{4})", fname_norm, re.IGNORECASE)
    if periode_m:
        meta["periode"] = periode_m.group(1)

    skip_prefixes = (
        "FORM REIMBURSEMENT", "Tanggal Deskripsi", "PAYMENT",
    )
    awaiting_approver = False
    ttd_date_re = re.compile(r"(\d{1,2}/\d{1,2}/\d{4})")
    # Format: "1 20 Juli 2026 Pembelian bensin Rp300,000" — number, date, desc, Rp amount
    tx_re = re.compile(
        r"^(\d+)\s+(\d{1,2}\s+\w+\s+\d{4})\s+(.+?)\s+Rp?\s*([\d,.]+)$",
        re.IGNORECASE,
    )
    # Fallback: number, date, desc, amount WITHOUT Rp prefix
    tx_re_no_prefix = re.compile(
        r"^(\d+)\s+(\d{1,2}\s+\w+\s+\d{4})\s+(.+?)\s+([\d,.]+)$",
    )

    items = []
    with pdfplumber.open(pdf_path) as pdf:
        if not pdf.pages:
            return {"meta": meta, "items": items}
        lines = [l.strip() for l in (pdf.pages[0].extract_text() or "").split("\n") if l.strip()]

    # Scan items: accumulate transactions
    for line in lines:
        if line.startswith("Nama Karyawan"):
            meta["nama_karyawan"] = line.replace("Nama Karyawan", "").strip()
        elif line.startswith("Divisi"):
            meta["divisi"] = line.replace("Divisi", "").strip()
        elif line.startswith("Saldo"):
            meta["saldo_awal"] = _parse_rp_amount(line)
        elif line.startswith("Sisa Dana"):
            amt = _parse_rp_amount(line.replace("-", ""))
            meta["sisa_dana"] = -amt if "-" in line else amt
        elif "Total yang Digantikan" in line:
            meta["total_digantikan"] = _parse_rp_amount(line)
        elif line.upper().startswith("PAYMENT"):
            meta["payment_info"] = line
            bm = re.search(r"(MANDIRI|BRI|BCA|BSI|BNI)\s+(\d+)", line, re.I)
            if bm:
                meta["bank"] = bm.group(1).upper()
                meta["nomor_rekening"] = bm.group(2)
            am = re.search(r"AN\s+(.+)$", line, re.I)
            if am:
                meta["atas_nama"] = am.group(1).strip()
            continue
        elif "Tanda Tangan Pemohon" in line:
            dm = ttd_date_re.search(line)
            if dm:
                meta["tanggal_ttd_pemohon"] = _parse_slash_date(dm.group(1))
            continue
        elif "yang menyetujui" in line.lower():
            dm = ttd_date_re.search(line)
            if dm:
                meta["tanggal_ttd_menyetujui"] = _parse_slash_date(dm.group(1))
            awaiting_approver = True
            continue
        elif awaiting_approver and not meta["yang_menyetujui"]:
            if not line.startswith("Tanda Tangan"):
                meta["yang_menyetujui"] = line.strip()
            awaiting_approver = False
            continue

        if any(line.startswith(p) for p in skip_prefixes):
            continue
        if line.startswith("Tanda Tangan"):
            continue
        if meta["nama_karyawan"] and line.upper() == meta["nama_karyawan"].upper():
            continue

        m = tx_re.match(line)
        if not m:
            m = tx_re_no_prefix.match(line)
        if m:
            no_val = m.group(1)
            date_val = m.group(2)
            desc = m.group(3).strip()
            raw_amt = m.group(4)
            # skip if desc looks like a date
            if re.match(r"^\d{1,2}\s+\w+\s+\d{4}$", desc):
                continue
            amt = _parse_rp_amount(raw_amt)
            if amt <= 0:
                continue
            items.append({
                "no": no_val,
                "tanggal": _parse_id_date(date_val),
                "tanggal_display": date_val,
                "deskripsi": desc,
                "pengajuan": desc,
                "jumlah": amt,
                "kredit": amt,
                "debit": 0,
                "status": "DIAJUKAN",
                "rekening": "PETTY CASH",
                "tipe_transaksi": "Pengeluaran",
            })

    meta["total_kredit"] = sum(i["kredit"] for i in items)
    if not meta["total_digantikan"]:
        meta["total_digantikan"] = meta["total_kredit"]
    meta["saldo_akhir"] = meta["sisa_dana"]

    return {"meta": meta, "items": items}


def parse_petty_cash_pdf(pdf_path: str, upload_id: int, filename: str = "") -> Dict[str, Any]:
    """Auto-detect format petty cash: Buku Besar Accurate atau Form Reimbursement."""
    import pdfplumber
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return {"meta": {}, "items": []}
            first_text = pdf.pages[0].extract_text() or ""
    except Exception as e:
        print("Petty cash PDF open error:", e)
        return {"meta": {}, "items": []}

    fname_up = filename.upper()
    if "FORM REIMBURSEMENT" in first_text or "FORM PETTY CASH" in fname_up or "REIMBURSEMENT" in fname_up:
        return parse_reimbursement_petty_cash(pdf_path, upload_id, filename)
    if "Rincian Buku Besar" in first_text or "11010101 - PETTY CASH" in first_text:
        return parse_buku_besar_petty_cash(pdf_path, upload_id, filename)
    return parse_reimbursement_petty_cash(pdf_path, upload_id, filename)

def parse_faktur_belum_lunas(pdf_path: str):
    """Parse PDF 'Faktur Belum Lunas' (Accurate) ke struktur kolom tagihan web.

    Mapping kolom PDF → database:
      - Baris sebelum PI.*           → pos (pemasok)
      - PI.2026.06.xxxxx             → no
      - Tanggal pertama pada baris PI → tanggal
      - PEMBELIAN ... / keterangan   → pengajuan
      - Jumlah (Total Utang)         → jumlah
      - MANDIRI/BRI/BCA/...          → bank
      - Nomor setelah bank           → nomor_rekening
      - Nama setelah nomor rekening  → atas_nama
    """
    import pdfplumber
    import re
    from dateutil import parser as date_parser

    BANK_RE = re.compile(r"\b(MANDIRI|BRI|BCA|BSI|VA|BNI)\s+(\d{5,})\s+(.+)$", re.IGNORECASE)
    PI_FULL_RE = re.compile(
        r"^(PI\.\d{4}\.\d{2}\.\d{5})\s+(\d{1,2}\s+\w+\s+\d{4})\s+(\d{1,2}\s+\w+\s+\d{4})\s+(.+)$"
    )
    PI_SHORT_RE = re.compile(r"^(PI\.\d{4}\.\d{2}\.\d{5})\s+(\d{1,2}\s+\w+\s+\d{4})\s+(.+)$")
    AMOUNT_RE = re.compile(r"(\d{1,3}(?:\.\d{3})+)")
    SUBTOTAL_RE = re.compile(r"^\d{1,3}(?:\.\d{3})+(?:\.\d{3})*\s+\d")
    SKIP_KEYWORDS = [
        "sppg wisma haji", "faktur belum lunas", "accurate", "tercetak", "halaman",
        "indonesian", "nomor #", "cabang :", "per tgl", "total utang", "rupiah",
        "daftar rekening pemasok", "jatuh tempo keterangan",
    ]

    def _parse_date(tgl_str: str):
        if not tgl_str:
            return None
        tgl = tgl_str.strip()
        norm = tgl.lower()
        for indo, eng in [("mei", "may"), ("agu", "aug"), ("agt", "aug"), ("okt", "oct"), ("des", "dec")]:
            norm = norm.replace(indo, eng)
        try:
            dt = date_parser.parse(norm, dayfirst=True, fuzzy=True)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
        try:
            p = tgl.lower().split()
            d = int(p[0])
            mon_str = p[1][:3]
            mon_map = {
                "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "mei": 5,
                "jun": 6, "jul": 7, "aug": 8, "agu": 8, "sep": 9,
                "oct": 10, "okt": 10, "nov": 11, "dec": 12, "des": 12,
            }
            mon = mon_map.get(mon_str, 1)
            y = int(p[2])
            return f"{y:04d}-{mon:02d}-{d:02d}"
        except Exception:
            return None

    def _should_skip(line: str) -> bool:
        low = line.lower()
        return any(kw in low for kw in SKIP_KEYWORDS)

    def _is_subtotal(line: str) -> bool:
        return bool(SUBTOTAL_RE.match(line))

    def _is_keterangan_continuation(line: str) -> bool:
        if line.startswith("PI.") or _is_subtotal(line) or _should_skip(line):
            return False
        if AMOUNT_RE.search(line) or BANK_RE.search(line):
            return False
        return len(line) <= 40

    def _is_supplier_line(line: str) -> bool:
        if line.startswith("PI.") or _is_subtotal(line) or _should_skip(line):
            return False
        if len(line) <= 2 or re.match(r"^\d", line):
            return False
        low = line.lower()
        if any(kw in low for kw in ["pembelian", "total", "peb", "report", "0 0", "rupiah", "wifi bulan"]):
            return False
        if re.search(r"\d", line) and len(line) < 28:
            return False
        if "," in line and len(line) < 30:
            return False
        return True

    def _is_supplier_continuation(line: str) -> bool:
        if line.startswith("PI.") or _is_subtotal(line) or _should_skip(line):
            return False
        if AMOUNT_RE.search(line) or BANK_RE.search(line):
            return False
        return len(line.split()) <= 3 and len(line) <= 24

    def _parse_pi_line(line: str, pemasok: str):
        m = PI_FULL_RE.match(line)
        if m:
            no, tgl_str, _, rest = m.group(1), m.group(2), m.group(3), m.group(4)
        else:
            m = PI_SHORT_RE.match(line)
            if not m:
                return None
            no, tgl_str, rest = m.group(1), m.group(2), m.group(3)

        tanggal = _parse_date(tgl_str)
        bank = ""
        rek = ""
        atas_nama = ""
        bank_m = BANK_RE.search(rest)
        if bank_m:
            bank = bank_m.group(1).upper()
            rek = bank_m.group(2)
            atas_nama = re.sub(r"\s+\d+\s+\d+\s*$", "", bank_m.group(3).strip()).strip()
            rest = rest[: bank_m.start()].strip()

        amount_m = AMOUNT_RE.search(rest)
        jumlah = 0
        keterangan = rest
        if amount_m:
            try:
                jumlah = int(amount_m.group(1).replace(".", ""))
            except Exception:
                jumlah = 0
            keterangan = rest[: amount_m.start()].strip()

        keterangan = re.sub(r"\s+", " ", keterangan).strip()
        if not keterangan or keterangan.isdigit():
            keterangan = f"Tagihan {pemasok}" if pemasok else f"Tagihan {no}"
        if jumlah <= 0:
            return None

        return {
            "no": no,
            "pos": pemasok or None,
            "pengajuan": keterangan,
            "jumlah": jumlah,
            "tanggal": tanggal,
            "bank": bank,
            "nomor_rekening": rek,
            "atas_nama": atas_nama or None,
            "status": "DIAJUKAN",
        }

    with pdfplumber.open(pdf_path) as pdf:
        all_lines = []
        for page in pdf.pages:
            text = page.extract_text() or ""
            all_lines.extend([l.strip() for l in text.split("\n") if l.strip()])

    items = []
    current_supplier = None
    pending_supplier_parts: List[str] = []
    awaiting_keterangan = False

    for line in all_lines:
        if not line or _should_skip(line) or _is_subtotal(line):
            awaiting_keterangan = False
            continue

        if line.startswith("PI."):
            pemasok = " ".join(pending_supplier_parts).strip() if pending_supplier_parts else current_supplier
            item = _parse_pi_line(line, pemasok)
            if item:
                items.append(item)
                if pemasok:
                    current_supplier = pemasok
                awaiting_keterangan = True
            pending_supplier_parts = []
            continue

        if awaiting_keterangan and items and _is_keterangan_continuation(line):
            items[-1]["pengajuan"] = f"{items[-1]['pengajuan']} {line}".strip()
            continue

        awaiting_keterangan = False

        if pending_supplier_parts and _is_supplier_continuation(line):
            pending_supplier_parts.append(line)
            continue

        if _is_supplier_line(line):
            pending_supplier_parts = [line]
            continue

    return items




_BANK_KEYS = ("MANDIRI", "BRI", "BCA", "BSI", "BNI", "BTN", "DANAMON", "CIMB", "PERMATA")

_HEADER_ANCHORS = {
    # urutan = posisi relatif anchor di header baris pertama tabel
    "no": ("no",),
    "tanggal": ("tanggal",),
    "deskripsi": ("deskripsi",),
    "jumlah": ("jumlah", "rp"),
    "norek": ("no.", "rekening", "no.rekening"),
    "nama": ("nama", "rekening"),
    "bank": ("bank",),
    "ket": ("keterangan",),
}


def _detect_column_boundaries(page) -> Optional[Dict[str, tuple]]:
    """Deteksi koordinat X kolom dari header baris tabel.

    Mencari baris yang mengandung anchor header (No, Tanggal, Deskripsi,
    Jumlah, No.Rekening, Nama Rekening, Bank, Keterangan). Boundary tiap
    kolom = titik tengah antara anchor ini dan anchor sebelah kanannya.
    Toleransi Y dilonggarkan ke 10pt agar multi-baris header (mis.
    "Jumlah" + "(Rp)") tetap dianggap satu grup.
    """
    try:
        words = page.extract_words(keep_blank_chars=False, use_text_flow=True)
    except TypeError:
        words = page.extract_words()
    if not words:
        return None

    candidates: List[Dict[str, Any]] = []
    for w in words:
        t = w["text"].lower().strip(".:()")
        if not t:
            continue
        if t in ("no",):
            candidates.append({"word": w, "col": "no", "x0": w["x0"]})
        elif t == "tanggal":
            candidates.append({"word": w, "col": "tanggal", "x0": w["x0"]})
        elif t == "deskripsi":
            candidates.append({"word": w, "col": "deskripsi", "x0": w["x0"]})
        elif t in ("jumlah", "rp"):
            candidates.append({"word": w, "col": "jumlah", "x0": w["x0"]})
        elif t in ("no.rekening", "rekening"):
            candidates.append({"word": w, "col": "rekening_x", "x0": w["x0"]})
        elif t == "nama":
            candidates.append({"word": w, "col": "nama_x", "x0": w["x0"]})
        elif t == "bank":
            candidates.append({"word": w, "col": "bank", "x0": w["x0"]})
        elif t == "keterangan":
            candidates.append({"word": w, "col": "ket", "x0": w["x0"]})

    if not candidates:
        return None

    # Cluster berdasarkan Y dengan toleransi 10pt
    sorted_cands = sorted(candidates, key=lambda c: c["word"]["top"])
    clusters: List[List[Dict[str, Any]]] = []
    current: List[Dict[str, Any]] = []
    current_top: Optional[float] = None
    for c in sorted_cands:
        if current_top is None or abs(c["word"]["top"] - current_top) <= 10.0:
            current.append(c)
            current_top = c["word"]["top"] if current_top is None else (current_top + c["word"]["top"]) / 2
        else:
            clusters.append(current)
            current = [c]
            current_top = c["word"]["top"]
    if current:
        clusters.append(current)

    # Pilih cluster header tabel: harus punya 'no', 'tanggal', 'deskripsi'.
    # Ambil yang paling BAWAH (supaya header form di atas tidak ikut).
    table_header_candidates = [
        cl for cl in clusters
        if "no" in {c["col"] for c in cl}
        and "tanggal" in {c["col"] for c in cl}
        and "deskripsi" in {c["col"] for c in cl}
    ]
    header_row = table_header_candidates[-1] if table_header_candidates else max(clusters, key=len)

    header_top = sum(c["word"]["top"] for c in header_row) / len(header_row)
    # Tarik anchor Jumlah yang beda baris (mis. '(Rp)') — toleransi 15pt
    for c in candidates:
        if c["col"] == "jumlah" and abs(c["word"]["top"] - header_top) <= 15:
            header_row.append(c)

    sorted_x = sorted(header_row, key=lambda c: c["x0"])

    col_x: Dict[str, float] = {}
    norek_no_anchor_x: Optional[float] = None
    rekening_xs: List[float] = []
    for c in sorted_x:
        col = c["col"]
        x0 = c["x0"]
        if col == "no" and x0 > 60:
            # Anchor 'no' di kolom kanan (untuk 'No. Rekening')
            norek_no_anchor_x = x0
            continue
        if col in ("tanggal", "deskripsi", "jumlah", "bank", "ket"):
            if col not in col_x:
                col_x[col] = x0
        elif col == "nama_x":
            col_x["nama"] = x0
        elif col == "rekening_x":
            rekening_xs.append(x0)
    if "no" not in col_x:
        # Cari anchor 'no' terkiri
        no_left = min(
            (c["x0"] for c in header_row if c["col"] == "no"),
            default=None,
        )
        if no_left is not None:
            col_x["no"] = no_left

    # norek: gunakan norek_no_anchor_x ("No.") kalau ada, else rekening_xs[0]
    norek_x = norek_no_anchor_x or (rekening_xs[0] if rekening_xs else None)
    if norek_x is not None:
        col_x["norek"] = norek_x

    # nama: rekening_xs[1] (Rekening ke-2 milik 'Nama Rekening'),
    # atau norek_x + offset jika tidak ada
    if len(rekening_xs) >= 2:
        col_x["nama"] = rekening_xs[1]
    elif "nama" not in col_x and norek_x is not None:
        col_x["nama"] = norek_x + 80

    if "no" not in col_x or "tanggal" not in col_x or "deskripsi" not in col_x:
        return None
    if "jumlah" not in col_x or "norek" not in col_x or "nama" not in col_x:
        return None

    anchor_order = ["no", "tanggal", "deskripsi", "jumlah", "norek", "nama", "bank", "ket"]
    present = [a for a in anchor_order if a in col_x]

    bounds: Dict[str, tuple] = {}
    for i, key in enumerate(present):
        if i == 0:
            left = 0
        else:
            left = (col_x[present[i - 1]] + col_x[key]) / 2
        if i + 1 >= len(present):
            right = float("inf")
        else:
            right = (col_x[key] + col_x[present[i + 1]]) / 2
        bounds[key] = (left, right)

    return bounds


def _classify_pdf_words(words: List[Dict[str, Any]], bounds: Dict[str, tuple]) -> Dict[str, List[Dict[str, Any]]]:
    cols = {k: [] for k in ["no", "tanggal", "deskripsi", "jumlah",
                             "norek", "nama", "bank", "ket"]}
    for w in words:
        x0 = w["x0"]
        for key, (lo, hi) in bounds.items():
            if lo <= x0 < hi:
                cols[key].append(w)
                break
    return cols


def _row_y_groups(words: List[Dict[str, Any]], tolerance: float = 3.0) -> List[List[Dict[str, Any]]]:
    """Kelompokkan kata-kata ke baris berdasarkan posisi Y (top)."""
    if not words:
        return []
    sorted_words = sorted(words, key=lambda w: (round(w["top"] / tolerance), w["x0"]))
    rows: List[List[Dict[str, Any]]] = []
    current: List[Dict[str, Any]] = []
    current_top = None
    for w in sorted_words:
        if current_top is None or abs(w["top"] - current_top) <= tolerance:
            current.append(w)
            current_top = w["top"] if current_top is None else (current_top + w["top"]) / 2
        else:
            rows.append(current)
            current = [w]
            current_top = w["top"]
    if current:
        rows.append(current)
    return rows


def parse_pengajuan_dana_mitra_pdf(pdf_path: str, filename: str = "") -> Dict[str, Any]:
    """Parse formulir PDF 'Pengajuan Dana Mitra SPPG Wisma Haji'.

    Strategi: layout-based extraction dari kolom di halaman pertama saja.
    Baris data dikenali dari anchor kolom 'No' (angka 1-7). Teks yang
    deskripsinya multi-baris digabung otomatis. Halaman 2 (footer ttd)
    di-skip.
    """
    import re
    import pdfplumber

    meta: Dict[str, Any] = {"filename": filename}
    items: List[Dict[str, Any]] = []
    total_pengajuan = 0

    with pdfplumber.open(pdf_path) as pdf:
        if not pdf.pages:
            return {"meta": meta, "items": items}

        page = pdf.pages[0]
        full_text = page.extract_text() or ""
        all_lines = [l.strip() for l in full_text.split("\n") if l.strip()]

        if not any("pengajuan dana mitra" in ln.lower() for ln in all_lines):
            return {"meta": {}, "items": []}

        # --- Extract meta dari teks (halaman 1) ---
        for line in all_lines:
            low = line.lower()
            if "no. form" in low or "no form" in low:
                m = re.search(r"No\.?\s*Form:?\s*([^\s]+)", line, re.IGNORECASE)
                if m:
                    meta["no_form"] = m.group(1)
                m2 = re.search(
                    r"Tanggal\s+Pengajuan:?\s*(.+)$",
                    line,
                    re.IGNORECASE,
                )
                if m2:
                    meta["tanggal_pengajuan"] = _parse_id_date(m2.group(1).strip())
            if low.startswith("pemohon") or "pemohon:" in low:
                m = re.search(r"Pemohon:?\s*(.+?)(?:\s+Devisi|$)", line, re.IGNORECASE)
                if m:
                    name = m.group(1).strip()
                    if name and "," not in name:
                        meta["pemohon"] = name
                m2 = re.search(r"Devisi:?\s*(.+)$", line, re.IGNORECASE)
                if m2:
                    meta["divisi"] = m2.group(1).strip()
            if "total pengajuan" in low:
                m = re.search(r"(\d{1,3}(?:\.\d{3})+)", line)
                if m:
                    try:
                        total_pengajuan = int(m.group(1).replace(".", ""))
                    except ValueError:
                        pass

        # --- Layout-based row extraction ---
        try:
            words = page.extract_words(keep_blank_chars=False, use_text_flow=True)
        except TypeError:
            words = page.extract_words()

        bounds = _detect_column_boundaries(page)
        if not bounds:
            return {"meta": meta, "items": items}

        cols = _classify_pdf_words(words, bounds)

        no_word_min_x = min(
            (w["x0"] for w in cols["no"] if w["text"].isdigit() and 1 <= int(w["text"]) <= 7),
            default=999,
        )
        no_words = sorted(
            [
                w for w in cols["no"]
                if w["text"].isdigit()
                and 1 <= int(w["text"]) <= 7
                and w["x0"] < no_word_min_x + 15
            ],
            key=lambda w: (w["top"], w["x0"]),
        )
        no_numbers: List[Dict[str, Any]] = []
        for row_w in no_words:
            no_numbers.append({"no": int(row_w["text"]), "top": row_w["top"], "x0": row_w["x0"]})

        deduped: List[Dict[str, Any]] = []
        for anchor in no_numbers:
            if deduped and abs(anchor["top"] - deduped[-1]["top"]) <= 2.0:
                if anchor["x0"] < deduped[-1]["x0"]:
                    deduped[-1] = anchor
                continue
            deduped.append(anchor)
        no_numbers = deduped

        def _words_in_y(words: List[Dict[str, Any]], y_top: float, y_bottom: float):
            return [w for w in words if y_top <= w["top"] <= y_bottom]

        month_token_set = {
            "jan", "feb", "mar", "apr", "mei", "jun", "jul",
            "agu", "agt", "aug", "sep", "okt", "oct", "nov", "des", "dec",
            "januari", "februari", "maret", "april", "juni", "juli",
            "agustus", "september", "oktober", "november", "desember",
        }

        # Sort by Y. Pass 1: field terstruktur di pita sempit di sekitar nomor baris.
        # Pass 2: teks deskripsi multi-baris di-assign ke baris terdekat (bukan midpoint
        # kaku) agar overflow baris tidak "nyasar" ke baris lain. Token tanggal
        # di-exclude by posisi (bukan by teks) supaya "Juni" di deskripsi tidak hilang.
        no_numbers.sort(key=lambda a: a["top"])

        def _word_key(w: Dict[str, Any]) -> tuple:
            return (round(w["top"], 1), round(w["x0"], 1), w["text"])

        def _is_same_word(a: Optional[Dict[str, Any]], b: Dict[str, Any]) -> bool:
            if not a:
                return False
            return (
                abs(a["top"] - b["top"]) < 1.0
                and abs(a["x0"] - b["x0"]) < 1.0
                and a["text"] == b["text"]
            )

        rows_buf: List[Dict[str, Any]] = []
        date_token_keys = set()

        for idx, anchor in enumerate(no_numbers):
            no_val = anchor["no"]
            # Pita ketat untuk field sejajar nomor baris (tanggal/jumlah/rek/bank/nama)
            y_top = anchor["top"] - 4
            y_bottom = anchor["top"] + 12
            row_words = _words_in_y(words, y_top, y_bottom)

            day_tok = month_tok = year_tok = None
            for w in sorted(row_words, key=lambda w: w["x0"]):
                t = w["text"]
                if t == str(no_val) and abs(w["x0"] - anchor["x0"]) < 5:
                    continue
                if day_tok is None and t.isdigit() and 1 <= int(t) <= 31 and w["x0"] < 145:
                    day_tok = w
                elif month_tok is None and t.lower()[:3] in month_token_set and w["x0"] < 200:
                    month_tok = w
                elif year_tok is None and t.isdigit() and len(t) == 4 and t.startswith("20"):
                    year_tok = w

            for tok in (day_tok, month_tok, year_tok):
                if tok:
                    date_token_keys.add(_word_key(tok))

            tgl_text = ""
            if day_tok and month_tok and year_tok:
                tgl_text = f"{day_tok['text']} {month_tok['text']} {year_tok['text']}"

            jumlah = 0
            for w in sorted(row_words, key=lambda w: w["x0"]):
                raw = w["text"].replace(".", "").replace(",", "")
                if raw.isdigit() and 4 <= len(raw) <= 12 and 350 <= w["x0"] <= 500:
                    try:
                        v = int(raw)
                        if v > jumlah:
                            jumlah = v
                    except ValueError:
                        pass

            nomor_rekening = ""
            for w in sorted(row_words, key=lambda w: w["x0"]):
                if w["text"].isdigit() and len(w["text"]) >= 8 and 470 <= w["x0"] <= 560:
                    nomor_rekening = w["text"]
                    break

            bank = ""
            for w in sorted(row_words, key=lambda w: w["x0"]):
                if w["text"].upper() in _BANK_KEYS and w["x0"] > 600:
                    bank = w["text"].upper()
                    break

            exclude_for_nama = {
                day_tok["text"] if day_tok else None,
                month_tok["text"] if month_tok else None,
                year_tok["text"] if year_tok else None,
                nomor_rekening,
                str(no_val),
                bank,
            }
            nama_tokens = []
            for w in sorted(row_words, key=lambda w: w["x0"]):
                t = w["text"]
                if 560 <= w["x0"] <= 700 and t not in exclude_for_nama:
                    if t.upper() in _BANK_KEYS or t.isdigit():
                        continue
                    nama_tokens.append(w)
            atas_nama = " ".join(w["text"] for w in nama_tokens).strip()

            # Keterangan kolom kanan (jika terisi di PDF)
            ket_left = bounds.get("bank", (0, 0))[1] if bounds.get("bank") else 720
            ket_tokens = [
                w for w in row_words
                if w["x0"] >= ket_left and not w["text"].isdigit()
                and w["text"].upper() not in _BANK_KEYS
            ]
            ket = " ".join(w["text"] for w in sorted(ket_tokens, key=lambda w: w["x0"])).strip()

            rows_buf.append({
                "no": no_val,
                "top": anchor["top"],
                "x0": anchor["x0"],
                "tgl_text": tgl_text,
                "jumlah": jumlah,
                "nomor_rekening": nomor_rekening,
                "atas_nama": atas_nama,
                "bank": bank,
                "ket": ket,
                "desc_tokens": [],
                "day_tok": day_tok,
                "month_tok": month_tok,
                "year_tok": year_tok,
            })

        # Pass 2: assign token deskripsi (area tengah) ke baris terdekat by Y
        for w in words:
            t = w["text"]
            if not (130 <= w["x0"] <= 410):
                continue
            if _word_key(w) in date_token_keys:
                continue
            # Skip nomor baris di kolom No
            if t.isdigit() and 1 <= int(t) <= 7 and w["x0"] < 100:
                continue
            if t.isdigit() and len(t) >= 4:
                continue  # jumlah / norek
            if t.upper() in _BANK_KEYS:
                continue
            # Skip token tanggal yang sama posisi
            if any(
                _is_same_word(r.get(k), w)
                for r in rows_buf
                for k in ("day_tok", "month_tok", "year_tok")
            ):
                continue

            # Baris terdekat: prefer anchor di atas atau sama Y, fallback terdekat
            best = None
            best_score = None
            for r in rows_buf:
                dy = w["top"] - r["top"]
                # Skor: jarak absolut, tapi beri penalti bila token jauh di atas anchor
                # (lebih mungkin milik baris sebelumnya)
                if dy < -10:
                    score = abs(dy) + 20
                else:
                    score = abs(dy)
                if best_score is None or score < best_score:
                    best_score = score
                    best = r
            if best is not None:
                best["desc_tokens"].append(w)

        for r in rows_buf:
            desc_tokens = sorted(r["desc_tokens"], key=lambda w: (w["top"], w["x0"]))
            # Bangun baris teks deskripsi; potongan murni dalam kurung → ket
            line_map: Dict[float, List[str]] = {}
            for w in desc_tokens:
                key = round(w["top"], 0)
                line_map.setdefault(key, []).append(w["text"])

            deskripsi_lines = []
            ket_notes = []
            for _, toks in sorted(line_map.items()):
                line = " ".join(toks).strip()
                if not line:
                    continue
                # Catatan dalam kurung murni (contoh: "(Mei)", "(110.000 dan 105.000)")
                if re.fullmatch(r"\([^)]*\)", line):
                    ket_notes.append(line.strip("() ").strip())
                else:
                    deskripsi_lines.append(line)

            deskripsi = " ".join(deskripsi_lines).strip()
            # Gabung sisa kurung di ujung deskripsi ke ket jika ada
            trailing = re.search(r"\s+(\([^)]*\))\s*$", deskripsi)
            if trailing:
                ket_notes.append(trailing.group(1).strip("() ").strip())
                deskripsi = deskripsi[: trailing.start()].strip()

            ket = r["ket"] or None
            if ket_notes:
                note = "; ".join(ket_notes)
                ket = f"{ket}; {note}" if ket else note

            if r["jumlah"] <= 0 or not deskripsi:
                continue

            items.append({
                "no": str(r["no"]),
                "tanggal": _parse_id_date(r["tgl_text"]) if r["tgl_text"] else None,
                "pengajuan": deskripsi,
                "jumlah": r["jumlah"],
                "nomor_rekening": r["nomor_rekening"] or None,
                "atas_nama": r["atas_nama"] or None,
                "bank": r["bank"] or None,
                "rekening": "PENGAJUAN DANA MITRA",
                "status": "DIAJUKAN",
                "ket": ket or None,
            })

    if meta.get("no_form") and meta.get("tanggal_pengajuan"):
        tgl_label = meta["tanggal_pengajuan"]
        if meta.get("pemohon"):
            meta["periode"] = f"Form #{meta['no_form']} — {meta['pemohon']} ({tgl_label})"
        else:
            meta["periode"] = f"Form #{meta['no_form']} — {tgl_label}"
    elif meta.get("pemohon"):
        meta["periode"] = f"Pengajuan {meta['pemohon']}"
    else:
        meta["periode"] = filename or "Pengajuan Dana Mitra"

    if total_pengajuan:
        meta["total_pengajuan"] = total_pengajuan

    return {"meta": meta, "items": items}


def parse_upload_file(file_path: str, filename: str, kategori: str = "tagihan", upload_id: int = 0):
    """Parse Excel, CSV or PDF. Use special parser per kategori."""
    rows = []
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''

    if kategori == "gaji_relawan" and ext == "csv":
        try:
            return parse_gaji_relawan_csv(file_path, filename)
        except Exception as e:
            print("Gaji relawan CSV parse error:", e)
            return {"meta": {}, "items": []}

    if kategori == "gaji_relawan" and ext in ("xlsx", "xls"):
        try:
            return parse_gaji_relawan_xlsx(file_path, filename)
        except Exception as e:
            print("Gaji relawan XLSX parse error:", e)
            return {"meta": {}, "items": []}

    if kategori == "insentif_pic" and ext == "csv":
        try:
            return parse_insentif_pic_csv(file_path, filename)
        except Exception as e:
            print("Insentif PIC CSV parse error:", e)
            return {"meta": {}, "items": []}

    if kategori == "insentif_pic" and ext in ("xlsx", "xls"):
        try:
            return parse_insentif_pic_xlsx(file_path, filename)
        except Exception as e:
            print("Insentif PIC XLSX parse error:", e)
            return {"meta": {}, "items": []}
    if kategori == "insentif_mitra" and ext == "csv":
        try:
            return parse_insentif_mitra_csv(file_path, filename)
        except Exception as e:
            print("Insentif Mitra CSV parse error:", e)
            return {"meta": {}, "items": []}

    if kategori == "insentif_mitra" and ext in ("xlsx", "xls"):
        try:
            return parse_insentif_mitra_xlsx(file_path, filename)
        except Exception as e:
            print("Insentif Mitra XLSX parse error:", e)
            return {"meta": {}, "items": []}

    if ext == 'pdf' and kategori == "petty_cash":
        try:
            result = parse_petty_cash_pdf(file_path, upload_id, filename)
            return result
        except Exception as e:
            print("Petty cash PDF parse error:", e)
            return {"meta": {}, "items": []}

    if ext == 'pdf' and kategori == "pengajuan_dana_mitra":
        try:
            return parse_pengajuan_dana_mitra_pdf(file_path, filename)
        except Exception as e:
            print("Pengajuan dana mitra PDF parse error:", e)
            return {"meta": {}, "items": []}

    if ext == 'pdf':
        # Always try the special parser for known template PDFs first
        try:
            special = parse_faktur_belum_lunas(file_path)
            if special and any(item.get("jumlah", 0) > 0 for item in special):
                rows = special
            else:
                # General fallback for other PDFs: try to pull amounts from text
                import pdfplumber
                import re
                with pdfplumber.open(file_path) as pdf:
                    all_text = ""
                    for page in pdf.pages:
                        t = page.extract_text()
                        if t:
                            all_text += t + "\n"
                    lines = [l.strip() for l in all_text.split("\n") if l.strip()]
                    for line in lines[:20]:
                        m = re.search(r"([\d\.,]{4,})", line)  # look for reasonably sized numbers
                        if m:
                            try:
                                amt = int(m.group(1).replace(".", "").replace(",", ""))
                            except:
                                amt = None
                            if amt and amt > 100:
                                rows.append({
                                    "pengajuan": line[:200],
                                    "jumlah": amt,
                                    "tanggal": None,
                                    "status": "DIAJUKAN"
                                })
                    if not rows:
                        pengajuan = "Laporan dari PDF: " + " | ".join(lines[:3])
                        rows.append({
                            "pengajuan": pengajuan[:200],
                            "jumlah": 0,
                            "tanggal": None,
                            "status": "DIAJUKAN"
                        })
        except Exception as e:
            print("PDF parse error:", e)
            rows = []
    elif ext in ['xlsx', 'xls']:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                headers = [str(h).strip().lower() if h else '' for h in row]
                continue
            if not any(row):
                continue
            item = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                item[h] = val
            rows.append(item)
    else:
        # CSV
        import csv
        with open(file_path, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append({k.strip().lower(): v for k, v in row.items() if k})

    return rows
