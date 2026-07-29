"""Pengajuan dana mitra routes."""
import os
from datetime import date
from urllib.parse import urlencode

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response

from app.auth.session import (
    _member_owns_pdm_item,
    can_member_upload,
    is_admin,
    is_viewer,
    redirect_with_flash,
    render_template,
    require_login,
)
from app.config import UPLOAD_DIR
from app.db import get_db
from app.parsers.upload import parse_pengajuan_dana_mitra_pdf
from app.services.transfer_reports import (
    get_pengajuan_dana_mitra,
    get_pengajuan_dana_mitra_laporans,
    get_pengajuan_dana_mitra_tanggal_options,
    recalc_pengajuan_dana_mitra_laporan,
)
from app.services.report_sync import build_keuangan_link_context, filter_items_by_date_range
from app.utils.formatters import format_rupiah, format_tanggal_display, format_tanggal_pengajuan

router = APIRouter()


@router.get("/pengajuan-dana-mitra", response_class=HTMLResponse)
async def pengajuan_dana_mitra_page(
    request: Request,
    user=Depends(require_login),
    upload_id: int = 0,
    search: str = "",
    status: str = "",
    rekening: str = "",
    tanggal: str = "",
    dari: str = "",
    sampai: str = "",
    message: str = "",
    success: bool = False,
    page: int = 1,
):
    periode_options = get_pengajuan_dana_mitra_laporans(active_only=True)
    laporans = periode_options
    laporan = None
    active_upload_ids = {lap["upload_id"] for lap in periode_options}

    filters = {}
    if search:
        filters["search"] = search
    if status:
        filters["status"] = status
    if rekening:
        filters["rekening"] = rekening
    if tanggal:
        filters["tanggal"] = tanggal

    if upload_id and upload_id not in active_upload_ids:
        upload_id = 0

    view_all = upload_id == 0

    if upload_id and upload_id in active_upload_ids:
        laporan = next((lap for lap in periode_options if lap["upload_id"] == upload_id), None)
        filters["upload_id"] = upload_id

    data = get_pengajuan_dana_mitra(filters)
    data = filter_items_by_date_range(data, dari, sampai)

    periode_map = {lap["upload_id"]: lap.get("periode") or lap.get("filename") for lap in laporans}
    conn = get_db()
    user_id = user.get("user_id")
    for item in data:
        uid = item.get("upload_id")
        item["periode_label"] = periode_map.get(uid, "Input Manual") if uid else "Input Manual"
        if is_admin(user):
            item["can_delete"] = True
        else:
            item["can_delete"] = _member_owns_pdm_item(conn, item["id"], user_id)

    if not data:
        main_status = "—"
    else:
        all_terbayar = all((item.get("status") or "").upper() in ("DIBAYARKAN", "TERBAYAR") for item in data)
        main_status = "Sukses" if all_terbayar else "Belum Lunas"

    pdm_filter = "kategori = 'pengajuan_dana_mitra'"
    statuses = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT status FROM tagihan WHERE status IS NOT NULL AND {pdm_filter}"
        ).fetchall()
    ]
    rekenings = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT rekening FROM tagihan WHERE rekening IS NOT NULL AND {pdm_filter}"
        ).fetchall()
    ]
    conn.close()

    total_filtered = sum(d["jumlah"] for d in data)
    avg_gaji = int(total_filtered / len(data)) if data else 0
    max_item = max(data, key=lambda x: x["jumlah"]) if data else None
    min_item = min(data, key=lambda x: x["jumlah"]) if data else None
    total_all_batches = sum(lap.get("total_gaji") or 0 for lap in laporans)
    total_all_mitra = sum(lap.get("jumlah_penerima") or lap.get("record_count") or 0 for lap in laporans)

    paid_statuses = {"DIBAYARKAN", "TERBAYAR", "LUNAS"}
    total_terbayar = sum(
        d["jumlah"] for d in data if (d.get("status") or "").upper() in paid_statuses
    )
    total_tertagih = total_filtered - total_terbayar

    # Pagination: 10 baris per halaman, dihitung dari total baris yang
    # difilter. Nilai total_* tetap dihitung dari SELURUH data, bukan per-halaman.
    per_page = 10
    total_rows = len(data)
    total_pages = max(1, (total_rows + per_page - 1) // per_page)
    try:
        page = max(1, int(page))
    except (TypeError, ValueError):
        page = 1
    page = min(page, total_pages)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paged_items = data[start_idx:end_idx]

    # Query string filter (tanpa page) agar link pagination mempertahankan filter aktif
    filter_params = {}
    if search:
        filter_params["search"] = search
    if status:
        filter_params["status"] = status
    if rekening:
        filter_params["rekening"] = rekening
    if tanggal:
        filter_params["tanggal"] = tanggal
    if (dari or "").strip():
        filter_params["dari"] = dari.strip()
    if (sampai or "").strip():
        filter_params["sampai"] = sampai.strip()
    pagination_qs = urlencode(filter_params)

    return render_template(request, "pengajuan_dana_mitra.html", {
        "user": user,
        "items": paged_items,
        "all_items_count": total_rows,
        "page": page,
        "total_pages": total_pages,
        "per_page": per_page,
        "pagination_qs": pagination_qs,
        "laporans": laporans,
        "laporan": laporan,
        "view_all": view_all,
        "selected_upload_id": upload_id if upload_id else 0,
        "total_filtered": total_filtered,
        "total_terbayar": total_terbayar,
        "total_tertagih": total_tertagih,
        "total_all_batches": total_all_batches,
        "total_all_mitra": total_all_mitra,
        "avg_gaji": avg_gaji,
        "max_item": max_item,
        "min_item": min_item,
        "main_status": main_status,
        "format_rupiah": format_rupiah,
        "format_tanggal": format_tanggal_display,
        "format_tanggal_pengajuan": format_tanggal_pengajuan,
        "filters": {
            "search": search,
            "status": status,
            "rekening": rekening,
            "tanggal": tanggal,
            "periode": upload_id,
            "dari": (dari or "").strip(),
            "sampai": (sampai or "").strip(),
        },
        **build_keuangan_link_context(dari, sampai, module_key="pengajuan_dana_mitra"),
        "periode_options": periode_options,
        "status_options": sorted([s for s in statuses if s]),
        "rekening_options": sorted([r for r in rekenings if r]),
        "tanggal_pengajuan_options": get_pengajuan_dana_mitra_tanggal_options(),
        "can_delete_upload": can_member_upload(user),
        "message": message,
        "success": success,
    })


@router.post("/pengajuan-dana-mitra/upload")
async def pengajuan_dana_mitra_upload(
    request: Request,
    user=Depends(require_login),
    file: UploadFile = File(...),
):
    """Upload PDF Formulir Pengajuan Dana Mitra — ekstrak otomatis baris rincian."""
    if is_viewer(user):
        return redirect_with_flash(
            request,
            "/pengajuan-dana-mitra",
            "Akses ditolak. Akun Anda hanya dapat melihat data.",
        )

    ext = (file.filename or "").lower().rsplit(".", 1)[-1] if "." in (file.filename or "") else ""
    if ext != "pdf":
        return redirect_with_flash(
            request,
            "/pengajuan-dana-mitra",
            "Hanya file PDF Formulir Pengajuan Dana Mitra yang didukung",
        )

    safe_name = f"{date.today().isoformat()}_{file.filename.replace(' ', '_')}"
    file_path = os.path.join(UPLOAD_DIR, safe_name)
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    conn = get_db()
    conn.execute(
        "INSERT INTO uploads (filename, created_by) VALUES (?, ?)",
        (file.filename, user["id"]),
    )
    upload_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    try:
        parsed = parse_pengajuan_dana_mitra_pdf(file_path, file.filename)
    except Exception as e:
        conn.close()
        return redirect_with_flash(request, "/pengajuan-dana-mitra", f"Error parsing file: {str(e)}")

    item_list = parsed.get("items", []) if isinstance(parsed, dict) else []
    pdm_meta = parsed.get("meta", {}) if isinstance(parsed, dict) else {}

    if not item_list:
        conn.close()
        return redirect_with_flash(
            request,
            "/pengajuan-dana-mitra",
            "Tidak ada data yang bisa dibaca dari PDF. Pastikan format Formulir Pengajuan Dana Mitra.",
        )

    inserted = 0
    for item in item_list:
        pengajuan = item.get("pengajuan") or ""
        if not pengajuan:
            continue
        jumlah = int(item.get("jumlah") or 0)
        if jumlah <= 0:
            continue

        dup = conn.execute("""
            SELECT 1 FROM tagihan
            WHERE upload_id = ? AND COALESCE(no, '') = COALESCE(?, '') AND pengajuan = ? AND jumlah = ?
            LIMIT 1
        """, (upload_id, item.get("no"), pengajuan, jumlah)).fetchone()
        if dup:
            continue

        conn.execute("""
            INSERT INTO tagihan
            (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, ket, kategori, upload_id, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pengajuan_dana_mitra', ?, ?)
        """, (
            item.get("no"),
            pengajuan,
            jumlah,
            item.get("status") or "DIAJUKAN",
            item.get("rekening") or "PENGAJUAN DANA MITRA",
            item.get("tanggal") or pdm_meta.get("tanggal_pengajuan"),
            item.get("atas_nama"),
            item.get("nomor_rekening"),
            item.get("bank"),
            item.get("ket"),
            upload_id,
            user["id"],
        ))
        inserted += 1

    total_gaji = sum(int(it.get("jumlah") or 0) for it in item_list)
    first_bank = item_list[0].get("bank") if item_list else "MANDIRI"
    first_rek = item_list[0].get("nomor_rekening") if item_list else None

    conn.execute("""
        INSERT OR REPLACE INTO pengajuan_dana_mitra_laporan
        (upload_id, tanggal_pembayaran, rekening_sumber, jumlah_penerima, total_gaji,
         periode, bank, kota, filename)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        upload_id,
        pdm_meta.get("tanggal_pengajuan"),
        first_rek,
        inserted,
        total_gaji,
        pdm_meta.get("periode") or file.filename,
        first_bank,
        pdm_meta.get("divisi"),
        file.filename,
    ))
    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (inserted, upload_id))
    conn.commit()
    conn.close()

    pemohon = pdm_meta.get("pemohon") or ""
    msg = f"Berhasil mengekstrak {inserted} baris dari PDF"
    if pemohon:
        msg += f" — Pemohon: {pemohon}"
    # Redirect ke view_all agar batch lama tetap terlihat — bukan hanya batch baru
    return redirect_with_flash(request, "/pengajuan-dana-mitra", msg, success=True)


@router.post("/pengajuan-dana-mitra")
async def create_pengajuan_dana_mitra(
    request: Request,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form("DIAJUKAN"),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        INSERT INTO tagihan (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, kategori, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pengajuan_dana_mitra', ?)
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or "DIAJUKAN",
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        user["id"],
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/pengajuan-dana-mitra", status_code=303)


@router.post("/pengajuan-dana-mitra/{item_id}/update")
async def update_pengajuan_dana_mitra(
    item_id: int,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form(""),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        UPDATE tagihan SET
            no = ?, pengajuan = ?, jumlah = ?, status = ?, rekening = ?,
            tanggal = ?, atas_nama = ?, nomor_rekening = ?, bank = ?,
            kategori = 'pengajuan_dana_mitra', updated_at = CURRENT_TIMESTAMP
        WHERE id = ? AND kategori = 'pengajuan_dana_mitra'
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or None,
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        item_id,
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/pengajuan-dana-mitra", status_code=303)


@router.post("/pengajuan-dana-mitra/{item_id}/delete")
async def delete_pengajuan_dana_mitra(item_id: int, request: Request, user=Depends(require_login)):
    conn = get_db()
    if not is_admin(user) and not _member_owns_pdm_item(conn, item_id, user["id"]):
        conn.close()
        return redirect_with_flash(
            request,
            "/pengajuan-dana-mitra",
            "Akses ditolak. Hanya data upload Anda yang dapat dihapus.",
        )
    row = conn.execute(
        "SELECT upload_id FROM tagihan WHERE id = ? AND kategori = 'pengajuan_dana_mitra'", (item_id,)
    ).fetchone()
    conn.execute("DELETE FROM tagihan WHERE id = ? AND kategori = 'pengajuan_dana_mitra'", (item_id,))
    if row and row[0]:
        recalc_pengajuan_dana_mitra_laporan(conn, row[0])
    conn.commit()
    conn.close()
    return RedirectResponse("/pengajuan-dana-mitra", status_code=303)


@router.post("/api/pengajuan-dana-mitra/bulk-delete")
async def api_pengajuan_dana_mitra_bulk_delete(request: Request, user=Depends(require_login)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        if not ids:
            return JSONResponse({"error": "no valid ids"}, status_code=400)
        conn = get_db()
        if not is_admin(user):
            ids = [i for i in ids if _member_owns_pdm_item(conn, i, user["id"])]
            if not ids:
                conn.close()
                return JSONResponse(
                    {"error": "Akses ditolak. Hanya data upload Anda yang dapat dihapus."},
                    status_code=403,
                )
        placeholders = ",".join("?" * len(ids))
        upload_rows = conn.execute(
            f"""SELECT DISTINCT upload_id FROM tagihan
                WHERE id IN ({placeholders}) AND kategori = 'pengajuan_dana_mitra' AND upload_id IS NOT NULL""",
            ids,
        ).fetchall()
        conn.execute(
            f"DELETE FROM tagihan WHERE id IN ({placeholders}) AND kategori = 'pengajuan_dana_mitra'",
            ids,
        )
        deleted = conn.total_changes
        for row in upload_rows:
            recalc_pengajuan_dana_mitra_laporan(conn, row[0])
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "deleted": deleted})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@router.post("/api/pengajuan-dana-mitra/bulk-update")
async def api_pengajuan_dana_mitra_bulk_update(request: Request, user=Depends(require_login)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        status = data.get("status")
        if not ids or status is None:
            return JSONResponse({"error": "no valid ids or status"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        from app.services.approval import bulk_set_tagihan_status

        updated = bulk_set_tagihan_status(conn, ids, status, user, kategori="pengajuan_dana_mitra")
        conn.commit()
        conn.close()
        if updated == 0:
            return JSONResponse(
                {"error": "Tidak ada status yang diubah. DIAJUKAN harus disetujui KA SPPG sebelum DIBAYARKAN."},
                status_code=400,
            )
        return JSONResponse({"success": True, "updated": updated})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@router.get("/export/pengajuan-dana-mitra/csv")
async def export_pengajuan_dana_mitra_csv(user=Depends(require_login)):
    import csv
    from io import StringIO

    data = get_pengajuan_dana_mitra()
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["NO", "TANGGAL", "KETERANGAN", "NAMA REKENING", "NOMOR REKENING", "BANK", "JUMLAH", "STATUS"])
    for d in data:
        writer.writerow([
            d["no"] or "",
            d["tanggal"] or "",
            d["pengajuan"],
            d["atas_nama"] or "",
            d["nomor_rekening"] or "",
            d["bank"] or "",
            d["jumlah"],
            d["status"] or "",
        ])
    output.seek(0)
    filename = f"pengajuan_dana_mitra_sppg_{date.today().isoformat()}.csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/pengajuan-dana-mitra/xlsx")
async def export_pengajuan_dana_mitra_xlsx(user=Depends(require_login)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    data = get_pengajuan_dana_mitra()
    wb = Workbook()
    ws = wb.active
    ws.title = "Pengajuan Dana Mitra"

    header_fill = PatternFill(start_color="071E49", end_color="071E49", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws.merge_cells("B1:I1")
    ws["B1"] = "PENGAJUAN DANA MITRA - SPPG WISMA HAJI MADIUN"
    ws["B1"].font = Font(bold=True, size=14, color="071E49")
    ws["B1"].alignment = Alignment(horizontal="center")

    headers = ["NO", "TANGGAL", "KETERANGAN", "NAMA REKENING", "NOMOR REKENING", "BANK", "JUMLAH", "STATUS"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for idx, d in enumerate(data, start=4):
        row_data = [
            d["no"] or "",
            d["tanggal"] or "",
            d["pengajuan"],
            d["atas_nama"] or "",
            d["nomor_rekening"] or "",
            d["bank"] or "",
            d["jumlah"],
            d["status"] or "",
        ]
        for col, val in enumerate(row_data, start=2):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border
            if col == 8:
                cell.number_format = "#,##0"
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [7, 14, 35, 22, 18, 14, 15, 14]
    for i, w in enumerate(widths, start=2):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Pengajuan_Dana_Mitra_SPPG_{date.today().isoformat()}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
