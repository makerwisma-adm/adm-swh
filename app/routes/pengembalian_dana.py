"""Pengembalian dana routes."""
import csv
from datetime import date
from io import StringIO

from fastapi import APIRouter, Depends, Form, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response

from app.auth.session import is_admin, redirect_with_flash, render_template, require_admin, require_login
from app.constants import (
    PENGEMBALIAN_DANA_HARI,
    PENGEMBALIAN_DANA_JUMLAH,
    PENGEMBALIAN_DANA_PER_HARI,
)
from app.db import get_db
from app.services.transfer_reports import (
    get_pengembalian_dana,
    get_pengembalian_dana_laporans,
    recalc_pengembalian_dana_laporan,
)
from app.services.report_sync import build_keuangan_link_context, filter_items_by_date_range
from app.utils.formatters import format_rupiah, format_tanggal_display

router = APIRouter()


@router.get("/pengembalian-dana", response_class=HTMLResponse)
async def pengembalian_dana_page(
    request: Request,
    user=Depends(require_login),
    upload_id: int = 0,
    search: str = "",
    status: str = "",
    rekening: str = "",
    tanggal: str = "",
    dari: str = "",
    sampai: str = "",
    message: str = "",
    success: bool = False,
):
    periode_options = get_pengembalian_dana_laporans(active_only=True)
    laporans = periode_options
    laporan = None
    active_upload_ids = {lap["upload_id"] for lap in periode_options}

    filters = {}
    if search:
        filters["search"] = search
    if status:
        filters["status"] = status
    if rekening:
        filters["rekening"] = rekening
    if tanggal:
        filters["tanggal"] = tanggal

    if upload_id and upload_id not in active_upload_ids:
        upload_id = 0

    view_all = upload_id == 0

    if upload_id and upload_id in active_upload_ids:
        laporan = next((lap for lap in periode_options if lap["upload_id"] == upload_id), None)
        filters["upload_id"] = upload_id

    data = get_pengembalian_dana(filters)
    data = filter_items_by_date_range(data, dari, sampai)

    periode_map = {lap["upload_id"]: lap.get("periode") or lap.get("filename") for lap in laporans}
    for item in data:
        uid = item.get("upload_id")
        item["periode_label"] = periode_map.get(uid, "Input Manual") if uid else "Input Manual"

    if not data:
        main_status = "—"
    else:
        all_terbayar = all((item.get("status") or "").upper() in ("DIBAYARKAN", "TERBAYAR") for item in data)
        main_status = "Sukses" if all_terbayar else "Belum Lunas"

    conn = get_db()
    pd_filter = "kategori = 'pengembalian_dana'"
    statuses = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT status FROM tagihan WHERE status IS NOT NULL AND {pd_filter}"
        ).fetchall()
    ]
    rekenings = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT rekening FROM tagihan WHERE rekening IS NOT NULL AND {pd_filter}"
        ).fetchall()
    ]
    conn.close()

    total_filtered = sum(d["jumlah"] for d in data)
    avg_gaji = int(total_filtered / len(data)) if data else 0
    max_item = max(data, key=lambda x: x["jumlah"]) if data else None
    min_item = min(data, key=lambda x: x["jumlah"]) if data else None
    total_all_batches = sum(lap.get("total_gaji") or 0 for lap in laporans)
    total_all_mitra = sum(lap.get("jumlah_penerima") or lap.get("record_count") or 0 for lap in laporans)

    return render_template(request, "pengembalian_dana.html", {
        "user": user,
        "items": data,
        "laporans": laporans,
        "laporan": laporan,
        "view_all": view_all,
        "selected_upload_id": upload_id if upload_id else 0,
        "total_filtered": total_filtered,
        "pengembalian_dana_jumlah": PENGEMBALIAN_DANA_JUMLAH,
        "pengembalian_dana_per_hari": PENGEMBALIAN_DANA_PER_HARI,
        "pengembalian_dana_hari": PENGEMBALIAN_DANA_HARI,
        "total_all_batches": total_all_batches,
        "total_all_mitra": total_all_mitra,
        "avg_gaji": avg_gaji,
        "max_item": max_item,
        "min_item": min_item,
        "main_status": main_status,
        "format_rupiah": format_rupiah,
        "format_tanggal": format_tanggal_display,
        "filters": {
            "search": search,
            "status": status,
            "rekening": rekening,
            "tanggal": tanggal,
            "periode": upload_id,
            "dari": (dari or "").strip(),
            "sampai": (sampai or "").strip(),
        },
        **build_keuangan_link_context(dari, sampai, module_key="pengembalian_dana"),
        "periode_options": periode_options,
        "status_options": sorted([s for s in statuses if s]),
        "rekening_options": sorted([r for r in rekenings if r]),
        "message": message,
        "success": success,
    })


@router.post("/pengembalian-dana")
async def create_pengembalian_dana(
    request: Request,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form("DIAJUKAN"),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        INSERT INTO tagihan (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, kategori, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pengembalian_dana', ?)
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or "DIAJUKAN",
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        user["id"],
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/pengembalian-dana", status_code=303)


@router.post("/pengembalian-dana/{item_id}/update")
async def update_pengembalian_dana(
    item_id: int,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form(""),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        UPDATE tagihan SET
            no = ?, pengajuan = ?, jumlah = ?, status = ?, rekening = ?,
            tanggal = ?, atas_nama = ?, nomor_rekening = ?, bank = ?,
            kategori = 'pengembalian_dana', updated_at = CURRENT_TIMESTAMP
        WHERE id = ? AND kategori = 'pengembalian_dana'
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or None,
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        item_id,
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/pengembalian-dana", status_code=303)


@router.post("/pengembalian-dana/{item_id}/delete")
async def delete_pengembalian_dana(item_id: int, user=Depends(require_login)):
    conn = get_db()
    row = conn.execute(
        "SELECT upload_id FROM tagihan WHERE id = ? AND kategori = 'pengembalian_dana'", (item_id,)
    ).fetchone()
    conn.execute("DELETE FROM tagihan WHERE id = ? AND kategori = 'pengembalian_dana'", (item_id,))
    if row and row[0]:
        recalc_pengembalian_dana_laporan(conn, row[0])
    conn.commit()
    conn.close()
    return RedirectResponse("/pengembalian-dana", status_code=303)


@router.post("/api/pengembalian-dana/bulk-delete")
async def api_pengembalian_dana_bulk_delete(request: Request, user=Depends(require_admin)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        if not ids:
            return JSONResponse({"error": "no valid ids"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        upload_rows = conn.execute(
            f"""SELECT DISTINCT upload_id FROM tagihan
                WHERE id IN ({placeholders}) AND kategori = 'pengembalian_dana' AND upload_id IS NOT NULL""",
            ids,
        ).fetchall()
        conn.execute(
            f"DELETE FROM tagihan WHERE id IN ({placeholders}) AND kategori = 'pengembalian_dana'",
            ids,
        )
        deleted = conn.total_changes
        for row in upload_rows:
            recalc_pengembalian_dana_laporan(conn, row[0])
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "deleted": deleted})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@router.post("/api/pengembalian-dana/bulk-update")
async def api_pengembalian_dana_bulk_update(request: Request, user=Depends(require_admin)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        status = data.get("status")
        if not ids or status is None:
            return JSONResponse({"error": "no valid ids or status"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        from app.services.approval import bulk_set_tagihan_status

        updated = bulk_set_tagihan_status(conn, ids, status, user, kategori="pengembalian_dana")
        conn.commit()
        conn.close()
        if updated == 0:
            return JSONResponse(
                {"error": "Tidak ada status yang diubah. DIAJUKAN harus disetujui KA SPPG sebelum DIBAYARKAN."},
                status_code=400,
            )
        return JSONResponse({"success": True, "updated": updated})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@router.get("/export/pengembalian-dana/csv")
async def export_pengembalian_dana_csv(user=Depends(require_login)):
    data = get_pengembalian_dana()
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["NO", "TANGGAL", "KETERANGAN", "NAMA REKENING", "NOMOR REKENING", "BANK", "JUMLAH", "STATUS"])
    for d in data:
        writer.writerow([
            d["no"] or "",
            d["tanggal"] or "",
            d["pengajuan"],
            d["atas_nama"] or "",
            d["nomor_rekening"] or "",
            d["bank"] or "",
            d["jumlah"],
            d["status"] or "",
        ])
    output.seek(0)
    filename = f"pengembalian_dana_sppg_{date.today().isoformat()}.csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/pengembalian-dana/xlsx")
async def export_pengembalian_dana_xlsx(user=Depends(require_login)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    data = get_pengembalian_dana()
    wb = Workbook()
    ws = wb.active
    ws.title = "Pengembalian Dana"

    header_fill = PatternFill(start_color="071E49", end_color="071E49", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws.merge_cells("B1:I1")
    ws["B1"] = "PENGEMBALIAN DANA / REFUND - SPPG WISMA HAJI MADIUN"
    ws["B1"].font = Font(bold=True, size=14, color="071E49")
    ws["B1"].alignment = Alignment(horizontal="center")

    headers = ["NO", "TANGGAL", "KETERANGAN", "NAMA REKENING", "NOMOR REKENING", "BANK", "JUMLAH", "STATUS"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for idx, d in enumerate(data, start=4):
        row_data = [
            d["no"] or "",
            d["tanggal"] or "",
            d["pengajuan"],
            d["atas_nama"] or "",
            d["nomor_rekening"] or "",
            d["bank"] or "",
            d["jumlah"],
            d["status"] or "",
        ]
        for col, val in enumerate(row_data, start=2):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border
            if col == 8:
                cell.number_format = "#,##0"
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [7, 14, 35, 22, 18, 14, 15, 14]
    for i, w in enumerate(widths, start=2):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Pengembalian_Dana_SPPG_{date.today().isoformat()}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
