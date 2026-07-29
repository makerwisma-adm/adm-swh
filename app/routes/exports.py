"""Tagihan export routes."""
import csv
from datetime import date
from io import StringIO

from fastapi import APIRouter, Depends
from fastapi.responses import Response

from app.auth.session import require_login
from app.exports.tagihan import _build_tagihan_pdf
from app.services.tagihan import enrich_tagihan_item, get_all_tagihan

router = APIRouter()


@router.get("/export/pdf")
async def export_pdf(user=Depends(require_login)):
    data = get_all_tagihan()
    content = _build_tagihan_pdf(data)
    filename = f"Laporan_Tagihan_SPPG_{date.today().isoformat()}.pdf"
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/csv")
async def export_csv(user=Depends(require_login)):
    data = get_all_tagihan()
    output = StringIO()
    writer = csv.writer(output)

    writer.writerow(["NO", "KETERANGAN", "JUMLAH", "CHARGES", "TOTAL", "STATUS", "BANK", "NOMOR REKENING", "ATAS NAMA REK.", "TANGGAL"])

    for raw in data:
        d = enrich_tagihan_item(raw)
        writer.writerow([
            d["no"] or "",
            d["pengajuan"],
            d["jumlah"],
            d["charges"],
            d["total"],
            d["status"] or "",
            d["bank"] or "",
            d["nomor_rekening"] or "",
            d["atas_nama"] or "",
            d["tanggal"] or "",
        ])

    output.seek(0)
    filename = f"tagihan_sppg_{date.today().isoformat()}.csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/xlsx")
async def export_xlsx(user=Depends(require_login)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    data = get_all_tagihan()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tagihan Mitra"

    # Header styling similar to original
    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Title row
    ws.merge_cells('B1:J1')
    ws['B1'] = "TAGIHAN MITRA - SPPG WISMA HAJI MADIUN"
    ws['B1'].font = Font(bold=True, size=14, color="0F766E")
    ws['B1'].alignment = Alignment(horizontal='center')

    # Headers (row 3 to match original structure)
    headers = ["NO", "KETERANGAN", "JUMLAH", "CHARGES", "TOTAL", "STATUS", "BANK", "NOMOR REKENING", "ATAS NAMA REK.", "TANGGAL"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data rows
    for idx, raw in enumerate(data, start=4):
        d = enrich_tagihan_item(raw)
        row_data = [
            d["no"] or "",
            d["pengajuan"],
            d["jumlah"],
            d["charges"],
            d["total"],
            d["status"] or "",
            d["bank"] or "",
            d["nomor_rekening"] or "",
            d["atas_nama"] or "",
            d["tanggal"] or "",
        ]
        for col, val in enumerate(row_data, start=2):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border
            if col in (4, 5, 6):  # JUMLAH, CHARGES, TOTAL
                cell.number_format = '#,##0'
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # Column widths (close to original)
    widths = [7, 40, 15, 12, 15, 16, 14, 18, 24, 14]
    for i, w in enumerate(widths, start=2):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = 'A4'

    # Total row
    total_row = 4 + len(data)
    ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=3)
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal='right', vertical='center')
    for col in (4, 5, 6):
        total_cell = ws.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_row-1})")
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Tagihan_Mitra_SPPG_{date.today().isoformat()}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
