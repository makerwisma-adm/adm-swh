#!/usr/bin/env python3
"""
Sistem Pelaporan Keuangan SPPG Wisma Haji Madiun
Web app untuk pelaporan tagihan mitra (Maker SPPG)
"""

import sqlite3
import os
from datetime import datetime, date, timedelta
from decimal import Decimal
from typing import Optional, List, Dict, Any

from fastapi import FastAPI, Request, Form, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, Response
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBasic
from starlette.middleware.sessions import SessionMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from passlib.hash import pbkdf2_sha256
from itsdangerous import URLSafeTimedSerializer, BadSignature
from dateutil import parser as date_parser

from db import get_db, db_info

# Config
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
IS_VERCEL = os.getenv("VERCEL") == "1"
PUBLIC_APP_URL = os.getenv(
    "PUBLIC_APP_URL",
    "https://adm-swh.vercel.app" if IS_VERCEL else "http://localhost:8001",
).rstrip("/")
PUBLIC_LOGIN_URL = f"{PUBLIC_APP_URL}/masuk"
SECRET_KEY = os.getenv("SECRET_KEY", "sppg-wisma-haji-madiun-2026-super-secret-key-change-in-prod")
SESSION_COOKIE_NAME = "sppg_session"

# Auth (using pbkdf2 for better compatibility)
serializer = URLSafeTimedSerializer(SECRET_KEY)

# FastAPI
app = FastAPI(title="Pelaporan Keuangan SPPG Wisma Haji", version="1.0")

templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))
# Vercel + Jinja2 3.x: nonaktifkan cache template (hindari TypeError unhashable dict)
templates.env.cache_size = 0
templates.env.cache = None
templates.env.auto_reload = True

# Static & uploads: di Vercel disajikan dari public/ oleh CDN
if not IS_VERCEL:
    if os.path.exists(os.path.join(BASE_DIR, "static")):
        app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
    _local_uploads = os.path.join(BASE_DIR, "uploads")
    if os.path.exists(_local_uploads):
        app.mount("/uploads", StaticFiles(directory=_local_uploads), name="uploads")

UPLOAD_DIR = "/tmp/sppg-uploads" if IS_VERCEL else os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(os.path.join(UPLOAD_DIR, "nota"), exist_ok=True)
os.makedirs(os.path.join(UPLOAD_DIR, "lampiran"), exist_ok=True)


@app.exception_handler(HTTPException)
async def auth_redirect_handler(request: Request, exc: HTTPException):
    """Redirect ke login saat belum auth — hindari respons JSON {"detail":"Found"}."""
    if exc.status_code in (301, 302, 303, 307, 308):
        location = (exc.headers or {}).get("Location")
        if location:
            return RedirectResponse(url=location, status_code=exc.status_code)
    return JSONResponse({"detail": exc.detail}, status_code=exc.status_code)


# ===================== DATABASE =====================

def init_db():
    conn = get_db()
    c = conn.cursor()

    # Users
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            full_name TEXT NOT NULL,
            role TEXT DEFAULT 'user'
        )
    """)

    # Tagihan Mitra (core table matching Excel)
    c.execute("""
        CREATE TABLE IF NOT EXISTS tagihan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            no TEXT,
            pengajuan TEXT NOT NULL,
            jumlah INTEGER NOT NULL,
            status TEXT,
            rekening TEXT,
            tanggal TEXT,
            atas_nama TEXT,
            nomor_rekening TEXT,
            bank TEXT,
            pos TEXT,
            kategori TEXT DEFAULT 'tagihan',
            upload_id INTEGER,
            created_by INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS uploads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            pos TEXT,
            periode TEXT,
            record_count INTEGER DEFAULT 0,
            created_by INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Add pos column if upgrading old database
    try:
        c.execute("ALTER TABLE tagihan ADD COLUMN pos TEXT")
    except:
        pass
    try:
        c.execute("ALTER TABLE tagihan ADD COLUMN upload_id INTEGER")
    except:
        pass
    try:
        c.execute("ALTER TABLE tagihan ADD COLUMN kategori TEXT")
    except:
        pass
    try:
        c.execute("ALTER TABLE tagihan ADD COLUMN nota_path TEXT")
    except:
        pass
    for col in ("pict_path", "bukti_path"):
        try:
            c.execute(f"ALTER TABLE tagihan ADD COLUMN {col} TEXT")
        except:
            pass
    for col, typedef in [
        ("debit", "INTEGER DEFAULT 0"),
        ("kredit", "INTEGER DEFAULT 0"),
        ("saldo_akhir", "INTEGER"),
        ("tipe_transaksi", "TEXT"),
    ]:
        try:
            c.execute(f"ALTER TABLE tagihan ADD COLUMN {col} {typedef}")
        except:
            pass

    c.execute("""
        CREATE TABLE IF NOT EXISTS petty_cash_laporan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            upload_id INTEGER UNIQUE,
            nama_karyawan TEXT,
            divisi TEXT,
            saldo_awal INTEGER DEFAULT 0,
            sisa_dana INTEGER DEFAULT 0,
            total_digantikan INTEGER DEFAULT 0,
            payment_info TEXT,
            bank TEXT,
            nomor_rekening TEXT,
            atas_nama TEXT,
            periode TEXT,
            filename TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    for col, typedef in [
        ("report_type", "TEXT DEFAULT 'reimbursement'"),
        ("total_debit", "INTEGER DEFAULT 0"),
        ("total_kredit", "INTEGER DEFAULT 0"),
        ("saldo_akhir", "INTEGER DEFAULT 0"),
        ("yang_menyetujui", "TEXT"),
        ("tanggal_ttd_pemohon", "TEXT"),
        ("tanggal_ttd_menyetujui", "TEXT"),
    ]:
        try:
            c.execute(f"ALTER TABLE petty_cash_laporan ADD COLUMN {col} {typedef}")
        except:
            pass

    c.execute("""
        CREATE TABLE IF NOT EXISTS gaji_relawan_laporan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            upload_id INTEGER UNIQUE,
            tanggal_pembayaran TEXT,
            rekening_sumber TEXT,
            jumlah_penerima INTEGER DEFAULT 0,
            total_gaji INTEGER DEFAULT 0,
            periode TEXT,
            bank TEXT DEFAULT 'MANDIRI',
            kota TEXT,
            filename TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS insentif_pic_laporan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            upload_id INTEGER UNIQUE,
            tanggal_pembayaran TEXT,
            rekening_sumber TEXT,
            jumlah_penerima INTEGER DEFAULT 0,
            total_gaji INTEGER DEFAULT 0,
            periode TEXT,
            bank TEXT DEFAULT 'MANDIRI',
            kota TEXT,
            filename TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS insentif_mitra_laporan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            upload_id INTEGER UNIQUE,
            tanggal_pembayaran TEXT,
            rekening_sumber TEXT,
            jumlah_penerima INTEGER DEFAULT 0,
            total_gaji INTEGER DEFAULT 0,
            periode TEXT,
            bank TEXT DEFAULT 'MANDIRI',
            kota TEXT,
            filename TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS pengembalian_dana_laporan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            upload_id INTEGER UNIQUE,
            tanggal_pembayaran TEXT,
            rekening_sumber TEXT,
            jumlah_penerima INTEGER DEFAULT 0,
            total_gaji INTEGER DEFAULT 0,
            periode TEXT,
            bank TEXT DEFAULT 'MANDIRI',
            kota TEXT,
            filename TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS pengajuan_dana_mitra_laporan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            upload_id INTEGER UNIQUE,
            tanggal_pembayaran TEXT,
            rekening_sumber TEXT,
            jumlah_penerima INTEGER DEFAULT 0,
            total_gaji INTEGER DEFAULT 0,
            periode TEXT,
            bank TEXT DEFAULT 'MANDIRI',
            kota TEXT,
            filename TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS petty_cash_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            upload_id INTEGER NOT NULL,
            no TEXT,
            pengajuan TEXT NOT NULL,
            jumlah INTEGER DEFAULT 0,
            debit INTEGER DEFAULT 0,
            kredit INTEGER DEFAULT 0,
            saldo_akhir INTEGER,
            tipe_transaksi TEXT,
            tanggal TEXT,
            nota_path TEXT,
            ket TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    try:
        c.execute("ALTER TABLE petty_cash_items ADD COLUMN ket TEXT")
    except Exception:
        pass

    # Seed users if empty
    c.execute("SELECT COUNT(*) FROM users")
    if c.fetchone()[0] == 0:
        default_users = [
            ("swhm", "A0312", "Maker SWH", "admin"),
            ("admin", "sppg123", "Admin Keuangan", "admin"),  # backup
        ]
        for u, p, name, role in default_users:
            ph = pbkdf2_sha256.hash(p)
            c.execute(
                "INSERT INTO users (username, password_hash, full_name, role) VALUES (?,?,?,?)",
                (u, ph, name, role)
            )

    c.execute(
        "UPDATE users SET password_hash = ?, full_name = ? WHERE username = 'swhm'",
        (pbkdf2_sha256.hash("A0312"), "Adam Primaskoro"),
    )

    icha_ph = pbkdf2_sha256.hash("sppg123")
    if c.execute("SELECT id FROM users WHERE username = 'icha'").fetchone():
        c.execute(
            "UPDATE users SET password_hash = ?, full_name = ?, role = ? WHERE username = 'icha'",
            (icha_ph, "Icha Salsabila", "member"),
        )
    elif c.execute("SELECT id FROM users WHERE username = 'member'").fetchone():
        c.execute(
            "UPDATE users SET username = 'icha', password_hash = ?, full_name = ?, role = ? WHERE username = 'member'",
            (icha_ph, "Icha Salsabila", "member"),
        )
    else:
        c.execute(
            "INSERT INTO users (username, password_hash, full_name, role) VALUES (?,?,?,?)",
            ("icha", icha_ph, "Icha Salsabila", "member"),
        )

    ulil_ph = pbkdf2_sha256.hash("swh123")
    if c.execute("SELECT id FROM users WHERE username = 'ulil'").fetchone():
        c.execute(
            "UPDATE users SET password_hash = ?, full_name = ?, role = ? WHERE username = 'ulil'",
            (ulil_ph, "Moch. Ulil Amri", "viewer"),
        )
    else:
        c.execute(
            "INSERT INTO users (username, password_hash, full_name, role) VALUES (?,?,?,?)",
            ("ulil", ulil_ph, "Moch. Ulil Amri", "viewer"),
        )

    wisma_ph = pbkdf2_sha256.hash("a123")
    if c.execute("SELECT id FROM users WHERE username = 'wisma'").fetchone():
        c.execute(
            "UPDATE users SET password_hash = ?, full_name = ?, role = ? WHERE username = 'wisma'",
            (wisma_ph, "Bapak Herman", "viewer"),
        )
    else:
        c.execute(
            "INSERT INTO users (username, password_hash, full_name, role) VALUES (?,?,?,?)",
            ("wisma", wisma_ph, "Bapak Herman", "viewer"),
        )

    conn.commit()
    conn.close()

def seed_initial_tagihan():
    """No longer auto-seeds demo data. User starts with empty data."""
    # Data is now empty by default. User will add/upload fresh data with proper categories.
    pass

def migrate_petty_cash_from_tagihan():
    """Pindahkan data petty cash lama dari tabel tagihan ke tabel terpisah."""
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM petty_cash_items")
    if c.fetchone()[0] > 0:
        conn.close()
        return
    rows = c.execute(
        "SELECT * FROM tagihan WHERE kategori = 'petty_cash' ORDER BY id"
    ).fetchall()
    for r in rows:
        c.execute("""
            INSERT INTO petty_cash_items
            (upload_id, no, pengajuan, jumlah, debit, kredit, saldo_akhir, tipe_transaksi, tanggal, nota_path, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            r["upload_id"] or 0,
            r["no"],
            r["pengajuan"],
            r["jumlah"],
            r["debit"] if "debit" in r.keys() else 0,
            r["kredit"] if "kredit" in r.keys() else 0,
            r["saldo_akhir"] if "saldo_akhir" in r.keys() else None,
            r["tipe_transaksi"] if "tipe_transaksi" in r.keys() else None,
            r["tanggal"],
            r["nota_path"] if "nota_path" in r.keys() else None,
            r["created_at"],
        ))
    if rows:
        c.execute("DELETE FROM tagihan WHERE kategori = 'petty_cash'")
    conn.commit()
    conn.close()

init_db()
migrate_petty_cash_from_tagihan()
seed_initial_tagihan()

# ===================== HELPERS =====================

def format_rupiah(amount: int) -> str:
    if amount is None:
        return "Rp0"
    if amount < 0:
        return "-Rp" + f"{abs(amount):,.0f}".replace(",", ".")
    return "Rp" + f"{amount:,.0f}".replace(",", ".")

ID_MONTH_NAMES = (
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
)


def format_tanggal_display(tgl_str: str) -> str:
    """Format tanggal: Tanggal - Bulan(huruf) - Tahun(angka), contoh 11 - Juni - 2026."""
    if not tgl_str:
        return "—"
    raw = tgl_str.strip()
    iso = raw if len(raw) == 10 and raw[4] == "-" and raw[7] == "-" else _parse_id_date(raw)
    if not iso:
        return tgl_str
    try:
        dt = datetime.strptime(iso, "%Y-%m-%d")
        bulan = ID_MONTH_NAMES[dt.month] if 1 <= dt.month <= 12 else str(dt.month)
        return f"{dt.day} - {bulan} - {dt.year}"
    except Exception:
        return tgl_str


def format_tanggal_pengajuan(tgl_str: str) -> str:
    """Format seperti formulir PDF: 20 Mei 2026."""
    if not tgl_str:
        return "—"
    raw = tgl_str.strip()
    iso = raw if len(raw) == 10 and raw[4] == "-" and raw[7] == "-" else _parse_id_date(raw)
    if not iso:
        return tgl_str
    try:
        dt = datetime.strptime(iso, "%Y-%m-%d")
        bulan = ID_MONTH_NAMES[dt.month] if 1 <= dt.month <= 12 else str(dt.month)
        return f"{dt.day} {bulan} {dt.year}"
    except Exception:
        return tgl_str


def _parse_slash_date(tgl_str: str) -> Optional[str]:
    """Parse tanggal format M/D/YYYY dari tanda tangan PDF, contoh 6/10/2026."""
    if not tgl_str:
        return None
    raw = tgl_str.strip()
    for fmt in ("%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _parse_id_date(tgl_str: str) -> Optional[str]:
    if not tgl_str:
        return None
    import re
    norm = tgl_str.strip().lower()
    for indo, eng in [
        ("januari", "jan"), ("februari", "feb"), ("maret", "mar"), ("april", "apr"),
        ("mei", "may"), ("juni", "jun"), ("juli", "jul"), ("agustus", "aug"),
        ("september", "sep"), ("oktober", "oct"), ("november", "nov"), ("desember", "dec"),
    ]:
        norm = re.sub(r"\b" + indo + r"\b", eng, norm)
    try:
        dt = date_parser.parse(norm, dayfirst=True)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return None

def _parse_rp_amount(text: str) -> int:
    import re
    m = re.search(r"Rp\s*([\d.,]+)", text)
    if not m:
        return 0
    raw = m.group(1).replace(".", "").replace(",", "")
    try:
        return int(raw)
    except ValueError:
        return 0

def get_petty_cash_laporans() -> List[Dict[str, Any]]:
    conn = get_db()
    rows = conn.execute("""
        SELECT p.*, u.record_count, u.created_at as upload_at
        FROM petty_cash_laporan p
        LEFT JOIN uploads u ON u.id = p.upload_id
        ORDER BY p.id DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_petty_cash_items(upload_id: int) -> List[Dict[str, Any]]:
    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM petty_cash_items
        WHERE upload_id = ?
        ORDER BY id ASC
    """, (upload_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def filter_petty_cash_items(
    items: List[Dict[str, Any]],
    search: str = "",
    tanggal: str = "",
    jenis: str = "",
) -> List[Dict[str, Any]]:
    """Filter petty cash rows by search, date, and transaction type."""
    result = items
    jenis = (jenis or "").strip().lower()
    if jenis == "pemasukan":
        result = [i for i in result if (i.get("debit") or 0) > 0]
    elif jenis == "pengeluaran":
        result = [
            i for i in result
            if (i.get("kredit") or 0) > 0
            or ((i.get("debit") or 0) == 0 and (i.get("jumlah") or 0) > 0)
        ]
    if tanggal:
        result = [i for i in result if (i.get("tanggal") or "") == tanggal]
    if search:
        s = search.strip().lower()
        result = [
            i for i in result
            if s in (i.get("pengajuan") or "").lower()
            or s in (i.get("tipe_transaksi") or "").lower()
        ]
    return result


def sum_petty_pengeluaran(items: List[Dict[str, Any]]) -> int:
    """Total pengeluaran = jumlah semua baris kredit / biaya yang ditambahkan."""
    total = sum(int(i.get("kredit") or 0) for i in items)
    if total > 0:
        return total
    return sum(
        int(i.get("jumlah") or 0)
        for i in items
        if int(i.get("debit") or 0) == 0
    )


def get_current_user(request: Request):
    """Ambil user dari session middleware saja — hindari konflik cookie ganda."""
    try:
        data = request.session.get("user")
        if not data or not data.get("user_id"):
            return None
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE id = ?", (data["user_id"],)).fetchone()
        conn.close()
        return dict(user) if user else None
    except Exception:
        return None


PORTAL_MODULES = [
    {"href": "/dashboard", "icon": "fa-chart-line", "title": "Dashboard", "desc": "Ringkasan keuangan"},
    {"href": "/tagihan", "icon": "fa-file-invoice-dollar", "title": "Laporan Tagihan", "desc": "Data tagihan mitra"},
    {"href": "/petty-cash", "icon": "fa-wallet", "title": "Petty Cash", "desc": "Buku besar & reimbursement"},
    {"href": "/insentif-mitra", "icon": "fa-handshake", "title": "Laporan Insentif Mitra", "desc": "Pembayaran insentif mitra"},
    {"href": "/pengembalian-dana", "icon": "fa-rotate-left", "title": "Pengembalian Dana", "desc": "Pengembalian dana / refund"},
    {"href": "/pengajuan-dana-mitra", "icon": "fa-file-invoice", "title": "Pengajuan Dana Mitra", "desc": "Pengajuan dana mitra SPPG"},
    {"href": "/gaji-relawan", "icon": "fa-users", "title": "Gaji Relawan", "desc": "Pembayaran relawan"},
    {"href": "/insentif-pic", "icon": "fa-user-tie", "title": "Insentif PIC", "desc": "Pembayaran insentif PIC"},
]


ROLE_ADMIN = "admin"
ROLE_MEMBER = "member"
ROLE_VIEWER = "viewer"
AUTH_ONLY_PATHS = {"/masuk", "/login", "/logout"}
MEMBER_WRITE_PATHS = AUTH_ONLY_PATHS | {
    "/upload",
    "/pengajuan-dana-mitra/upload",
    "/petty-cash/upload",
    "/gaji-relawan/upload",
    "/insentif-pic/upload",
    "/insentif-mitra/upload",
}


def user_role(user: Optional[Dict]) -> str:
    if not user:
        return ""
    return (user.get("role") or ROLE_MEMBER).lower()


def is_admin(user: Optional[Dict]) -> bool:
    return user_role(user) == ROLE_ADMIN


def is_viewer(user: Optional[Dict]) -> bool:
    return user_role(user) == ROLE_VIEWER


def can_member_upload(user: Optional[Dict]) -> bool:
    """Member dengan hak upload (bukan viewer read-only)."""
    return user_role(user) == ROLE_MEMBER


def _member_owns_pdm_item(conn, item_id: int, user_id: int) -> bool:
    """Member hanya boleh hapus data pengajuan dana mitra yang ia upload."""
    row = conn.execute(
        """
        SELECT t.created_by, t.upload_id, u.created_by AS upload_creator
        FROM tagihan t
        LEFT JOIN uploads u ON u.id = t.upload_id
        WHERE t.id = ? AND t.kategori = 'pengajuan_dana_mitra'
        """,
        (item_id,),
    ).fetchone()
    if not row:
        return False
    if row["created_by"] == user_id:
        return True
    if row["upload_id"] and row["upload_creator"] == user_id:
        return True
    return False


def _is_member_pdm_delete_path(path: str) -> bool:
    import re
    if path == "/api/pengajuan-dana-mitra/bulk-delete":
        return True
    return bool(re.match(r"^/pengajuan-dana-mitra/\d+/delete$", path))


def _is_member_petty_cash_write_path(path: str) -> bool:
    """Member boleh upload PDF, nota, dan keterangan di halaman Petty Cash."""
    import re
    if path == "/petty-cash/upload":
        return True
    if re.match(r"^/api/petty-cash/\d+/nota$", path):
        return True
    if re.match(r"^/api/petty-cash/\d+/ket$", path):
        return True
    return False


TAGIHAN_ATTACHMENT_FIELDS = {"pict", "nota", "bukti"}
TAGIHAN_ATTACHMENT_KATEGORI = {"pengajuan_dana_mitra", "insentif_mitra"}


def _is_member_tagihan_attachment_path(path: str) -> bool:
    import re
    return bool(
        re.match(
            r"^/api/tagihan/\d+/(pict|nota|bukti)$",
            path,
        )
    )


def redirect_with_flash(
    request: Request,
    url: str,
    message: str = "",
    success: bool = False,
    status_code: int = 303,
) -> RedirectResponse:
    """Redirect dengan pesan sekali pakai (session flash), tanpa query ?message=."""
    if message:
        request.session["flash"] = {"message": message, "success": success}
    return RedirectResponse(url, status_code=status_code)


def render_template(request: Request, name: str, context: Dict = None, status_code: int = 200):
    """Starlette baru: TemplateResponse(request, name, context)."""
    ctx = {k: v for k, v in (context or {}).items() if k != "request"}
    flash = request.session.pop("flash", None)
    if flash:
        ctx["message"] = flash.get("message", "")
        ctx["success"] = bool(flash.get("success", False))
    elif ctx.get("message"):
        ctx["strip_message_params"] = True
    ctx.setdefault("public_app_url", PUBLIC_APP_URL)
    ctx.setdefault("public_login_url", PUBLIC_LOGIN_URL)
    ctx.setdefault("is_vercel", IS_VERCEL)
    user = ctx.get("user") or get_current_user(request)
    if user:
        ctx["user"] = user
        ctx["is_admin"] = is_admin(user)
        ctx["can_edit"] = is_admin(user)
        ctx["can_upload"] = is_admin(user) or can_member_upload(user)
        ctx["can_download"] = is_admin(user) or can_member_upload(user)
        ctx.setdefault("can_delete_upload", can_member_upload(user))
        ctx["user_role"] = user_role(user)
    else:
        ctx.setdefault("is_admin", False)
        ctx.setdefault("can_edit", False)
        ctx.setdefault("can_upload", False)
        ctx.setdefault("can_download", False)
        ctx.setdefault("can_delete_upload", False)
        ctx.setdefault("user_role", "")
    return templates.TemplateResponse(request, name, ctx, status_code=status_code)


def _login_context(**extra):
    return {
        "portal_modules": PORTAL_MODULES,
        "is_vercel": IS_VERCEL,
        "public_app_url": PUBLIC_APP_URL,
        "public_login_url": PUBLIC_LOGIN_URL,
        **extra,
    }


def _safe_next_url(next_url: str) -> str:
    """Hanya izinkan redirect internal."""
    if not next_url or not next_url.startswith("/") or next_url.startswith("//"):
        return "/dashboard"
    if next_url.startswith("/login") or next_url.startswith("/masuk") or next_url.startswith("/logout"):
        return "/dashboard"
    return next_url


def require_login(request: Request):
    user = get_current_user(request)
    if not user:
        next_path = request.url.path
        if request.url.query:
            next_path = f"{next_path}?{request.url.query}"
        from urllib.parse import quote
        location = f"/masuk?next={quote(next_path, safe='')}"
        raise HTTPException(status_code=302, headers={"Location": location})
    return user


def require_admin(request: Request):
    user = require_login(request)
    if not is_admin(user):
        raise HTTPException(status_code=403, detail="Akses ditolak. Hanya admin yang dapat mengubah data.")
    return user


class MemberWriteGuardMiddleware(BaseHTTPMiddleware):
    """Member: upload di Tagihan, Petty Cash, Gaji Relawan & Pengajuan Dana Mitra. Viewer: hanya lihat."""

    async def dispatch(self, request: Request, call_next):
        if request.method in ("POST", "PUT", "PATCH", "DELETE"):
            path = request.url.path.rstrip("/") or "/"
            user = get_current_user(request)
            if user and not is_admin(user):
                if is_viewer(user):
                    if path not in AUTH_ONLY_PATHS:
                        if path.startswith("/api/"):
                            return JSONResponse(
                                {"error": "Akses ditolak. Akun Anda hanya dapat melihat data."},
                                status_code=403,
                            )
                        return redirect_with_flash(
                            request,
                            "/dashboard",
                            "Akses ditolak. Akun Anda hanya dapat melihat data.",
                        )
                elif (
                    path not in MEMBER_WRITE_PATHS
                    and not _is_member_pdm_delete_path(path)
                    and not _is_member_petty_cash_write_path(path)
                    and not _is_member_tagihan_attachment_path(path)
                ):
                    if path.startswith("/api/"):
                        return JSONResponse(
                            {"error": "Akses ditolak. Member hanya dapat upload di halaman Tagihan, Petty Cash, Gaji Relawan, dan Pengajuan Dana Mitra."},
                            status_code=403,
                        )
                    return redirect_with_flash(
                        request,
                        "/tagihan",
                        "Akses ditolak. Anda tidak dapat mengubah data.",
                    )
        return await call_next(request)


app.add_middleware(MemberWriteGuardMiddleware)
app.add_middleware(
    SessionMiddleware,
    secret_key=SECRET_KEY,
    session_cookie=SESSION_COOKIE_NAME,
    max_age=60 * 60 * 24 * 7,
    same_site="lax",
    https_only=IS_VERCEL,
)

def get_all_tagihan(filters: Dict = None) -> List[Dict]:
    conn = get_db()
    query = "SELECT * FROM tagihan"
    params = []
    where = []
    if filters and filters.get("kategori"):
        where.append("kategori = ?")
        params.append(filters["kategori"])
    else:
        where.append("(kategori IS NULL OR kategori = '' OR kategori = 'tagihan')")

    if filters:
        if filters.get("search"):
            where.append("(pengajuan LIKE ? OR atas_nama LIKE ? OR bank LIKE ? OR pos LIKE ? OR no LIKE ? OR nomor_rekening LIKE ?)")
            s = f"%{filters['search']}%"
            params.extend([s, s, s, s, s, s])
        if filters.get("status"):
            where.append("status = ?")
            params.append(filters["status"])
        if filters.get("rekening"):
            where.append("rekening = ?")
            params.append(filters["rekening"])
        if filters.get("tanggal"):
            where.append("tanggal = ?")
            params.append(filters["tanggal"])
        if filters.get("upload_id"):
            where.append("upload_id = ?")
            params.append(filters["upload_id"])

    if where:
        query += " WHERE " + " AND ".join(where)

    query += " ORDER BY COALESCE(tanggal, '9999-12-31') DESC, id DESC"

    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_summary() -> Dict[str, Any]:
    conn = get_db()
    c = conn.cursor()
    exclude_pc = "WHERE (kategori IS NULL OR kategori = '' OR kategori = 'tagihan')"

    # Total keseluruhan (tanpa petty cash)
    c.execute(f"SELECT COALESCE(SUM(jumlah), 0) FROM tagihan {exclude_pc}")
    total = c.fetchone()[0] or 0

    c.execute(f"SELECT COALESCE(SUM(jumlah), 0) FROM tagihan {exclude_pc} AND status = 'TERBAYAR'")
    terbayar = c.fetchone()[0] or 0

    c.execute(f"SELECT COALESCE(SUM(jumlah), 0) FROM tagihan {exclude_pc} AND status = 'DIAJUKAN'")
    diajukan = c.fetchone()[0] or 0

    c.execute(f"SELECT COUNT(*) FROM tagihan {exclude_pc}")
    jumlah_item = c.fetchone()[0] or 0

    # By rekening
    c.execute(f"""
        SELECT COALESCE(rekening, 'Belum Ditentukan') as r, COALESCE(SUM(jumlah),0) 
        FROM tagihan {exclude_pc} GROUP BY r ORDER BY 2 DESC
    """)
    by_rekening = [{"rekening": r[0], "total": r[1]} for r in c.fetchall()]

    # By status counts
    c.execute(f"SELECT status, COUNT(*), COALESCE(SUM(jumlah),0) FROM tagihan {exclude_pc} GROUP BY status")
    by_status = [{"status": r[0] or "Belum Ada Status", "count": r[1], "total": r[2]} for r in c.fetchall()]

    conn.close()

    return {
        "total": total,
        "terbayar": terbayar,
        "diajukan": diajukan,
        "jumlah_item": jumlah_item,
        "by_rekening": by_rekening,
        "by_status": by_status,
    }

# ===================== ROUTES =====================

@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    user = get_current_user(request)
    if user:
        return RedirectResponse("/dashboard")
    return RedirectResponse("/masuk")


@app.get("/masuk", response_class=HTMLResponse)
async def login_page(request: Request, next: str = "", message: str = "", error: str = ""):
    if message:
        qs = f"?next={next}" if next else ""
        return RedirectResponse(f"/masuk{qs}", status_code=303)
    user = get_current_user(request)
    return render_template(request, "login.html", _login_context(
        user=user,
        error=error or None,
        message=None,
        next_url=_safe_next_url(next),
    ))


@app.get("/login", response_class=HTMLResponse)
async def login_redirect(request: Request, next: str = "", message: str = "", error: str = ""):
    from urllib.parse import urlencode
    params = {}
    if next:
        params["next"] = next
    if error:
        params["error"] = error
    qs = f"?{urlencode(params)}" if params else ""
    return RedirectResponse(f"/masuk{qs}", status_code=303)


@app.post("/masuk", response_class=HTMLResponse)
@app.post("/login", response_class=HTMLResponse)
async def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    next: str = Form("/dashboard"),
):
    username = username.strip()
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()

    if not user or not pbkdf2_sha256.verify(password, user["password_hash"]):
        return render_template(request, "login.html", _login_context(
            user=None,
            error="Username atau password salah. Periksa kembali lalu coba lagi.",
            message=None,
            next_url=_safe_next_url(next),
            username_value=username,
        ), status_code=401)

    request.session.clear()
    request.session["user"] = {
        "user_id": user["id"],
        "username": user["username"],
        "full_name": user["full_name"],
        "role": user["role"] or ROLE_MEMBER,
    }

    dest = _safe_next_url(next)
    return RedirectResponse(dest, status_code=303)


@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    resp = RedirectResponse("/masuk", status_code=303)
    resp.delete_cookie(SESSION_COOKIE_NAME)
    return resp

@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request, user=Depends(require_login)):
    summary = get_summary()
    recent = get_all_tagihan()[:6]

    # Monthly breakdown for chart (last 6 months)
    conn = get_db()
    monthly = conn.execute("""
        SELECT 
            strftime('%Y-%m', COALESCE(tanggal, created_at)) as bulan,
            COALESCE(SUM(jumlah), 0) as total
        FROM tagihan
        WHERE (kategori IS NULL OR kategori = '' OR kategori = 'tagihan')
        GROUP BY bulan
        ORDER BY bulan DESC
        LIMIT 6
    """).fetchall()
    conn.close()

    monthly_data = [{"bulan": m[0], "total": m[1]} for m in reversed(monthly)]

    return render_template(request, "dashboard.html", {
        "user": user,
        "summary": summary,
        "recent": recent,
        "monthly": monthly_data,
        "format_rupiah": format_rupiah,
        "today": date.today().isoformat(),
    })

@app.get("/tagihan", response_class=HTMLResponse)
async def tagihan_page(
    request: Request,
    user=Depends(require_login),
    search: str = "",
    status: str = "",
    rekening: str = "",
    tanggal: str = "",
    kategori: str = "",
    message: str = "",
    success: bool = False
):
    if kategori == "petty_cash":
        return RedirectResponse("/petty-cash", status_code=303)
    if kategori == "gaji_relawan":
        return RedirectResponse("/gaji-relawan", status_code=303)
    if kategori == "insentif_pic":
        return RedirectResponse("/insentif-pic", status_code=303)
    if kategori == "insentif_mitra":
        return RedirectResponse("/insentif-mitra", status_code=303)
    if kategori == "pengembalian_dana":
        return RedirectResponse("/pengembalian-dana", status_code=303)
    if kategori == "pengajuan_dana_mitra":
        return RedirectResponse("/pengajuan-dana-mitra", status_code=303)
    if kategori in ("pic", "tagihan_bulanan"):
        return RedirectResponse("/tagihan", status_code=303)

    filters = {}
    if search: filters["search"] = search
    if status: filters["status"] = status
    if rekening: filters["rekening"] = rekening
    if tanggal: filters["tanggal"] = tanggal
    if kategori: filters["kategori"] = kategori

    data = [enrich_tagihan_item(d) for d in get_all_tagihan(filters)]

    # Compute main status automatically based on current filtered data
    if not data:
        main_status = "—"
    else:
        all_terbayar = all((item.get("status") or "").upper() == "TERBAYAR" for item in data)
        main_status = "Sukses" if all_terbayar else "Tertagih"

    # Get filter options
    conn = get_db()
    pc_exclude = "AND (kategori IS NULL OR kategori = '' OR kategori = 'tagihan')"
    statuses = [r[0] for r in conn.execute(f"SELECT DISTINCT status FROM tagihan WHERE status IS NOT NULL {pc_exclude}").fetchall()]
    rekenings = [r[0] for r in conn.execute(f"SELECT DISTINCT rekening FROM tagihan WHERE rekening IS NOT NULL {pc_exclude}").fetchall()]
    kategori_list = []
    conn.close()

    total_filtered = sum(d["jumlah"] for d in data)
    total_charges = sum(d["charges"] for d in data)
    total_grand = total_filtered + total_charges

    return render_template(request, "tagihan.html", {
        "user": user,
        "tagihan": data,
        "total_filtered": total_filtered,
        "total_charges": total_charges,
        "total_grand": total_grand,
        "tagihan_charges_amount": TAGIHAN_CHARGES_AMOUNT,
        "main_status": main_status,
        "format_rupiah": format_rupiah,
        "format_tanggal": format_tanggal_display,
        "filters": {"search": search, "status": status, "rekening": rekening, "tanggal": tanggal, "kategori": kategori},
        "status_options": sorted([s for s in statuses if s]),
        "rekening_options": sorted([r for r in rekenings if r]),
        "kategori_options": sorted([k for k in kategori_list if k]),
        "message": message,
        "success": success,
    })

TAGIHAN_CHARGES_AMOUNT = 6500

FEE_PAYROL_PER_ORANG = 2500
INSENTIF_MITRA_PER_HARI = 6_000_000
INSENTIF_MITRA_HARI = 6
INSENTIF_MITRA_JUMLAH = INSENTIF_MITRA_PER_HARI * INSENTIF_MITRA_HARI
PENGEMBALIAN_DANA_PER_HARI = 6_000_000
PENGEMBALIAN_DANA_HARI = 6
PENGEMBALIAN_DANA_JUMLAH = PENGEMBALIAN_DANA_PER_HARI * PENGEMBALIAN_DANA_HARI


def _sort_key_no(item: Dict) -> int:
    try:
        raw = str(item.get("no") or "").strip()
        return int(raw) if raw else 999_999
    except (ValueError, TypeError):
        return 999_999


def get_gaji_relawan(filters: Dict = None) -> List[Dict]:
    f = dict(filters or {})
    f["kategori"] = "gaji_relawan"
    rows = get_all_tagihan(f)
    return sorted(rows, key=_sort_key_no)


def calc_fee_payrol(item_count: int) -> int:
    return FEE_PAYROL_PER_ORANG * max(item_count, 0)


def calc_tagihan_charges(item: Dict) -> int:
    """Biaya transfer Rp6.500 per baris, kecuali bank Mandiri."""
    bank = (item.get("bank") or "").strip().lower()
    if "mandiri" in bank:
        return 0
    return TAGIHAN_CHARGES_AMOUNT


def format_tagihan_rekening_export(item: Dict) -> str:
    """REKENING untuk export: gabung nomor rekening + atas nama (bukan label kategori)."""
    nomor = str(item.get("nomor_rekening") or "").strip()
    atas = str(item.get("atas_nama") or "").strip()
    bank = str(item.get("bank") or "").strip()
    if nomor and atas:
        prefix = f"{bank} " if bank else ""
        return f"{prefix}{nomor} a.n. {atas}".strip()
    if nomor:
        return f"{bank} {nomor}".strip() if bank else nomor
    if atas:
        return atas
    return str(item.get("rekening") or "").strip()


def enrich_tagihan_item(item: Dict) -> Dict:
    row = dict(item)
    jumlah = row.get("jumlah") or 0
    charges = calc_tagihan_charges(row)
    row["charges"] = charges
    row["total"] = jumlah + charges
    row["rekening_export"] = format_tagihan_rekening_export(row)
    return row


def get_gaji_relawan_laporans(active_only: bool = False) -> List[Dict[str, Any]]:
    """Daftar laporan gaji relawan. active_only=True: hanya upload yang masih punya data."""
    conn = get_db()
    query = """
        SELECT g.*, u.record_count, u.created_at AS upload_at,
               (SELECT COUNT(*) FROM tagihan t
                WHERE t.upload_id = g.upload_id AND t.kategori = 'gaji_relawan') AS item_count
        FROM gaji_relawan_laporan g
        LEFT JOIN uploads u ON u.id = g.upload_id
    """
    if active_only:
        query += """
        WHERE EXISTS (
            SELECT 1 FROM tagihan t
            WHERE t.upload_id = g.upload_id AND t.kategori = 'gaji_relawan'
        )
        """
    query += " ORDER BY g.id DESC"
    rows = conn.execute(query).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def recalc_gaji_relawan_laporan(conn, upload_id: int):
    """Hitung ulang total laporan setelah hapus baris — metadata upload tetap ada."""
    if not upload_id:
        return
    rows = conn.execute(
        "SELECT jumlah FROM tagihan WHERE upload_id = ? AND kategori = 'gaji_relawan'",
        (upload_id,),
    ).fetchall()
    count = len(rows)
    total = sum(r[0] or 0 for r in rows)
    if count == 0:
        conn.execute("DELETE FROM gaji_relawan_laporan WHERE upload_id = ?", (upload_id,))
    else:
        conn.execute("""
            UPDATE gaji_relawan_laporan SET jumlah_penerima = ?, total_gaji = ?
            WHERE upload_id = ?
        """, (count, total, upload_id))
    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (count, upload_id))


def _strip_nama_gelar(nama: str) -> str:
    """Hilangkan awalan gelar (IBU, BPK, BPA, SDRI, SDR, dll) dari nama relawan."""
    import re
    if not nama:
        return nama
    return re.sub(
        r"^(IBU|BPK|BPA|SDRI|SDR|TN|NY|NN)\.?\s+",
        "",
        nama.strip(),
        flags=re.I,
    ).strip()


def _parse_pic_periode(filename: str) -> Optional[str]:
    import re
    m = re.search(r"PERIODE\s*(\d+)", filename, re.I)
    if m:
        return f"Periode {m.group(1)}"
    m = re.search(r"(\d{1,2}\s+\w+\s+\d{4})", filename.replace("_", " "), re.I)
    if m:
        return m.group(1)
    return None


def _parse_mandiri_csv_date(raw: str) -> Optional[str]:
    raw = (raw or "").strip()
    if len(raw) == 8 and raw.isdigit():
        return f"{raw[0:4]}-{raw[4:6]}-{raw[6:8]}"
    return _parse_id_date(raw)


def _pic_cell_str(val) -> str:
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y%m%d")
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    if isinstance(val, int):
        return str(val)
    return str(val).strip()


def _pic_parse_amount(val) -> int:
    if val is None or val == "":
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(str(val).strip().replace(".", "").replace(",", "") or 0)
    except ValueError:
        return 0


def _parse_pic_transfer_rows(raw_rows: List, filename: str = "", default_label: str = "Gaji Relawan") -> Dict[str, Any]:
    """Parse baris format transfer massal Mandiri (header P + data per kolom)."""
    meta = {
        "tanggal_pembayaran": None,
        "rekening_sumber": None,
        "jumlah_penerima": 0,
        "total_gaji": 0,
        "periode": _parse_pic_periode(filename),
        "bank": "MANDIRI",
        "kota": "Madiun",
        "filename": filename,
    }
    items: List[Dict[str, Any]] = []

    rows = [[_pic_cell_str(c) for c in row] for row in raw_rows if row is not None]
    if not rows:
        return {"meta": meta, "items": items}

    header = rows[0]
    if header and (header[0] or "").strip().upper() == "P":
        meta["tanggal_pembayaran"] = _parse_mandiri_csv_date(header[1] if len(header) > 1 else "")
        meta["rekening_sumber"] = (header[2] if len(header) > 2 else "").strip() or None
        try:
            meta["jumlah_penerima"] = int((header[3] if len(header) > 3 else "0").strip() or 0)
        except ValueError:
            meta["jumlah_penerima"] = 0
        meta["total_gaji"] = _pic_parse_amount(header[4] if len(header) > 4 else 0)
        data_rows = rows[1:]
    else:
        data_rows = rows

    for row in data_rows:
        if not row or not any(cell for cell in row):
            continue
        nomor_rek = (row[0] if len(row) > 0 else "").strip()
        nama = _strip_nama_gelar((row[1] if len(row) > 1 else "").strip())
        if not nama and not nomor_rek:
            continue
        if not nama:
            continue

        jumlah = _pic_parse_amount(row[6] if len(row) > 6 else 0)
        if jumlah <= 0:
            continue

        bank = (row[11] if len(row) > 11 else "MANDIRI").strip() or "MANDIRI"
        kota = (row[12] if len(row) > 12 else "").strip() or meta.get("kota")
        if kota:
            meta["kota"] = kota

        periode_label = meta.get("periode") or default_label
        items.append({
            "no": str(len(items) + 1),
            "pengajuan": f"{nama} — {periode_label}",
            "atas_nama": nama,
            "nomor_rekening": nomor_rek or None,
            "bank": bank,
            "jumlah": jumlah,
            "tanggal": meta.get("tanggal_pembayaran"),
            "rekening": "TRANSFER MASSAL",
            "status": "DIAJUKAN",
        })

    if items and not meta["total_gaji"]:
        meta["total_gaji"] = sum(i["jumlah"] for i in items)
    if items and not meta["jumlah_penerima"]:
        meta["jumlah_penerima"] = len(items)

    return {"meta": meta, "items": items}


def parse_pic_transfer_csv(file_path: str, filename: str = "", default_label: str = "Gaji Relawan") -> Dict[str, Any]:
    """Parse CSV format transfer massal Mandiri (PIC), contoh CSV PIC PERIODE10."""
    import csv

    with open(file_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    return _parse_pic_transfer_rows(rows, filename, default_label)


def parse_pic_transfer_xlsx(file_path: str, filename: str = "", default_label: str = "Gaji Relawan") -> Dict[str, Any]:
    """Parse Excel (.xlsx/.xls) dengan struktur kolom yang sama seperti CSV transfer massal Mandiri."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return _parse_pic_transfer_rows(rows, filename, default_label)


PIC_TRANSFER_COL_COUNT = 43


def _nama_from_gaji_item(item: Dict) -> str:
    atas = (item.get("atas_nama") or "").strip()
    if atas:
        return atas
    pengajuan = item.get("pengajuan") or ""
    if " — " in pengajuan:
        return pengajuan.split(" — ", 1)[0].strip()
    return pengajuan.strip()


def _export_filename_from_laporan(laporan: Dict, ext: str) -> str:
    import re

    raw = (laporan.get("filename") or laporan.get("periode") or "gaji_relawan").strip()
    raw = re.sub(r'[<>:"/\\|?*]', "_", raw)
    base = raw.rsplit(".", 1)[0] if "." in raw else raw
    if ext == "csv" and base.lower().endswith((".xlsx", ".xls")):
        base = base.rsplit(".", 1)[0]
    return f"{base}.{ext}"


def resolve_gaji_relawan_export(upload_id: int = 0) -> tuple:
    """Ambil laporan + baris untuk export sesuai batch upload."""
    laporans = get_gaji_relawan_laporans(active_only=True)
    laporan = None

    if upload_id:
        laporan = next((lap for lap in laporans if lap["upload_id"] == upload_id), None)
        if not laporan:
            conn = get_db()
            row = conn.execute(
                """
                SELECT g.*, u.filename AS upload_filename
                FROM gaji_relawan_laporan g
                LEFT JOIN uploads u ON u.id = g.upload_id
                WHERE g.upload_id = ?
                """,
                (upload_id,),
            ).fetchone()
            conn.close()
            if row:
                laporan = dict(row)
    elif laporans:
        laporan = laporans[0]
        upload_id = laporan["upload_id"]

    if not laporan or not upload_id:
        return None, [], 0

    items = get_gaji_relawan({"upload_id": upload_id})
    return laporan, items, upload_id


def build_pic_transfer_export_rows(laporan: Dict, items: List[Dict]) -> List[List]:
    """Bangun baris export format CSV transfer massal Mandiri (sama seperti file upload)."""
    tgl_raw = (laporan.get("tanggal_pembayaran") or "").strip()
    tgl_export = tgl_raw.replace("-", "") if len(tgl_raw) == 10 and tgl_raw[4] == "-" else tgl_raw.replace("-", "")

    total_gaji = sum(int(i.get("jumlah") or 0) for i in items)
    jumlah_penerima = len(items)

    header = [
        "P",
        tgl_export,
        str(laporan.get("rekening_sumber") or ""),
        str(jumlah_penerima),
        str(total_gaji),
    ]
    header += [""] * (PIC_TRANSFER_COL_COUNT - len(header))
    rows = [header[:PIC_TRANSFER_COL_COUNT]]

    bank = laporan.get("bank") or "MANDIRI"
    kota = laporan.get("kota") or "Madiun"

    for item in items:
        row = [""] * PIC_TRANSFER_COL_COUNT
        row[0] = str(item.get("nomor_rekening") or "")
        row[1] = _nama_from_gaji_item(item)
        row[6] = str(int(item.get("jumlah") or 0))
        row[9] = "IBU"
        row[11] = bank
        row[12] = kota
        row[16] = "N"
        row[37] = "OUR"
        row[38] = "1"
        row[39] = "E"
        rows.append(row)
    return rows


def build_pic_transfer_xlsx_bytes(laporan: Dict, items: List[Dict]) -> bytes:
    """Bangun file Excel format transfer massal Mandiri — struktur sama dengan file upload."""
    from openpyxl import Workbook
    from io import BytesIO

    rows = build_pic_transfer_export_rows(laporan, items)
    wb = Workbook()
    ws = wb.active
    ws.title = "Transfer Massal"

    numeric_cols = {4, 5, 7}
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell_val = val
            if c_idx in numeric_cols and str(val).isdigit():
                cell_val = int(val)
            cell = ws.cell(row=r_idx, column=c_idx, value=cell_val)
            if isinstance(cell_val, int):
                cell.number_format = "#,##0"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def parse_gaji_relawan_csv(file_path: str, filename: str = "") -> Dict[str, Any]:
    return parse_pic_transfer_csv(file_path, filename, "Gaji Relawan")


def parse_gaji_relawan_xlsx(file_path: str, filename: str = "") -> Dict[str, Any]:
    return parse_pic_transfer_xlsx(file_path, filename, "Gaji Relawan")


def parse_insentif_pic_csv(file_path: str, filename: str = "") -> Dict[str, Any]:
    return parse_pic_transfer_csv(file_path, filename, "Insentif PIC")


def parse_insentif_pic_xlsx(file_path: str, filename: str = "") -> Dict[str, Any]:
    return parse_pic_transfer_xlsx(file_path, filename, "Insentif PIC")


def resolve_insentif_pic_export(upload_id: int = 0) -> tuple:
    """Ambil laporan + baris untuk export sesuai batch upload Insentif PIC."""
    laporans = get_insentif_pic_laporans(active_only=True)
    laporan = None

    if upload_id:
        laporan = next((lap for lap in laporans if lap["upload_id"] == upload_id), None)
        if not laporan:
            conn = get_db()
            row = conn.execute(
                """
                SELECT g.*, u.filename AS upload_filename
                FROM insentif_pic_laporan g
                LEFT JOIN uploads u ON u.id = g.upload_id
                WHERE g.upload_id = ?
                """,
                (upload_id,),
            ).fetchone()
            conn.close()
            if row:
                laporan = dict(row)
    elif laporans:
        laporan = laporans[0]
        upload_id = laporan["upload_id"]

    if not laporan or not upload_id:
        return None, [], 0

    items = get_insentif_pic({"upload_id": upload_id})
    return laporan, items, upload_id


def get_insentif_pic(filters: Dict = None) -> List[Dict]:
    f = dict(filters or {})
    f["kategori"] = "insentif_pic"
    rows = get_all_tagihan(f)
    return sorted(rows, key=_sort_key_no)


def get_insentif_pic_laporans(active_only: bool = False) -> List[Dict[str, Any]]:
    conn = get_db()
    query = """
        SELECT g.*, u.record_count, u.created_at AS upload_at,
               (SELECT COUNT(*) FROM tagihan t
                WHERE t.upload_id = g.upload_id AND t.kategori = 'insentif_pic') AS item_count
        FROM insentif_pic_laporan g
        LEFT JOIN uploads u ON u.id = g.upload_id
    """
    if active_only:
        query += """
        WHERE EXISTS (
            SELECT 1 FROM tagihan t
            WHERE t.upload_id = g.upload_id AND t.kategori = 'insentif_pic'
        )
        """
    query += " ORDER BY g.id DESC"
    rows = conn.execute(query).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def recalc_insentif_pic_laporan(conn, upload_id: int):
    if not upload_id:
        return
    rows = conn.execute(
        "SELECT jumlah FROM tagihan WHERE upload_id = ? AND kategori = 'insentif_pic'",
        (upload_id,),
    ).fetchall()
    count = len(rows)
    total = sum(r[0] or 0 for r in rows)
    if count == 0:
        conn.execute("DELETE FROM insentif_pic_laporan WHERE upload_id = ?", (upload_id,))
    else:
        conn.execute("""
            UPDATE insentif_pic_laporan SET jumlah_penerima = ?, total_gaji = ?
            WHERE upload_id = ?
        """, (count, total, upload_id))
    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (count, upload_id))




def parse_insentif_mitra_csv(file_path: str, filename: str = "") -> Dict[str, Any]:
    return parse_pic_transfer_csv(file_path, filename, "Insentif Mitra")


def parse_insentif_mitra_xlsx(file_path: str, filename: str = "") -> Dict[str, Any]:
    return parse_pic_transfer_xlsx(file_path, filename, "Insentif Mitra")


def resolve_insentif_mitra_export(upload_id: int = 0) -> tuple:
    """Ambil laporan + baris untuk export sesuai batch upload Insentif Mitra."""
    laporans = get_insentif_mitra_laporans(active_only=True)
    laporan = None

    if upload_id:
        laporan = next((lap for lap in laporans if lap["upload_id"] == upload_id), None)
        if not laporan:
            conn = get_db()
            row = conn.execute(
                """
                SELECT g.*, u.filename AS upload_filename
                FROM insentif_mitra_laporan g
                LEFT JOIN uploads u ON u.id = g.upload_id
                WHERE g.upload_id = ?
                """,
                (upload_id,),
            ).fetchone()
            conn.close()
            if row:
                laporan = dict(row)
    elif laporans:
        laporan = laporans[0]
        upload_id = laporan["upload_id"]

    if not laporan or not upload_id:
        return None, [], 0

    items = get_insentif_mitra({"upload_id": upload_id})
    return laporan, items, upload_id


def get_insentif_mitra(filters: Dict = None) -> List[Dict]:
    f = dict(filters or {})
    f["kategori"] = "insentif_mitra"
    rows = get_all_tagihan(f)
    return sorted(rows, key=_sort_key_no)


def get_insentif_mitra_laporans(active_only: bool = False) -> List[Dict[str, Any]]:
    conn = get_db()
    query = """
        SELECT g.*, u.record_count, u.created_at AS upload_at,
               (SELECT COUNT(*) FROM tagihan t
                WHERE t.upload_id = g.upload_id AND t.kategori = 'insentif_mitra') AS item_count
        FROM insentif_mitra_laporan g
        LEFT JOIN uploads u ON u.id = g.upload_id
    """
    if active_only:
        query += """
        WHERE EXISTS (
            SELECT 1 FROM tagihan t
            WHERE t.upload_id = g.upload_id AND t.kategori = 'insentif_mitra'
        )
        """
    query += " ORDER BY g.id DESC"
    rows = conn.execute(query).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def recalc_insentif_mitra_laporan(conn, upload_id: int):
    if not upload_id:
        return
    rows = conn.execute(
        "SELECT jumlah FROM tagihan WHERE upload_id = ? AND kategori = 'insentif_mitra'",
        (upload_id,),
    ).fetchall()
    count = len(rows)
    total = sum(r[0] or 0 for r in rows)
    if count == 0:
        conn.execute("DELETE FROM insentif_mitra_laporan WHERE upload_id = ?", (upload_id,))
    else:
        conn.execute("""
            UPDATE insentif_mitra_laporan SET jumlah_penerima = ?, total_gaji = ?
            WHERE upload_id = ?
        """, (count, total, upload_id))
    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (count, upload_id))


def get_pengembalian_dana(filters: Dict = None) -> List[Dict]:
    f = dict(filters or {})
    f["kategori"] = "pengembalian_dana"
    rows = get_all_tagihan(f)
    return sorted(rows, key=_sort_key_no)


def get_pengembalian_dana_laporans(active_only: bool = False) -> List[Dict[str, Any]]:
    conn = get_db()
    query = """
        SELECT g.*, u.record_count, u.created_at AS upload_at,
               (SELECT COUNT(*) FROM tagihan t
                WHERE t.upload_id = g.upload_id AND t.kategori = 'pengembalian_dana') AS item_count
        FROM pengembalian_dana_laporan g
        LEFT JOIN uploads u ON u.id = g.upload_id
    """
    if active_only:
        query += """
        WHERE EXISTS (
            SELECT 1 FROM tagihan t
            WHERE t.upload_id = g.upload_id AND t.kategori = 'pengembalian_dana'
        )
        """
    query += " ORDER BY g.id DESC"
    rows = conn.execute(query).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def recalc_pengembalian_dana_laporan(conn, upload_id: int):
    if not upload_id:
        return
    rows = conn.execute(
        "SELECT jumlah FROM tagihan WHERE upload_id = ? AND kategori = 'pengembalian_dana'",
        (upload_id,),
    ).fetchall()
    count = len(rows)
    total = sum(r[0] or 0 for r in rows)
    if count == 0:
        conn.execute("DELETE FROM pengembalian_dana_laporan WHERE upload_id = ?", (upload_id,))
    else:
        conn.execute("""
            UPDATE pengembalian_dana_laporan SET jumlah_penerima = ?, total_gaji = ?
            WHERE upload_id = ?
        """, (count, total, upload_id))
    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (count, upload_id))


def get_pengajuan_dana_mitra_tanggal_options() -> List[str]:
    """Daftar tanggal pengajuan unik dari laporan PDF & entri manual."""
    conn = get_db()
    dates = set()
    for row in conn.execute("""
        SELECT DISTINCT tanggal_pembayaran AS t FROM pengajuan_dana_mitra_laporan
        WHERE tanggal_pembayaran IS NOT NULL AND tanggal_pembayaran != ''
        UNION
        SELECT DISTINCT tanggal AS t FROM tagihan
        WHERE kategori = 'pengajuan_dana_mitra' AND tanggal IS NOT NULL AND tanggal != ''
    """).fetchall():
        if row[0]:
            dates.add(row[0])
    conn.close()
    return sorted(dates, reverse=True)


def get_pengajuan_dana_mitra(filters: Dict = None) -> List[Dict]:
    """Filter tanggal = Tanggal Pengajuan dari formulir PDF (header upload)."""
    f = dict(filters or {})
    tanggal_pengajuan = f.pop("tanggal", None)
    search = f.pop("search", None)
    status = f.pop("status", None)
    rekening = f.pop("rekening", None)
    upload_id = f.pop("upload_id", None)

    conn = get_db()
    query = """
        SELECT t.* FROM tagihan t
        LEFT JOIN pengajuan_dana_mitra_laporan g ON g.upload_id = t.upload_id
    """
    where = ["t.kategori = 'pengajuan_dana_mitra'"]
    params: List[Any] = []

    if search:
        where.append("(t.pengajuan LIKE ? OR t.atas_nama LIKE ? OR t.bank LIKE ?)")
        s = f"%{search}%"
        params.extend([s, s, s])
    if status:
        where.append("t.status = ?")
        params.append(status)
    if rekening:
        where.append("t.rekening = ?")
        params.append(rekening)
    if tanggal_pengajuan:
        where.append("(t.tanggal = ? OR g.tanggal_pembayaran = ?)")
        params.extend([tanggal_pengajuan, tanggal_pengajuan])
    if upload_id:
        where.append("t.upload_id = ?")
        params.append(upload_id)

    query += " WHERE " + " AND ".join(where)
    query += " ORDER BY COALESCE(t.tanggal, g.tanggal_pembayaran, '9999-12-31') DESC, t.id DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return sorted([dict(r) for r in rows], key=_sort_key_no)


def get_pengajuan_dana_mitra_laporans(active_only: bool = False) -> List[Dict[str, Any]]:
    conn = get_db()
    query = """
        SELECT g.*, u.record_count, u.created_at AS upload_at,
               (SELECT COUNT(*) FROM tagihan t
                WHERE t.upload_id = g.upload_id AND t.kategori = 'pengajuan_dana_mitra') AS item_count
        FROM pengajuan_dana_mitra_laporan g
        LEFT JOIN uploads u ON u.id = g.upload_id
    """
    if active_only:
        query += """
        WHERE EXISTS (
            SELECT 1 FROM tagihan t
            WHERE t.upload_id = g.upload_id AND t.kategori = 'pengajuan_dana_mitra'
        )
        """
    query += " ORDER BY g.id DESC"
    rows = conn.execute(query).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def recalc_pengajuan_dana_mitra_laporan(conn, upload_id: int):
    if not upload_id:
        return
    rows = conn.execute(
        "SELECT jumlah FROM tagihan WHERE upload_id = ? AND kategori = 'pengajuan_dana_mitra'",
        (upload_id,),
    ).fetchall()
    count = len(rows)
    total = sum(r[0] or 0 for r in rows)
    if count == 0:
        conn.execute("DELETE FROM pengajuan_dana_mitra_laporan WHERE upload_id = ?", (upload_id,))
    else:
        conn.execute("""
            UPDATE pengajuan_dana_mitra_laporan SET jumlah_penerima = ?, total_gaji = ?
            WHERE upload_id = ?
        """, (count, total, upload_id))
    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (count, upload_id))


@app.get("/gaji-relawan", response_class=HTMLResponse)
async def gaji_relawan_page(
    request: Request,
    user=Depends(require_login),
    upload_id: int = 0,
    search: str = "",
    status: str = "",
    rekening: str = "",
    tanggal: str = "",
    message: str = "",
    success: bool = False,
):
    periode_options = get_gaji_relawan_laporans(active_only=True)
    laporans = periode_options
    laporan = None
    active_upload_ids = {lap["upload_id"] for lap in periode_options}

    filters = {}
    if search:
        filters["search"] = search
    if status:
        filters["status"] = status
    if rekening:
        filters["rekening"] = rekening
    if tanggal:
        filters["tanggal"] = tanggal

    if upload_id and upload_id not in active_upload_ids:
        upload_id = 0

    view_all = upload_id == 0

    if upload_id and upload_id in active_upload_ids:
        laporan = next((lap for lap in periode_options if lap["upload_id"] == upload_id), None)
        filters["upload_id"] = upload_id

    data = get_gaji_relawan(filters)

    periode_map = {lap["upload_id"]: lap.get("periode") or lap.get("filename") for lap in laporans}
    for item in data:
        uid = item.get("upload_id")
        item["periode_label"] = periode_map.get(uid, "Input Manual") if uid else "Input Manual"

    if not data:
        main_status = "—"
    else:
        all_terbayar = all((item.get("status") or "").upper() == "TERBAYAR" for item in data)
        main_status = "Sukses" if all_terbayar else "Belum Lunas"

    conn = get_db()
    gr_filter = "kategori = 'gaji_relawan'"
    statuses = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT status FROM tagihan WHERE status IS NOT NULL AND {gr_filter}"
        ).fetchall()
    ]
    rekenings = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT rekening FROM tagihan WHERE rekening IS NOT NULL AND {gr_filter}"
        ).fetchall()
    ]
    conn.close()

    for item in data:
        item["fee_payrol"] = FEE_PAYROL_PER_ORANG
        item["total_bayar"] = (item.get("jumlah") or 0) + FEE_PAYROL_PER_ORANG

    total_filtered = sum(d["jumlah"] for d in data)
    total_fee_payrol = calc_fee_payrol(len(data))
    total_grand = total_filtered + total_fee_payrol
    avg_gaji = int(total_filtered / len(data)) if data else 0
    max_item = max(data, key=lambda x: x["jumlah"]) if data else None
    min_item = min(data, key=lambda x: x["jumlah"]) if data else None
    total_all_batches = sum(lap.get("total_gaji") or 0 for lap in laporans)
    total_all_relawan = sum(lap.get("jumlah_penerima") or lap.get("record_count") or 0 for lap in laporans)

    return render_template(request, "gaji_relawan.html", {
        "user": user,
        "items": data,
        "laporans": laporans,
        "laporan": laporan,
        "view_all": view_all,
        "selected_upload_id": upload_id if upload_id else 0,
        "total_filtered": total_filtered,
        "total_fee_payrol": total_fee_payrol,
        "total_grand": total_grand,
        "fee_payrol_amount": FEE_PAYROL_PER_ORANG,
        "total_all_batches": total_all_batches,
        "total_all_relawan": total_all_relawan,
        "avg_gaji": avg_gaji,
        "max_item": max_item,
        "min_item": min_item,
        "main_status": main_status,
        "format_rupiah": format_rupiah,
        "format_tanggal": format_tanggal_display,
        "filters": {"search": search, "status": status, "rekening": rekening, "tanggal": tanggal, "periode": upload_id},
        "periode_options": periode_options,
        "status_options": sorted([s for s in statuses if s]),
        "rekening_options": sorted([r for r in rekenings if r]),
        "message": message,
        "success": success,
    })


@app.post("/gaji-relawan/upload")
async def gaji_relawan_upload(
    request: Request,
    user=Depends(require_login),
    file: UploadFile = File(...),
):
    """Upload Excel daftar gaji relawan — format transfer massal Mandiri; download CSV mengekstrak ke format bank."""
    if is_viewer(user):
        return redirect_with_flash(
            request,
            "/gaji-relawan",
            "Akses ditolak. Akun Anda hanya dapat melihat data.",
        )
    if not is_admin(user) and not can_member_upload(user):
        return redirect_with_flash(
            request,
            "/gaji-relawan",
            "Akses ditolak. Anda tidak dapat mengunggah file.",
        )
    safe_name = f"{date.today().isoformat()}_{file.filename.replace(' ', '_')}"
    file_path = os.path.join(UPLOAD_DIR, safe_name)
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    ext = (file.filename or "").lower().rsplit(".", 1)[-1] if "." in (file.filename or "") else ""

    conn = get_db()
    conn.execute(
        "INSERT INTO uploads (filename, created_by) VALUES (?, ?)",
        (file.filename, user["id"]),
    )
    upload_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    try:
        if ext in ("xlsx", "xls"):
            parsed = parse_gaji_relawan_xlsx(file_path, file.filename)
        else:
            conn.close()
            return redirect_with_flash(
                request,
                "/gaji-relawan",
                "Format tidak didukung. Upload file Excel (.xlsx / .xls).",
            )
    except Exception as e:
        conn.close()
        return redirect_with_flash(request, "/gaji-relawan", f"Error parsing file: {str(e)}")

    item_list = parsed.get("items", []) if isinstance(parsed, dict) else []
    gr_meta = parsed.get("meta", {}) if isinstance(parsed, dict) else {}

    if not item_list:
        conn.close()
        return redirect_with_flash(request, "/gaji-relawan", "Tidak ada data yang bisa dibaca dari file")

    inserted = 0
    for item in item_list:
        pengajuan = item.get("pengajuan") or item.get("atas_nama") or ""
        if not pengajuan:
            continue
        jumlah = int(item.get("jumlah") or 0)
        if jumlah <= 0:
            continue

        dup = conn.execute("""
            SELECT 1 FROM tagihan
            WHERE upload_id = ? AND pengajuan = ? AND jumlah = ?
            LIMIT 1
        """, (upload_id, pengajuan, jumlah)).fetchone()
        if dup:
            continue

        conn.execute("""
            INSERT INTO tagihan
            (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, kategori, upload_id, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'gaji_relawan', ?, ?)
        """, (
            item.get("no"),
            pengajuan,
            jumlah,
            item.get("status") or "DIAJUKAN",
            item.get("rekening") or "TRANSFER MASSAL",
            item.get("tanggal") or gr_meta.get("tanggal_pembayaran"),
            item.get("atas_nama"),
            item.get("nomor_rekening"),
            item.get("bank"),
            upload_id,
            user["id"],
        ))
        inserted += 1

    conn.execute("""
        INSERT OR REPLACE INTO gaji_relawan_laporan
        (upload_id, tanggal_pembayaran, rekening_sumber, jumlah_penerima, total_gaji,
         periode, bank, kota, filename)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        upload_id,
        gr_meta.get("tanggal_pembayaran"),
        gr_meta.get("rekening_sumber"),
        gr_meta.get("jumlah_penerima") or inserted,
        gr_meta.get("total_gaji") or sum(i.get("jumlah", 0) for i in item_list),
        gr_meta.get("periode"),
        gr_meta.get("bank") or "MANDIRI",
        gr_meta.get("kota"),
        file.filename,
    ))
    conn.execute("UPDATE uploads SET record_count = ?, periode = ? WHERE id = ?",
                 (inserted, gr_meta.get("periode"), upload_id))
    conn.commit()
    conn.close()

    msg = f"Berhasil mengunggah {inserted} data gaji relawan"
    if gr_meta.get("periode"):
        msg += f" ({gr_meta['periode']})"
    return redirect_with_flash(request, f"/gaji-relawan?upload_id={upload_id}", msg, success=True)


@app.post("/gaji-relawan")
async def create_gaji_relawan(
    request: Request,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form("DIAJUKAN"),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        INSERT INTO tagihan (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, kategori, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'gaji_relawan', ?)
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or "DIAJUKAN",
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        user["id"],
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/gaji-relawan", status_code=303)


@app.post("/gaji-relawan/{item_id}/update")
async def update_gaji_relawan(
    item_id: int,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form(""),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        UPDATE tagihan SET
            no = ?, pengajuan = ?, jumlah = ?, status = ?, rekening = ?,
            tanggal = ?, atas_nama = ?, nomor_rekening = ?, bank = ?,
            kategori = 'gaji_relawan', updated_at = CURRENT_TIMESTAMP
        WHERE id = ? AND kategori = 'gaji_relawan'
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or None,
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        item_id,
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/gaji-relawan", status_code=303)


@app.post("/gaji-relawan/{item_id}/delete")
async def delete_gaji_relawan(item_id: int, user=Depends(require_login)):
    conn = get_db()
    row = conn.execute(
        "SELECT upload_id FROM tagihan WHERE id = ? AND kategori = 'gaji_relawan'", (item_id,)
    ).fetchone()
    conn.execute("DELETE FROM tagihan WHERE id = ? AND kategori = 'gaji_relawan'", (item_id,))
    if row and row[0]:
        recalc_gaji_relawan_laporan(conn, row[0])
    conn.commit()
    conn.close()
    return RedirectResponse("/gaji-relawan", status_code=303)


@app.post("/api/gaji-relawan/bulk-delete")
async def api_gaji_relawan_bulk_delete(request: Request, user=Depends(require_login)):
    """Hapus baris terpilih saja — upload & data periode lain tidak ikut terhapus."""
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        if not ids:
            return JSONResponse({"error": "no valid ids"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        upload_rows = conn.execute(
            f"""SELECT DISTINCT upload_id FROM tagihan
                WHERE id IN ({placeholders}) AND kategori = 'gaji_relawan' AND upload_id IS NOT NULL""",
            ids,
        ).fetchall()
        conn.execute(
            f"DELETE FROM tagihan WHERE id IN ({placeholders}) AND kategori = 'gaji_relawan'",
            ids,
        )
        deleted = conn.total_changes
        for row in upload_rows:
            recalc_gaji_relawan_laporan(conn, row[0])
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "deleted": deleted})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/gaji-relawan/bulk-update")
async def api_gaji_relawan_bulk_update(request: Request, user=Depends(require_login)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        status = data.get("status")
        if not ids or status is None:
            return JSONResponse({"error": "no valid ids or status"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        conn.execute(
            f"""UPDATE tagihan SET status = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id IN ({placeholders}) AND kategori = 'gaji_relawan'""",
            [status] + ids,
        )
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "updated": len(ids)})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


def _pdf_sanitize_text(text) -> str:
    if text is None:
        return ""
    return str(text).encode("latin-1", "replace").decode("latin-1")


def _pdf_fit_cell(pdf, text, width_mm: float) -> str:
    text = _pdf_sanitize_text(text)
    if not text:
        return ""
    usable = max(width_mm - 2, 4)
    if pdf.get_string_width(text) <= usable:
        return text
    ell = "..."
    trimmed = text
    while trimmed and pdf.get_string_width(trimmed + ell) > usable:
        trimmed = trimmed[:-1]
    return (trimmed + ell) if trimmed else ell


def _gaji_relawan_pdf_status(_status: str = "") -> str:
    """Label status di PDF rincian transaksi — selalu Sukses."""
    return "Sukses"


def _enrich_gaji_relawan_export_items(laporan: Dict, items: List[Dict]) -> List[Dict]:
    periode_label = laporan.get("periode") or laporan.get("filename") or "—"
    enriched = []
    for item in items:
        row = dict(item)
        row["periode_label"] = periode_label
        row["fee_payrol"] = FEE_PAYROL_PER_ORANG
        row["total_bayar"] = int(row.get("jumlah") or 0) + FEE_PAYROL_PER_ORANG
        enriched.append(row)
    return enriched


def _build_gaji_relawan_pdf(laporan: Dict, items: List[Dict]) -> bytes:
    """PDF rincian transaksi gaji relawan — ringkasan + tabel sama dengan halaman web."""
    from fpdf import FPDF

    items = _enrich_gaji_relawan_export_items(laporan, items)
    periode_label = laporan.get("periode") or laporan.get("filename") or "—"
    total_gaji = sum(int(i.get("jumlah") or 0) for i in items)
    total_fee = calc_fee_payrol(len(items))
    total_grand = total_gaji + total_fee

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False)
    pdf.set_margins(10, 10, 10)
    pdf.add_page()

    page_bottom = pdf.h - pdf.b_margin
    table_width = 277
    col_widths = [10, 58, 28, 36, 22, 32, 28, 32, 31]
    headers = ["NO", "NAMA RELAWAN", "PERIODE", "REK. TUJUAN", "BANK", "JUMLAH GAJI", "FEE PAYROL", "TOTAL", "STATUS"]
    row_h = 7
    header_h = 8

    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(7, 30, 73)
    pdf.cell(0, 9, "LAPORAN GAJI RELAWAN - SPPG WISMA HAJI MADIUN", ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(80, 80, 80)
    subtitle = f"Periode: {periode_label}  |  Dicetak: {date.today().strftime('%d %B %Y')}"
    if laporan.get("tanggal_pembayaran"):
        subtitle += f"  |  Tanggal bayar: {laporan['tanggal_pembayaran']}"
    pdf.cell(0, 6, _pdf_sanitize_text(subtitle), ln=True, align="C")
    pdf.ln(3)

    kpi_w = table_width / 4
    kpi_h = 18
    kpi_data = [
        ("JUMLAH PENERIMA", str(len(items)), "relawan / transaksi"),
        ("JUMLAH GAJI", format_rupiah(total_gaji), "total gaji relawan"),
        ("JUMLAH FEE PAYROL", format_rupiah(total_fee), f"{len(items)} x {format_rupiah(FEE_PAYROL_PER_ORANG)}"),
        ("GRAND TOTAL", format_rupiah(total_grand), "gaji + fee payrol"),
    ]
    kpi_y = pdf.get_y()
    for col_idx, (label, value, note) in enumerate(kpi_data):
        x0 = pdf.l_margin + col_idx * kpi_w
        pdf.set_fill_color(248, 250, 252)
        pdf.set_draw_color(226, 232, 240)
        pdf.rect(x0, kpi_y, kpi_w, kpi_h, style="DF")
        pdf.set_xy(x0 + 3, kpi_y + 2)
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_text_color(7, 30, 73)
        pdf.cell(kpi_w - 6, 4, label, ln=0)
        pdf.set_xy(x0 + 3, kpi_y + 6)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(kpi_w - 6, 6, _pdf_fit_cell(pdf, value, kpi_w - 6), ln=0)
        pdf.set_xy(x0 + 3, kpi_y + 12)
        pdf.set_font("Helvetica", "", 6)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(kpi_w - 6, 4, _pdf_fit_cell(pdf, note, kpi_w - 6), ln=0)
    pdf.set_xy(pdf.l_margin, kpi_y + kpi_h + 4)

    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(7, 30, 73)
    pdf.cell(0, 7, "Rincian Transaksi", ln=True)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, f"{len(items)} baris pembayaran", ln=True)
    pdf.ln(2)

    def _draw_table_header():
        pdf.set_x(pdf.l_margin)
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(7, 30, 73)
        pdf.set_text_color(255, 255, 255)
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], header_h, header, border=1, align="C", ln=0, fill=True)
        pdf.ln(header_h)

    def _ensure_row_space(height: float):
        if pdf.get_y() + height > page_bottom:
            pdf.add_page()
            _draw_table_header()
            pdf.set_font("Helvetica", "", 7)
            pdf.set_text_color(30, 30, 30)

    def _draw_data_row(cells: List[str], aligns: List[str], fill: bool):
        _ensure_row_space(row_h)
        pdf.set_x(pdf.l_margin)
        pdf.set_fill_color(248, 250, 252) if fill else pdf.set_fill_color(255, 255, 255)
        for i, (text, align) in enumerate(zip(cells, aligns)):
            pdf.cell(
                col_widths[i],
                row_h,
                _pdf_fit_cell(pdf, text, col_widths[i]),
                border=1,
                align=align,
                ln=0,
                fill=True,
            )
        pdf.ln(row_h)

    _draw_table_header()
    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(30, 30, 30)
    fill = False
    for idx, item in enumerate(items, start=1):
        cells = [
            str(item.get("no") or idx),
            _nama_from_gaji_item(item),
            item.get("periode_label") or periode_label,
            str(item.get("nomor_rekening") or ""),
            str(item.get("bank") or ""),
            format_rupiah(int(item.get("jumlah") or 0)),
            format_rupiah(int(item.get("fee_payrol") or FEE_PAYROL_PER_ORANG)),
            format_rupiah(int(item.get("total_bayar") or 0)),
            _gaji_relawan_pdf_status(item.get("status")),
        ]
        aligns = ["C", "L", "L", "L", "C", "R", "R", "R", "C"]
        _draw_data_row(cells, aligns, fill)
        fill = not fill

    _ensure_row_space(row_h + 2)
    pdf.set_x(pdf.l_margin)
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(236, 253, 245)
    pdf.set_text_color(7, 30, 73)
    pdf.cell(sum(col_widths[:5]), row_h, "TOTAL", border=1, align="R", ln=0, fill=True)
    pdf.cell(col_widths[5], row_h, _pdf_fit_cell(pdf, format_rupiah(total_gaji), col_widths[5]), border=1, align="R", ln=0, fill=True)
    pdf.cell(col_widths[6], row_h, _pdf_fit_cell(pdf, format_rupiah(total_fee), col_widths[6]), border=1, align="R", ln=0, fill=True)
    pdf.cell(col_widths[7], row_h, _pdf_fit_cell(pdf, format_rupiah(total_grand), col_widths[7]), border=1, align="R", ln=0, fill=True)
    pdf.cell(col_widths[8], row_h, "", border=1, ln=0, fill=True)
    pdf.ln(row_h)

    return bytes(pdf.output())


@app.get("/export/gaji-relawan/pdf")
async def export_gaji_relawan_pdf(user=Depends(require_login), upload_id: int = 0):
    laporan, items, _ = resolve_gaji_relawan_export(upload_id)
    if not laporan or not items:
        return Response("Tidak ada data gaji relawan untuk diunduh.", status_code=404)

    content = _build_gaji_relawan_pdf(laporan, items)
    filename = _export_filename_from_laporan(laporan, "pdf")
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/export/gaji-relawan/csv")
async def export_gaji_relawan_csv(user=Depends(require_login), upload_id: int = 0):
    import csv
    from io import StringIO

    laporan, items, _ = resolve_gaji_relawan_export(upload_id)
    if not laporan or not items:
        return Response("Tidak ada data gaji relawan untuk diunduh.", status_code=404)

    output = StringIO()
    writer = csv.writer(output)
    for row in build_pic_transfer_export_rows(laporan, items):
        writer.writerow(row)
    output.seek(0)
    filename = _export_filename_from_laporan(laporan, "csv")
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/export/gaji-relawan/xlsx")
async def export_gaji_relawan_xlsx(user=Depends(require_login), upload_id: int = 0):
    laporan, items, _ = resolve_gaji_relawan_export(upload_id)
    if not laporan or not items:
        return Response("Tidak ada data gaji relawan untuk diunduh.", status_code=404)

    filename = _export_filename_from_laporan(laporan, "xlsx")
    return Response(
        content=build_pic_transfer_xlsx_bytes(laporan, items),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/insentif-pic", response_class=HTMLResponse)
async def insentif_pic_page(
    request: Request,
    user=Depends(require_login),
    upload_id: int = 0,
    search: str = "",
    status: str = "",
    rekening: str = "",
    tanggal: str = "",
    message: str = "",
    success: bool = False,
):
    periode_options = get_insentif_pic_laporans(active_only=True)
    laporans = periode_options
    laporan = None
    active_upload_ids = {lap["upload_id"] for lap in periode_options}

    filters = {}
    if search:
        filters["search"] = search
    if status:
        filters["status"] = status
    if rekening:
        filters["rekening"] = rekening
    if tanggal:
        filters["tanggal"] = tanggal

    if upload_id and upload_id not in active_upload_ids:
        upload_id = 0

    view_all = upload_id == 0

    if upload_id and upload_id in active_upload_ids:
        laporan = next((lap for lap in periode_options if lap["upload_id"] == upload_id), None)
        filters["upload_id"] = upload_id

    data = get_insentif_pic(filters)

    periode_map = {lap["upload_id"]: lap.get("periode") or lap.get("filename") for lap in laporans}
    for item in data:
        uid = item.get("upload_id")
        item["periode_label"] = periode_map.get(uid, "Input Manual") if uid else "Input Manual"

    if not data:
        main_status = "—"
    else:
        all_terbayar = all((item.get("status") or "").upper() == "TERBAYAR" for item in data)
        main_status = "Sukses" if all_terbayar else "Belum Lunas"

    conn = get_db()
    ip_filter = "kategori = 'insentif_pic'"
    statuses = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT status FROM tagihan WHERE status IS NOT NULL AND {ip_filter}"
        ).fetchall()
    ]
    rekenings = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT rekening FROM tagihan WHERE rekening IS NOT NULL AND {ip_filter}"
        ).fetchall()
    ]
    conn.close()

    for item in data:
        item["fee_payrol"] = FEE_PAYROL_PER_ORANG
        item["total_bayar"] = (item.get("jumlah") or 0) + FEE_PAYROL_PER_ORANG

    total_filtered = sum(d["jumlah"] for d in data)
    total_fee_payrol = calc_fee_payrol(len(data))
    total_grand = total_filtered + total_fee_payrol
    avg_gaji = int(total_filtered / len(data)) if data else 0
    max_item = max(data, key=lambda x: x["jumlah"]) if data else None
    min_item = min(data, key=lambda x: x["jumlah"]) if data else None
    total_all_batches = sum(lap.get("total_gaji") or 0 for lap in laporans)
    total_all_pic = sum(lap.get("jumlah_penerima") or lap.get("record_count") or 0 for lap in laporans)

    return render_template(request, "insentif_pic.html", {
        "user": user,
        "items": data,
        "laporans": laporans,
        "laporan": laporan,
        "view_all": view_all,
        "selected_upload_id": upload_id if upload_id else 0,
        "total_filtered": total_filtered,
        "total_fee_payrol": total_fee_payrol,
        "total_grand": total_grand,
        "fee_payrol_amount": FEE_PAYROL_PER_ORANG,
        "total_all_batches": total_all_batches,
        "total_all_pic": total_all_pic,
        "avg_gaji": avg_gaji,
        "max_item": max_item,
        "min_item": min_item,
        "main_status": main_status,
        "format_rupiah": format_rupiah,
        "format_tanggal": format_tanggal_display,
        "filters": {"search": search, "status": status, "rekening": rekening, "tanggal": tanggal, "periode": upload_id},
        "periode_options": periode_options,
        "status_options": sorted([s for s in statuses if s]),
        "rekening_options": sorted([r for r in rekenings if r]),
        "message": message,
        "success": success,
    })


@app.post("/insentif-pic/upload")
async def insentif_pic_upload(
    request: Request,
    user=Depends(require_login),
    file: UploadFile = File(...),
):
    """Upload Excel insentif PIC — format transfer massal Mandiri; download CSV mengekstrak ke format bank."""
    if is_viewer(user):
        return redirect_with_flash(
            request,
            "/insentif-pic",
            "Akses ditolak. Akun Anda hanya dapat melihat data.",
        )
    if not is_admin(user) and not can_member_upload(user):
        return redirect_with_flash(
            request,
            "/insentif-pic",
            "Akses ditolak. Anda tidak dapat mengunggah file.",
        )
    safe_name = f"{date.today().isoformat()}_{file.filename.replace(' ', '_')}"
    file_path = os.path.join(UPLOAD_DIR, safe_name)
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    ext = (file.filename or "").lower().rsplit(".", 1)[-1] if "." in (file.filename or "") else ""

    conn = get_db()
    conn.execute(
        "INSERT INTO uploads (filename, created_by) VALUES (?, ?)",
        (file.filename, user["id"]),
    )
    upload_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    try:
        if ext in ("xlsx", "xls"):
            parsed = parse_insentif_pic_xlsx(file_path, file.filename)
        else:
            conn.close()
            return redirect_with_flash(
                request,
                "/insentif-pic",
                "Format tidak didukung. Upload file Excel (.xlsx / .xls).",
            )
    except Exception as e:
        conn.close()
        return redirect_with_flash(request, "/insentif-pic", f"Error parsing file: {str(e)}")

    item_list = parsed.get("items", []) if isinstance(parsed, dict) else []
    ip_meta = parsed.get("meta", {}) if isinstance(parsed, dict) else {}

    if not item_list:
        conn.close()
        return redirect_with_flash(request, "/insentif-pic", "Tidak ada data yang bisa dibaca dari file")

    inserted = 0
    for item in item_list:
        pengajuan = item.get("pengajuan") or item.get("atas_nama") or ""
        if not pengajuan:
            continue
        jumlah = int(item.get("jumlah") or 0)
        if jumlah <= 0:
            continue

        dup = conn.execute("""
            SELECT 1 FROM tagihan
            WHERE upload_id = ? AND pengajuan = ? AND jumlah = ?
            LIMIT 1
        """, (upload_id, pengajuan, jumlah)).fetchone()
        if dup:
            continue

        conn.execute("""
            INSERT INTO tagihan
            (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, kategori, upload_id, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'insentif_pic', ?, ?)
        """, (
            item.get("no"),
            pengajuan,
            jumlah,
            item.get("status") or "DIAJUKAN",
            item.get("rekening") or "TRANSFER MASSAL",
            item.get("tanggal") or ip_meta.get("tanggal_pembayaran"),
            item.get("atas_nama"),
            item.get("nomor_rekening"),
            item.get("bank"),
            upload_id,
            user["id"],
        ))
        inserted += 1

    conn.execute("""
        INSERT OR REPLACE INTO insentif_pic_laporan
        (upload_id, tanggal_pembayaran, rekening_sumber, jumlah_penerima, total_gaji,
         periode, bank, kota, filename)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        upload_id,
        ip_meta.get("tanggal_pembayaran"),
        ip_meta.get("rekening_sumber"),
        ip_meta.get("jumlah_penerima") or inserted,
        ip_meta.get("total_gaji") or sum(i.get("jumlah", 0) for i in item_list),
        ip_meta.get("periode"),
        ip_meta.get("bank") or "MANDIRI",
        ip_meta.get("kota"),
        file.filename,
    ))
    conn.execute("UPDATE uploads SET record_count = ?, periode = ? WHERE id = ?",
                 (inserted, ip_meta.get("periode"), upload_id))
    conn.commit()
    conn.close()

    msg = f"Berhasil mengunggah {inserted} data insentif PIC"
    if ip_meta.get("periode"):
        msg += f" ({ip_meta['periode']})"
    return redirect_with_flash(request, f"/insentif-pic?upload_id={upload_id}", msg, success=True)


@app.post("/insentif-pic")
async def create_insentif_pic(
    request: Request,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form("DIAJUKAN"),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        INSERT INTO tagihan (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, kategori, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'insentif_pic', ?)
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or "DIAJUKAN",
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        user["id"],
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/insentif-pic", status_code=303)


@app.post("/insentif-pic/{item_id}/update")
async def update_insentif_pic(
    item_id: int,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form(""),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        UPDATE tagihan SET
            no = ?, pengajuan = ?, jumlah = ?, status = ?, rekening = ?,
            tanggal = ?, atas_nama = ?, nomor_rekening = ?, bank = ?,
            kategori = 'insentif_pic', updated_at = CURRENT_TIMESTAMP
        WHERE id = ? AND kategori = 'insentif_pic'
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or None,
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        item_id,
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/insentif-pic", status_code=303)


@app.post("/insentif-pic/{item_id}/delete")
async def delete_insentif_pic(item_id: int, user=Depends(require_login)):
    conn = get_db()
    row = conn.execute(
        "SELECT upload_id FROM tagihan WHERE id = ? AND kategori = 'insentif_pic'", (item_id,)
    ).fetchone()
    conn.execute("DELETE FROM tagihan WHERE id = ? AND kategori = 'insentif_pic'", (item_id,))
    if row and row[0]:
        recalc_insentif_pic_laporan(conn, row[0])
    conn.commit()
    conn.close()
    return RedirectResponse("/insentif-pic", status_code=303)


@app.post("/api/insentif-pic/bulk-delete")
async def api_insentif_pic_bulk_delete(request: Request, user=Depends(require_login)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        if not ids:
            return JSONResponse({"error": "no valid ids"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        upload_rows = conn.execute(
            f"""SELECT DISTINCT upload_id FROM tagihan
                WHERE id IN ({placeholders}) AND kategori = 'insentif_pic' AND upload_id IS NOT NULL""",
            ids,
        ).fetchall()
        conn.execute(
            f"DELETE FROM tagihan WHERE id IN ({placeholders}) AND kategori = 'insentif_pic'",
            ids,
        )
        deleted = conn.total_changes
        for row in upload_rows:
            recalc_insentif_pic_laporan(conn, row[0])
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "deleted": deleted})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/insentif-pic/bulk-update")
async def api_insentif_pic_bulk_update(request: Request, user=Depends(require_login)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        status = data.get("status")
        if not ids or status is None:
            return JSONResponse({"error": "no valid ids or status"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        conn.execute(
            f"""UPDATE tagihan SET status = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id IN ({placeholders}) AND kategori = 'insentif_pic'""",
            [status] + ids,
        )
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "updated": len(ids)})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/export/insentif-pic/csv")
async def export_insentif_pic_csv(user=Depends(require_login), upload_id: int = 0):
    import csv
    from io import StringIO

    laporan, items, _ = resolve_insentif_pic_export(upload_id)
    if not laporan or not items:
        return Response("Tidak ada data insentif PIC untuk diunduh.", status_code=404)

    output = StringIO()
    writer = csv.writer(output)
    for row in build_pic_transfer_export_rows(laporan, items):
        writer.writerow(row)
    output.seek(0)
    filename = _export_filename_from_laporan(laporan, "csv")
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/export/insentif-pic/xlsx")
async def export_insentif_pic_xlsx(user=Depends(require_login), upload_id: int = 0):
    laporan, items, _ = resolve_insentif_pic_export(upload_id)
    if not laporan or not items:
        return Response("Tidak ada data insentif PIC untuk diunduh.", status_code=404)

    filename = _export_filename_from_laporan(laporan, "xlsx")
    return Response(
        content=build_pic_transfer_xlsx_bytes(laporan, items),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/insentif-mitra", response_class=HTMLResponse)
async def insentif_mitra_page(
    request: Request,
    user=Depends(require_login),
    upload_id: int = 0,
    search: str = "",
    status: str = "",
    rekening: str = "",
    tanggal: str = "",
    message: str = "",
    success: bool = False,
):
    periode_options = get_insentif_mitra_laporans(active_only=True)
    laporans = periode_options
    laporan = None
    active_upload_ids = {lap["upload_id"] for lap in periode_options}

    filters = {}
    if search:
        filters["search"] = search
    if status:
        filters["status"] = status
    if rekening:
        filters["rekening"] = rekening
    if tanggal:
        filters["tanggal"] = tanggal

    if upload_id and upload_id not in active_upload_ids:
        upload_id = 0

    view_all = upload_id == 0

    if upload_id and upload_id in active_upload_ids:
        laporan = next((lap for lap in periode_options if lap["upload_id"] == upload_id), None)
        filters["upload_id"] = upload_id

    data = get_insentif_mitra(filters)

    periode_map = {lap["upload_id"]: lap.get("periode") or lap.get("filename") for lap in laporans}
    for item in data:
        uid = item.get("upload_id")
        item["periode_label"] = periode_map.get(uid, "Input Manual") if uid else "Input Manual"

    if not data:
        main_status = "—"
    else:
        all_terbayar = all((item.get("status") or "").upper() == "TERBAYAR" for item in data)
        main_status = "Sukses" if all_terbayar else "Belum Lunas"

    conn = get_db()
    ip_filter = "kategori = 'insentif_mitra'"
    statuses = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT status FROM tagihan WHERE status IS NOT NULL AND {ip_filter}"
        ).fetchall()
    ]
    rekenings = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT rekening FROM tagihan WHERE rekening IS NOT NULL AND {ip_filter}"
        ).fetchall()
    ]
    conn.close()

    for item in data:
        item["fee_payrol"] = FEE_PAYROL_PER_ORANG
        item["total_bayar"] = (item.get("jumlah") or 0) + FEE_PAYROL_PER_ORANG

    total_filtered = sum(d["jumlah"] for d in data)
    total_fee_payrol = calc_fee_payrol(len(data))
    total_grand = total_filtered + total_fee_payrol
    avg_gaji = int(total_filtered / len(data)) if data else 0
    max_item = max(data, key=lambda x: x["jumlah"]) if data else None
    min_item = min(data, key=lambda x: x["jumlah"]) if data else None
    total_all_batches = sum(lap.get("total_gaji") or 0 for lap in laporans)
    total_all_mitra = sum(lap.get("jumlah_penerima") or lap.get("record_count") or 0 for lap in laporans)

    return render_template(request, "insentif_mitra.html", {
        "user": user,
        "items": data,
        "laporans": laporans,
        "laporan": laporan,
        "view_all": view_all,
        "selected_upload_id": upload_id if upload_id else 0,
        "total_filtered": total_filtered,
        "total_fee_payrol": total_fee_payrol,
        "total_grand": total_grand,
        "fee_payrol_amount": FEE_PAYROL_PER_ORANG,
        "insentif_mitra_jumlah": INSENTIF_MITRA_JUMLAH,
        "insentif_mitra_per_hari": INSENTIF_MITRA_PER_HARI,
        "insentif_mitra_hari": INSENTIF_MITRA_HARI,
        "total_all_batches": total_all_batches,
        "total_all_mitra": total_all_mitra,
        "avg_gaji": avg_gaji,
        "max_item": max_item,
        "min_item": min_item,
        "main_status": main_status,
        "format_rupiah": format_rupiah,
        "format_tanggal": format_tanggal_display,
        "filters": {"search": search, "status": status, "rekening": rekening, "tanggal": tanggal, "periode": upload_id},
        "periode_options": periode_options,
        "status_options": sorted([s for s in statuses if s]),
        "rekening_options": sorted([r for r in rekenings if r]),
        "message": message,
        "success": success,
    })


@app.post("/insentif-mitra/upload")
async def insentif_mitra_upload(
    request: Request,
    user=Depends(require_login),
    file: UploadFile = File(...),
):
    """Upload Excel insentif mitra — format transfer massal Mandiri; download CSV mengekstrak ke format bank."""
    if is_viewer(user):
        return redirect_with_flash(
            request,
            "/insentif-mitra",
            "Akses ditolak. Akun Anda hanya dapat melihat data.",
        )
    if not is_admin(user) and not can_member_upload(user):
        return redirect_with_flash(
            request,
            "/insentif-mitra",
            "Akses ditolak. Anda tidak dapat mengunggah file.",
        )
    safe_name = f"{date.today().isoformat()}_{file.filename.replace(' ', '_')}"
    file_path = os.path.join(UPLOAD_DIR, safe_name)
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    ext = (file.filename or "").lower().rsplit(".", 1)[-1] if "." in (file.filename or "") else ""

    conn = get_db()
    conn.execute(
        "INSERT INTO uploads (filename, created_by) VALUES (?, ?)",
        (file.filename, user["id"]),
    )
    upload_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    try:
        if ext in ("xlsx", "xls"):
            parsed = parse_insentif_mitra_xlsx(file_path, file.filename)
        else:
            conn.close()
            return redirect_with_flash(
                request,
                "/insentif-mitra",
                "Format tidak didukung. Upload file Excel (.xlsx / .xls).",
            )
    except Exception as e:
        conn.close()
        return redirect_with_flash(request, "/insentif-mitra", f"Error parsing file: {str(e)}")

    item_list = parsed.get("items", []) if isinstance(parsed, dict) else []
    ip_meta = parsed.get("meta", {}) if isinstance(parsed, dict) else {}

    if not item_list:
        conn.close()
        return redirect_with_flash(request, "/insentif-mitra", "Tidak ada data yang bisa dibaca dari file")

    inserted = 0
    for item in item_list:
        pengajuan = item.get("pengajuan") or item.get("atas_nama") or ""
        if not pengajuan:
            continue
        jumlah = int(item.get("jumlah") or 0)
        if jumlah <= 0:
            continue

        dup = conn.execute("""
            SELECT 1 FROM tagihan
            WHERE upload_id = ? AND pengajuan = ? AND jumlah = ?
            LIMIT 1
        """, (upload_id, pengajuan, jumlah)).fetchone()
        if dup:
            continue

        conn.execute("""
            INSERT INTO tagihan
            (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, kategori, upload_id, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'insentif_mitra', ?, ?)
        """, (
            item.get("no"),
            pengajuan,
            jumlah,
            item.get("status") or "DIAJUKAN",
            item.get("rekening") or "TRANSFER MASSAL",
            item.get("tanggal") or ip_meta.get("tanggal_pembayaran"),
            item.get("atas_nama"),
            item.get("nomor_rekening"),
            item.get("bank"),
            upload_id,
            user["id"],
        ))
        inserted += 1

    conn.execute("""
        INSERT OR REPLACE INTO insentif_mitra_laporan
        (upload_id, tanggal_pembayaran, rekening_sumber, jumlah_penerima, total_gaji,
         periode, bank, kota, filename)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        upload_id,
        ip_meta.get("tanggal_pembayaran"),
        ip_meta.get("rekening_sumber"),
        ip_meta.get("jumlah_penerima") or inserted,
        ip_meta.get("total_gaji") or sum(i.get("jumlah", 0) for i in item_list),
        ip_meta.get("periode"),
        ip_meta.get("bank") or "MANDIRI",
        ip_meta.get("kota"),
        file.filename,
    ))
    conn.execute("UPDATE uploads SET record_count = ?, periode = ? WHERE id = ?",
                 (inserted, ip_meta.get("periode"), upload_id))
    conn.commit()
    conn.close()

    msg = f"Berhasil mengunggah {inserted} data insentif mitra"
    if ip_meta.get("periode"):
        msg += f" ({ip_meta['periode']})"
    return redirect_with_flash(request, f"/insentif-mitra?upload_id={upload_id}", msg, success=True)


@app.post("/insentif-mitra")
async def create_insentif_mitra(
    request: Request,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form("DIAJUKAN"),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        INSERT INTO tagihan (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, kategori, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'insentif_mitra', ?)
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or "DIAJUKAN",
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        user["id"],
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/insentif-mitra", status_code=303)


@app.post("/insentif-mitra/{item_id}/update")
async def update_insentif_mitra(
    item_id: int,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form(""),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        UPDATE tagihan SET
            no = ?, pengajuan = ?, jumlah = ?, status = ?, rekening = ?,
            tanggal = ?, atas_nama = ?, nomor_rekening = ?, bank = ?,
            kategori = 'insentif_mitra', updated_at = CURRENT_TIMESTAMP
        WHERE id = ? AND kategori = 'insentif_mitra'
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or None,
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        item_id,
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/insentif-mitra", status_code=303)


@app.post("/insentif-mitra/{item_id}/delete")
async def delete_insentif_mitra(item_id: int, user=Depends(require_login)):
    conn = get_db()
    row = conn.execute(
        "SELECT upload_id FROM tagihan WHERE id = ? AND kategori = 'insentif_mitra'", (item_id,)
    ).fetchone()
    conn.execute("DELETE FROM tagihan WHERE id = ? AND kategori = 'insentif_mitra'", (item_id,))
    if row and row[0]:
        recalc_insentif_mitra_laporan(conn, row[0])
    conn.commit()
    conn.close()
    return RedirectResponse("/insentif-mitra", status_code=303)


@app.post("/api/insentif-mitra/bulk-delete")
async def api_insentif_mitra_bulk_delete(request: Request, user=Depends(require_login)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        if not ids:
            return JSONResponse({"error": "no valid ids"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        upload_rows = conn.execute(
            f"""SELECT DISTINCT upload_id FROM tagihan
                WHERE id IN ({placeholders}) AND kategori = 'insentif_mitra' AND upload_id IS NOT NULL""",
            ids,
        ).fetchall()
        conn.execute(
            f"DELETE FROM tagihan WHERE id IN ({placeholders}) AND kategori = 'insentif_mitra'",
            ids,
        )
        deleted = conn.total_changes
        for row in upload_rows:
            recalc_insentif_mitra_laporan(conn, row[0])
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "deleted": deleted})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/insentif-mitra/bulk-update")
async def api_insentif_mitra_bulk_update(request: Request, user=Depends(require_login)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        status = data.get("status")
        if not ids or status is None:
            return JSONResponse({"error": "no valid ids or status"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        conn.execute(
            f"""UPDATE tagihan SET status = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id IN ({placeholders}) AND kategori = 'insentif_mitra'""",
            [status] + ids,
        )
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "updated": len(ids)})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/export/insentif-mitra/csv")
async def export_insentif_mitra_csv(user=Depends(require_login), upload_id: int = 0):
    import csv
    from io import StringIO

    laporan, items, _ = resolve_insentif_mitra_export(upload_id)
    if not laporan or not items:
        return Response("Tidak ada data insentif mitra untuk diunduh.", status_code=404)

    output = StringIO()
    writer = csv.writer(output)
    for row in build_pic_transfer_export_rows(laporan, items):
        writer.writerow(row)
    output.seek(0)
    filename = _export_filename_from_laporan(laporan, "csv")
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/export/insentif-mitra/xlsx")
async def export_insentif_mitra_xlsx(user=Depends(require_login), upload_id: int = 0):
    laporan, items, _ = resolve_insentif_mitra_export(upload_id)
    if not laporan or not items:
        return Response("Tidak ada data insentif mitra untuk diunduh.", status_code=404)

    filename = _export_filename_from_laporan(laporan, "xlsx")
    return Response(
        content=build_pic_transfer_xlsx_bytes(laporan, items),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/pengembalian-dana", response_class=HTMLResponse)
async def pengembalian_dana_page(
    request: Request,
    user=Depends(require_login),
    upload_id: int = 0,
    search: str = "",
    status: str = "",
    rekening: str = "",
    tanggal: str = "",
    message: str = "",
    success: bool = False,
):
    periode_options = get_pengembalian_dana_laporans(active_only=True)
    laporans = periode_options
    laporan = None
    active_upload_ids = {lap["upload_id"] for lap in periode_options}

    filters = {}
    if search:
        filters["search"] = search
    if status:
        filters["status"] = status
    if rekening:
        filters["rekening"] = rekening
    if tanggal:
        filters["tanggal"] = tanggal

    if upload_id and upload_id not in active_upload_ids:
        upload_id = 0

    view_all = upload_id == 0

    if upload_id and upload_id in active_upload_ids:
        laporan = next((lap for lap in periode_options if lap["upload_id"] == upload_id), None)
        filters["upload_id"] = upload_id

    data = get_pengembalian_dana(filters)

    periode_map = {lap["upload_id"]: lap.get("periode") or lap.get("filename") for lap in laporans}
    for item in data:
        uid = item.get("upload_id")
        item["periode_label"] = periode_map.get(uid, "Input Manual") if uid else "Input Manual"

    if not data:
        main_status = "—"
    else:
        all_terbayar = all((item.get("status") or "").upper() == "TERBAYAR" for item in data)
        main_status = "Sukses" if all_terbayar else "Belum Lunas"

    conn = get_db()
    pd_filter = "kategori = 'pengembalian_dana'"
    statuses = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT status FROM tagihan WHERE status IS NOT NULL AND {pd_filter}"
        ).fetchall()
    ]
    rekenings = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT rekening FROM tagihan WHERE rekening IS NOT NULL AND {pd_filter}"
        ).fetchall()
    ]
    conn.close()

    total_filtered = sum(d["jumlah"] for d in data)
    avg_gaji = int(total_filtered / len(data)) if data else 0
    max_item = max(data, key=lambda x: x["jumlah"]) if data else None
    min_item = min(data, key=lambda x: x["jumlah"]) if data else None
    total_all_batches = sum(lap.get("total_gaji") or 0 for lap in laporans)
    total_all_mitra = sum(lap.get("jumlah_penerima") or lap.get("record_count") or 0 for lap in laporans)

    return render_template(request, "pengembalian_dana.html", {
        "user": user,
        "items": data,
        "laporans": laporans,
        "laporan": laporan,
        "view_all": view_all,
        "selected_upload_id": upload_id if upload_id else 0,
        "total_filtered": total_filtered,
        "pengembalian_dana_jumlah": PENGEMBALIAN_DANA_JUMLAH,
        "pengembalian_dana_per_hari": PENGEMBALIAN_DANA_PER_HARI,
        "pengembalian_dana_hari": PENGEMBALIAN_DANA_HARI,
        "total_all_batches": total_all_batches,
        "total_all_mitra": total_all_mitra,
        "avg_gaji": avg_gaji,
        "max_item": max_item,
        "min_item": min_item,
        "main_status": main_status,
        "format_rupiah": format_rupiah,
        "format_tanggal": format_tanggal_display,
        "filters": {"search": search, "status": status, "rekening": rekening, "tanggal": tanggal, "periode": upload_id},
        "periode_options": periode_options,
        "status_options": sorted([s for s in statuses if s]),
        "rekening_options": sorted([r for r in rekenings if r]),
        "message": message,
        "success": success,
    })


@app.post("/pengembalian-dana")
async def create_pengembalian_dana(
    request: Request,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form("DIAJUKAN"),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        INSERT INTO tagihan (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, kategori, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pengembalian_dana', ?)
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or "DIAJUKAN",
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        user["id"],
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/pengembalian-dana", status_code=303)


@app.post("/pengembalian-dana/{item_id}/update")
async def update_pengembalian_dana(
    item_id: int,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form(""),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        UPDATE tagihan SET
            no = ?, pengajuan = ?, jumlah = ?, status = ?, rekening = ?,
            tanggal = ?, atas_nama = ?, nomor_rekening = ?, bank = ?,
            kategori = 'pengembalian_dana', updated_at = CURRENT_TIMESTAMP
        WHERE id = ? AND kategori = 'pengembalian_dana'
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or None,
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        item_id,
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/pengembalian-dana", status_code=303)


@app.post("/pengembalian-dana/{item_id}/delete")
async def delete_pengembalian_dana(item_id: int, user=Depends(require_login)):
    conn = get_db()
    row = conn.execute(
        "SELECT upload_id FROM tagihan WHERE id = ? AND kategori = 'pengembalian_dana'", (item_id,)
    ).fetchone()
    conn.execute("DELETE FROM tagihan WHERE id = ? AND kategori = 'pengembalian_dana'", (item_id,))
    if row and row[0]:
        recalc_pengembalian_dana_laporan(conn, row[0])
    conn.commit()
    conn.close()
    return RedirectResponse("/pengembalian-dana", status_code=303)


@app.post("/api/pengembalian-dana/bulk-delete")
async def api_pengembalian_dana_bulk_delete(request: Request, user=Depends(require_login)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        if not ids:
            return JSONResponse({"error": "no valid ids"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        upload_rows = conn.execute(
            f"""SELECT DISTINCT upload_id FROM tagihan
                WHERE id IN ({placeholders}) AND kategori = 'pengembalian_dana' AND upload_id IS NOT NULL""",
            ids,
        ).fetchall()
        conn.execute(
            f"DELETE FROM tagihan WHERE id IN ({placeholders}) AND kategori = 'pengembalian_dana'",
            ids,
        )
        deleted = conn.total_changes
        for row in upload_rows:
            recalc_pengembalian_dana_laporan(conn, row[0])
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "deleted": deleted})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/pengembalian-dana/bulk-update")
async def api_pengembalian_dana_bulk_update(request: Request, user=Depends(require_login)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        status = data.get("status")
        if not ids or status is None:
            return JSONResponse({"error": "no valid ids or status"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        conn.execute(
            f"""UPDATE tagihan SET status = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id IN ({placeholders}) AND kategori = 'pengembalian_dana'""",
            [status] + ids,
        )
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "updated": len(ids)})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/export/pengembalian-dana/csv")
async def export_pengembalian_dana_csv(user=Depends(require_login)):
    import csv
    from io import StringIO

    data = get_pengembalian_dana()
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["NO", "TANGGAL", "KETERANGAN", "NAMA REKENING", "NOMOR REKENING", "BANK", "JUMLAH", "STATUS"])
    for d in data:
        writer.writerow([
            d["no"] or "",
            d["tanggal"] or "",
            d["pengajuan"],
            d["atas_nama"] or "",
            d["nomor_rekening"] or "",
            d["bank"] or "",
            d["jumlah"],
            d["status"] or "",
        ])
    output.seek(0)
    filename = f"pengembalian_dana_sppg_{date.today().isoformat()}.csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/export/pengembalian-dana/xlsx")
async def export_pengembalian_dana_xlsx(user=Depends(require_login)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    data = get_pengembalian_dana()
    wb = Workbook()
    ws = wb.active
    ws.title = "Pengembalian Dana"

    header_fill = PatternFill(start_color="071E49", end_color="071E49", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws.merge_cells("B1:I1")
    ws["B1"] = "PENGEMBALIAN DANA / REFUND - SPPG WISMA HAJI MADIUN"
    ws["B1"].font = Font(bold=True, size=14, color="071E49")
    ws["B1"].alignment = Alignment(horizontal="center")

    headers = ["NO", "TANGGAL", "KETERANGAN", "NAMA REKENING", "NOMOR REKENING", "BANK", "JUMLAH", "STATUS"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for idx, d in enumerate(data, start=4):
        row_data = [
            d["no"] or "",
            d["tanggal"] or "",
            d["pengajuan"],
            d["atas_nama"] or "",
            d["nomor_rekening"] or "",
            d["bank"] or "",
            d["jumlah"],
            d["status"] or "",
        ]
        for col, val in enumerate(row_data, start=2):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border
            if col == 8:
                cell.number_format = "#,##0"
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [7, 14, 35, 22, 18, 14, 15, 14]
    for i, w in enumerate(widths, start=2):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Pengembalian_Dana_SPPG_{date.today().isoformat()}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/pengajuan-dana-mitra", response_class=HTMLResponse)
async def pengajuan_dana_mitra_page(
    request: Request,
    user=Depends(require_login),
    upload_id: int = 0,
    search: str = "",
    status: str = "",
    rekening: str = "",
    tanggal: str = "",
    message: str = "",
    success: bool = False,
):
    periode_options = get_pengajuan_dana_mitra_laporans(active_only=True)
    laporans = periode_options
    laporan = None
    active_upload_ids = {lap["upload_id"] for lap in periode_options}

    filters = {}
    if search:
        filters["search"] = search
    if status:
        filters["status"] = status
    if rekening:
        filters["rekening"] = rekening
    if tanggal:
        filters["tanggal"] = tanggal

    if upload_id and upload_id not in active_upload_ids:
        upload_id = 0

    view_all = upload_id == 0

    if upload_id and upload_id in active_upload_ids:
        laporan = next((lap for lap in periode_options if lap["upload_id"] == upload_id), None)
        filters["upload_id"] = upload_id

    data = get_pengajuan_dana_mitra(filters)

    periode_map = {lap["upload_id"]: lap.get("periode") or lap.get("filename") for lap in laporans}
    conn = get_db()
    user_id = user.get("user_id")
    for item in data:
        uid = item.get("upload_id")
        item["periode_label"] = periode_map.get(uid, "Input Manual") if uid else "Input Manual"
        if is_admin(user):
            item["can_delete"] = True
        else:
            item["can_delete"] = _member_owns_pdm_item(conn, item["id"], user_id)

    if not data:
        main_status = "—"
    else:
        all_terbayar = all((item.get("status") or "").upper() == "TERBAYAR" for item in data)
        main_status = "Sukses" if all_terbayar else "Belum Lunas"

    pdm_filter = "kategori = 'pengajuan_dana_mitra'"
    statuses = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT status FROM tagihan WHERE status IS NOT NULL AND {pdm_filter}"
        ).fetchall()
    ]
    rekenings = [
        r[0] for r in conn.execute(
            f"SELECT DISTINCT rekening FROM tagihan WHERE rekening IS NOT NULL AND {pdm_filter}"
        ).fetchall()
    ]
    conn.close()

    total_filtered = sum(d["jumlah"] for d in data)
    avg_gaji = int(total_filtered / len(data)) if data else 0
    max_item = max(data, key=lambda x: x["jumlah"]) if data else None
    min_item = min(data, key=lambda x: x["jumlah"]) if data else None
    total_all_batches = sum(lap.get("total_gaji") or 0 for lap in laporans)
    total_all_mitra = sum(lap.get("jumlah_penerima") or lap.get("record_count") or 0 for lap in laporans)

    return render_template(request, "pengajuan_dana_mitra.html", {
        "user": user,
        "items": data,
        "laporans": laporans,
        "laporan": laporan,
        "view_all": view_all,
        "selected_upload_id": upload_id if upload_id else 0,
        "total_filtered": total_filtered,
        "total_all_batches": total_all_batches,
        "total_all_mitra": total_all_mitra,
        "avg_gaji": avg_gaji,
        "max_item": max_item,
        "min_item": min_item,
        "main_status": main_status,
        "format_rupiah": format_rupiah,
        "format_tanggal": format_tanggal_display,
        "format_tanggal_pengajuan": format_tanggal_pengajuan,
        "filters": {"search": search, "status": status, "rekening": rekening, "tanggal": tanggal, "periode": upload_id},
        "periode_options": periode_options,
        "status_options": sorted([s for s in statuses if s]),
        "rekening_options": sorted([r for r in rekenings if r]),
        "tanggal_pengajuan_options": get_pengajuan_dana_mitra_tanggal_options(),
        "can_delete_upload": can_member_upload(user),
        "message": message,
        "success": success,
    })


@app.post("/pengajuan-dana-mitra/upload")
async def pengajuan_dana_mitra_upload(
    request: Request,
    user=Depends(require_login),
    file: UploadFile = File(...),
):
    """Upload PDF Formulir Pengajuan Dana Mitra — ekstrak otomatis baris rincian."""
    if is_viewer(user):
        return redirect_with_flash(
            request,
            "/pengajuan-dana-mitra",
            "Akses ditolak. Akun Anda hanya dapat melihat data.",
        )

    ext = (file.filename or "").lower().rsplit(".", 1)[-1] if "." in (file.filename or "") else ""
    if ext != "pdf":
        return redirect_with_flash(
            request,
            "/pengajuan-dana-mitra",
            "Hanya file PDF Formulir Pengajuan Dana Mitra yang didukung",
        )

    safe_name = f"{date.today().isoformat()}_{file.filename.replace(' ', '_')}"
    file_path = os.path.join(UPLOAD_DIR, safe_name)
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    conn = get_db()
    conn.execute(
        "INSERT INTO uploads (filename, created_by) VALUES (?, ?)",
        (file.filename, user["id"]),
    )
    upload_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    try:
        parsed = parse_pengajuan_dana_mitra_pdf(file_path, file.filename)
    except Exception as e:
        conn.close()
        return redirect_with_flash(request, "/pengajuan-dana-mitra", f"Error parsing file: {str(e)}")

    item_list = parsed.get("items", []) if isinstance(parsed, dict) else []
    pdm_meta = parsed.get("meta", {}) if isinstance(parsed, dict) else {}

    if not item_list:
        conn.close()
        return redirect_with_flash(
            request,
            "/pengajuan-dana-mitra",
            "Tidak ada data yang bisa dibaca dari PDF. Pastikan format Formulir Pengajuan Dana Mitra.",
        )

    inserted = 0
    for item in item_list:
        pengajuan = item.get("pengajuan") or ""
        if not pengajuan:
            continue
        jumlah = int(item.get("jumlah") or 0)
        if jumlah <= 0:
            continue

        dup = conn.execute("""
            SELECT 1 FROM tagihan
            WHERE upload_id = ? AND COALESCE(no, '') = COALESCE(?, '') AND pengajuan = ? AND jumlah = ?
            LIMIT 1
        """, (upload_id, item.get("no"), pengajuan, jumlah)).fetchone()
        if dup:
            continue

        conn.execute("""
            INSERT INTO tagihan
            (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, kategori, upload_id, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pengajuan_dana_mitra', ?, ?)
        """, (
            item.get("no"),
            pengajuan,
            jumlah,
            item.get("status") or "DIAJUKAN",
            item.get("rekening") or "PENGAJUAN DANA MITRA",
            pdm_meta.get("tanggal_pengajuan") or item.get("tanggal"),
            item.get("atas_nama"),
            item.get("nomor_rekening"),
            item.get("bank"),
            upload_id,
            user["id"],
        ))
        inserted += 1

    total_gaji = sum(int(it.get("jumlah") or 0) for it in item_list)
    first_bank = item_list[0].get("bank") if item_list else "MANDIRI"
    first_rek = item_list[0].get("nomor_rekening") if item_list else None

    conn.execute("""
        INSERT OR REPLACE INTO pengajuan_dana_mitra_laporan
        (upload_id, tanggal_pembayaran, rekening_sumber, jumlah_penerima, total_gaji,
         periode, bank, kota, filename)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        upload_id,
        pdm_meta.get("tanggal_pengajuan"),
        first_rek,
        inserted,
        total_gaji,
        pdm_meta.get("periode") or file.filename,
        first_bank,
        pdm_meta.get("divisi"),
        file.filename,
    ))
    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (inserted, upload_id))
    conn.commit()
    conn.close()

    pemohon = pdm_meta.get("pemohon") or ""
    msg = f"Berhasil mengekstrak {inserted} baris dari PDF"
    if pemohon:
        msg += f" — Pemohon: {pemohon}"
    return redirect_with_flash(request, f"/pengajuan-dana-mitra?upload_id={upload_id}", msg, success=True)


@app.post("/pengajuan-dana-mitra")
async def create_pengajuan_dana_mitra(
    request: Request,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form("DIAJUKAN"),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        INSERT INTO tagihan (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, kategori, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pengajuan_dana_mitra', ?)
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or "DIAJUKAN",
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        user["id"],
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/pengajuan-dana-mitra", status_code=303)


@app.post("/pengajuan-dana-mitra/{item_id}/update")
async def update_pengajuan_dana_mitra(
    item_id: int,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form(""),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        UPDATE tagihan SET
            no = ?, pengajuan = ?, jumlah = ?, status = ?, rekening = ?,
            tanggal = ?, atas_nama = ?, nomor_rekening = ?, bank = ?,
            kategori = 'pengajuan_dana_mitra', updated_at = CURRENT_TIMESTAMP
        WHERE id = ? AND kategori = 'pengajuan_dana_mitra'
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or None,
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        item_id,
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/pengajuan-dana-mitra", status_code=303)


@app.post("/pengajuan-dana-mitra/{item_id}/delete")
async def delete_pengajuan_dana_mitra(item_id: int, request: Request, user=Depends(require_login)):
    conn = get_db()
    if not is_admin(user) and not _member_owns_pdm_item(conn, item_id, user["id"]):
        conn.close()
        return redirect_with_flash(
            request,
            "/pengajuan-dana-mitra",
            "Akses ditolak. Hanya data upload Anda yang dapat dihapus.",
        )
    row = conn.execute(
        "SELECT upload_id FROM tagihan WHERE id = ? AND kategori = 'pengajuan_dana_mitra'", (item_id,)
    ).fetchone()
    conn.execute("DELETE FROM tagihan WHERE id = ? AND kategori = 'pengajuan_dana_mitra'", (item_id,))
    if row and row[0]:
        recalc_pengajuan_dana_mitra_laporan(conn, row[0])
    conn.commit()
    conn.close()
    return RedirectResponse("/pengajuan-dana-mitra", status_code=303)


@app.post("/api/pengajuan-dana-mitra/bulk-delete")
async def api_pengajuan_dana_mitra_bulk_delete(request: Request, user=Depends(require_login)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        if not ids:
            return JSONResponse({"error": "no valid ids"}, status_code=400)
        conn = get_db()
        if not is_admin(user):
            ids = [i for i in ids if _member_owns_pdm_item(conn, i, user["id"])]
            if not ids:
                conn.close()
                return JSONResponse(
                    {"error": "Akses ditolak. Hanya data upload Anda yang dapat dihapus."},
                    status_code=403,
                )
        placeholders = ",".join("?" * len(ids))
        upload_rows = conn.execute(
            f"""SELECT DISTINCT upload_id FROM tagihan
                WHERE id IN ({placeholders}) AND kategori = 'pengajuan_dana_mitra' AND upload_id IS NOT NULL""",
            ids,
        ).fetchall()
        conn.execute(
            f"DELETE FROM tagihan WHERE id IN ({placeholders}) AND kategori = 'pengajuan_dana_mitra'",
            ids,
        )
        deleted = conn.total_changes
        for row in upload_rows:
            recalc_pengajuan_dana_mitra_laporan(conn, row[0])
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "deleted": deleted})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/pengajuan-dana-mitra/bulk-update")
async def api_pengajuan_dana_mitra_bulk_update(request: Request, user=Depends(require_login)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        status = data.get("status")
        if not ids or status is None:
            return JSONResponse({"error": "no valid ids or status"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        conn.execute(
            f"""UPDATE tagihan SET status = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id IN ({placeholders}) AND kategori = 'pengajuan_dana_mitra'""",
            [status] + ids,
        )
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "updated": len(ids)})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/export/pengajuan-dana-mitra/csv")
async def export_pengajuan_dana_mitra_csv(user=Depends(require_login)):
    import csv
    from io import StringIO

    data = get_pengajuan_dana_mitra()
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["NO", "TANGGAL", "KETERANGAN", "NAMA REKENING", "NOMOR REKENING", "BANK", "JUMLAH", "STATUS"])
    for d in data:
        writer.writerow([
            d["no"] or "",
            d["tanggal"] or "",
            d["pengajuan"],
            d["atas_nama"] or "",
            d["nomor_rekening"] or "",
            d["bank"] or "",
            d["jumlah"],
            d["status"] or "",
        ])
    output.seek(0)
    filename = f"pengajuan_dana_mitra_sppg_{date.today().isoformat()}.csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/export/pengajuan-dana-mitra/xlsx")
async def export_pengajuan_dana_mitra_xlsx(user=Depends(require_login)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    data = get_pengajuan_dana_mitra()
    wb = Workbook()
    ws = wb.active
    ws.title = "Pengajuan Dana Mitra"

    header_fill = PatternFill(start_color="071E49", end_color="071E49", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws.merge_cells("B1:I1")
    ws["B1"] = "PENGAJUAN DANA MITRA - SPPG WISMA HAJI MADIUN"
    ws["B1"].font = Font(bold=True, size=14, color="071E49")
    ws["B1"].alignment = Alignment(horizontal="center")

    headers = ["NO", "TANGGAL", "KETERANGAN", "NAMA REKENING", "NOMOR REKENING", "BANK", "JUMLAH", "STATUS"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for idx, d in enumerate(data, start=4):
        row_data = [
            d["no"] or "",
            d["tanggal"] or "",
            d["pengajuan"],
            d["atas_nama"] or "",
            d["nomor_rekening"] or "",
            d["bank"] or "",
            d["jumlah"],
            d["status"] or "",
        ]
        for col, val in enumerate(row_data, start=2):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border
            if col == 8:
                cell.number_format = "#,##0"
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [7, 14, 35, 22, 18, 14, 15, 14]
    for i, w in enumerate(widths, start=2):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Pengajuan_Dana_Mitra_SPPG_{date.today().isoformat()}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/petty-cash", response_class=HTMLResponse)
async def petty_cash_page(
    request: Request,
    user=Depends(require_login),
    upload_id: int = 0,
    search: str = "",
    tanggal: str = "",
    jenis: str = "",
    message: str = "",
    success: bool = False,
):
    laporans = get_petty_cash_laporans()
    laporan = None
    all_items = []

    if upload_id:
        conn = get_db()
        row = conn.execute(
            "SELECT * FROM petty_cash_laporan WHERE upload_id = ?", (upload_id,)
        ).fetchone()
        conn.close()
        if row:
            laporan = dict(row)
            all_items = get_petty_cash_items(upload_id)
    elif laporans:
        laporan = laporans[0]
        all_items = get_petty_cash_items(laporan["upload_id"])

    tanggal_options = sorted(
        {i.get("tanggal") for i in all_items if i.get("tanggal")},
        reverse=True,
    )
    tanggal_min = tanggal_options[-1] if tanggal_options else ""
    tanggal_max = tanggal_options[0] if tanggal_options else ""
    has_filters = bool(search.strip() or tanggal or jenis.strip())
    items = filter_petty_cash_items(all_items, search, tanggal, jenis)

    total_pengeluaran = sum_petty_pengeluaran(items)
    total_penerimaan = sum(int(i.get("debit") or 0) for i in items)

    total_pengeluaran_all = sum_petty_pengeluaran(all_items)
    total_penerimaan_all = sum(int(i.get("debit") or 0) for i in all_items)
    is_reimbursement = laporan and (laporan.get("report_type") or "") == "reimbursement"
    if laporan and not has_filters and not is_reimbursement:
        if laporan.get("total_kredit"):
            total_pengeluaran = laporan.get("total_kredit")
        if laporan.get("total_debit"):
            total_penerimaan = laporan.get("total_debit")

    nota_count = sum(1 for i in all_items if i.get("nota_path"))

    return render_template(request, "petty_cash.html", {
        "user": user,
        "laporans": laporans,
        "laporan": laporan,
        "items": items,
        "all_items_count": len(all_items),
        "total_pengeluaran": total_pengeluaran,
        "total_penerimaan": total_penerimaan,
        "total_pengeluaran_all": total_pengeluaran_all or (laporan.get("total_kredit") if laporan else 0),
        "total_penerimaan_all": total_penerimaan_all or (laporan.get("total_debit") if laporan else 0),
        "format_rupiah": format_rupiah,
        "format_tanggal": format_tanggal_display,
        "selected_upload_id": laporan["upload_id"] if laporan else 0,
        "filters": {"search": search, "tanggal": tanggal, "jenis": jenis},
        "tanggal_options": tanggal_options,
        "tanggal_min": tanggal_min,
        "tanggal_max": tanggal_max,
        "has_filters": has_filters,
        "is_reimbursement": is_reimbursement,
        "nota_count": nota_count,
        "message": message,
        "success": success,
    })

@app.post("/petty-cash/upload")
async def petty_cash_upload(
    request: Request,
    user=Depends(require_login),
    file: UploadFile = File(...),
    periode: str = Form(""),
):
    """Upload PDF petty cash — terpisah dari modul Tagihan."""
    if is_viewer(user):
        return redirect_with_flash(
            request,
            "/petty-cash",
            "Akses ditolak. Akun Anda hanya dapat melihat data.",
        )
    if not is_admin(user) and not can_member_upload(user):
        return redirect_with_flash(
            request,
            "/petty-cash",
            "Akses ditolak. Anda tidak dapat mengunggah file.",
        )
    safe_name = f"{date.today().isoformat()}_{file.filename.replace(' ', '_')}"
    file_path = os.path.join(UPLOAD_DIR, safe_name)
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    conn = get_db()
    conn.execute("""
        INSERT INTO uploads (filename, periode, created_by)
        VALUES (?, ?, ?)
    """, (file.filename, periode or None, user["id"]))
    upload_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    try:
        parsed = parse_petty_cash_pdf(file_path, upload_id, file.filename)
    except Exception as e:
        conn.close()
        return redirect_with_flash(request, "/petty-cash", f"Error parsing file: {str(e)}")

    item_list = parsed.get("items", []) if isinstance(parsed, dict) else []
    petty_meta = parsed.get("meta", {}) if isinstance(parsed, dict) else {}

    if not item_list:
        conn.close()
        return redirect_with_flash(request, "/petty-cash", "Tidak ada data yang bisa dibaca dari file")

    inserted = 0
    for item in item_list:
        pengajuan = item.get("pengajuan") or item.get("deskripsi") or ""
        if not pengajuan:
            continue
        try:
            debit_val = int(item.get("debit") or 0)
            kredit_val = int(item.get("kredit") or 0)
            jumlah = int(item.get("jumlah") or 0)
        except Exception:
            debit_val = kredit_val = jumlah = 0
        if not jumlah:
            jumlah = kredit_val if kredit_val > 0 else debit_val
        if jumlah <= 0 and debit_val <= 0 and kredit_val <= 0:
            continue

        no = str(item.get("no") or "").strip() or None
        tanggal = str(item.get("tanggal") or "").strip() or None
        nota_path = item.get("nota_path")
        saldo_akhir = item.get("saldo_akhir")
        tipe_transaksi = item.get("tipe_transaksi") or item.get("tipe")

        dup = conn.execute("""
            SELECT 1 FROM petty_cash_items
            WHERE upload_id = ? AND COALESCE(no,'') = COALESCE(?, '') AND pengajuan = ?
            LIMIT 1
        """, (upload_id, no, pengajuan)).fetchone()
        if dup:
            continue

        conn.execute("""
            INSERT INTO petty_cash_items
            (upload_id, no, pengajuan, jumlah, debit, kredit, saldo_akhir, tipe_transaksi, tanggal, nota_path)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            upload_id, no, pengajuan, jumlah, debit_val, kredit_val,
            saldo_akhir, tipe_transaksi, tanggal, nota_path,
        ))
        inserted += 1

    if petty_meta:
        conn.execute("""
            INSERT OR REPLACE INTO petty_cash_laporan
            (upload_id, nama_karyawan, divisi, saldo_awal, sisa_dana, total_digantikan,
             payment_info, bank, nomor_rekening, atas_nama, yang_menyetujui,
             tanggal_ttd_pemohon, tanggal_ttd_menyetujui, periode, filename,
             report_type, total_debit, total_kredit, saldo_akhir)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            upload_id,
            petty_meta.get("nama_karyawan"),
            petty_meta.get("divisi"),
            petty_meta.get("saldo_awal") or 0,
            petty_meta.get("sisa_dana") or petty_meta.get("saldo_akhir") or 0,
            petty_meta.get("total_digantikan") or petty_meta.get("total_kredit") or 0,
            petty_meta.get("payment_info"),
            petty_meta.get("bank"),
            petty_meta.get("nomor_rekening"),
            petty_meta.get("atas_nama"),
            petty_meta.get("yang_menyetujui"),
            petty_meta.get("tanggal_ttd_pemohon"),
            petty_meta.get("tanggal_ttd_menyetujui"),
            petty_meta.get("periode") or periode,
            file.filename,
            petty_meta.get("report_type") or "reimbursement",
            petty_meta.get("total_debit") or 0,
            petty_meta.get("total_kredit") or 0,
            petty_meta.get("saldo_akhir") or petty_meta.get("sisa_dana") or 0,
        ))

    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (inserted, upload_id))
    conn.commit()
    conn.close()

    msg = f"Berhasil mengunggah {inserted} baris data Petty Cash"
    return redirect_with_flash(request, f"/petty-cash?upload_id={upload_id}", msg, success=True)

@app.post("/tagihan")
async def create_tagihan(
    request: Request,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form("DIAJUKAN"),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
    kategori: str = Form("tagihan"),
):
    conn = get_db()
    conn.execute("""
        INSERT INTO tagihan (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, kategori, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or "DIAJUKAN",
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        kategori or "tagihan",
        user["id"]
    ))
    conn.commit()
    conn.close()

    # If AJAX request, return JSON
    if request.headers.get("hx-request"):
        return JSONResponse({"success": True})

    return RedirectResponse("/tagihan", status_code=303)

@app.post("/tagihan/{tagihan_id}/update")
async def update_tagihan(
    tagihan_id: int,
    request: Request,
    user=Depends(require_login),
    no: str = Form(""),
    pengajuan: str = Form(...),
    jumlah: int = Form(...),
    status: str = Form(""),
    rekening: str = Form(""),
    tanggal: str = Form(""),
    atas_nama: str = Form(""),
    nomor_rekening: str = Form(""),
    bank: str = Form(""),
    kategori: str = Form(""),
):
    conn = get_db()
    conn.execute("""
        UPDATE tagihan SET
            no = ?,
            pengajuan = ?,
            jumlah = ?,
            status = ?,
            rekening = ?,
            tanggal = ?,
            atas_nama = ?,
            nomor_rekening = ?,
            bank = ?,
            kategori = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (
        no.strip() or None,
        pengajuan.strip(),
        jumlah,
        status or None,
        rekening.strip() or None,
        tanggal or None,
        atas_nama.strip() or None,
        nomor_rekening.strip() or None,
        bank.strip() or None,
        kategori or None,
        tagihan_id
    ))
    conn.commit()
    conn.close()
    return RedirectResponse("/tagihan", status_code=303)

@app.post("/tagihan/{tagihan_id}/delete")
async def delete_tagihan(tagihan_id: int, user=Depends(require_login)):
    conn = get_db()
    conn.execute("DELETE FROM tagihan WHERE id = ?", (tagihan_id,))
    conn.commit()
    conn.close()
    return RedirectResponse("/tagihan", status_code=303)

# Bulk actions
def recalc_petty_cash_laporan(conn, upload_id: int):
    """Hitung ulang total laporan setelah hapus transaksi."""
    rows = conn.execute(
        "SELECT debit, kredit, saldo_akhir FROM petty_cash_items WHERE upload_id = ? ORDER BY id",
        (upload_id,),
    ).fetchall()
    total_debit = sum(r[0] or 0 for r in rows)
    total_kredit = sum(r[1] or 0 for r in rows)
    saldo_akhir = rows[-1][2] if rows else 0
    conn.execute("""
        UPDATE petty_cash_laporan SET
            total_debit = ?, total_kredit = ?, total_digantikan = ?,
            sisa_dana = ?, saldo_akhir = ?
        WHERE upload_id = ?
    """, (total_debit, total_kredit, total_kredit, saldo_akhir, saldo_akhir, upload_id))
    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (len(rows), upload_id))

ALLOWED_NOTA_EXT = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".pdf"}
TAGIHAN_ATTACHMENT_COLS = {
    "pict": "pict_path",
    "nota": "nota_path",
    "bukti": "bukti_path",
}


def _delete_upload_file(rel_path: str):
    if not rel_path:
        return
    full = os.path.join(UPLOAD_DIR, rel_path)
    if os.path.isfile(full):
        try:
            os.remove(full)
        except OSError:
            pass


def _delete_nota_file(nota_path: str):
    _delete_upload_file(nota_path)


@app.post("/api/petty-cash/{item_id}/nota")
async def api_petty_cash_upload_nota(
    item_id: int,
    user=Depends(require_login),
    file: UploadFile = File(...),
):
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALLOWED_NOTA_EXT:
        return JSONResponse({"error": "Format file tidak didukung"}, status_code=400)

    conn = get_db()
    row = conn.execute(
        "SELECT id, upload_id, nota_path FROM petty_cash_items WHERE id = ?", (item_id,)
    ).fetchone()
    if not row:
        conn.close()
        return JSONResponse({"error": "Transaksi tidak ditemukan"}, status_code=404)

    nota_dir = os.path.join(UPLOAD_DIR, "nota", str(row["upload_id"]))
    os.makedirs(nota_dir, exist_ok=True)
    safe_ext = ".jpg" if ext in (".jpg", ".jpeg") else ext
    fname = f"nota_item_{item_id}{safe_ext}"
    rel_path = f"nota/{row['upload_id']}/{fname}"
    full_path = os.path.join(UPLOAD_DIR, rel_path)

    if row["nota_path"] and row["nota_path"] != rel_path:
        _delete_nota_file(row["nota_path"])

    content = await file.read()
    with open(full_path, "wb") as f:
        f.write(content)

    conn.execute(
        "UPDATE petty_cash_items SET nota_path = ? WHERE id = ?", (rel_path, item_id)
    )
    conn.commit()
    conn.close()
    return JSONResponse({"success": True, "nota_path": rel_path})


@app.patch("/api/petty-cash/{item_id}/ket")
async def api_petty_cash_update_ket(
    item_id: int,
    request: Request,
    user=Depends(require_login),
):
    try:
        data = await request.json()
        ket = (data.get("ket") or "").strip() or None
        conn = get_db()
        row = conn.execute(
            "SELECT id FROM petty_cash_items WHERE id = ?", (item_id,)
        ).fetchone()
        if not row:
            conn.close()
            return JSONResponse({"error": "Transaksi tidak ditemukan"}, status_code=404)
        conn.execute(
            "UPDATE petty_cash_items SET ket = ? WHERE id = ?", (ket, item_id)
        )
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "ket": ket or ""})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.delete("/api/petty-cash/{item_id}/nota")
async def api_petty_cash_delete_nota(item_id: int, user=Depends(require_login)):
    conn = get_db()
    row = conn.execute(
        "SELECT nota_path FROM petty_cash_items WHERE id = ?", (item_id,)
    ).fetchone()
    if not row:
        conn.close()
        return JSONResponse({"error": "Transaksi tidak ditemukan"}, status_code=404)

    if row["nota_path"]:
        _delete_nota_file(row["nota_path"])

    conn.execute("UPDATE petty_cash_items SET nota_path = NULL WHERE id = ?", (item_id,))
    conn.commit()
    conn.close()
    return JSONResponse({"success": True})


@app.post("/api/tagihan/{item_id}/{field}")
async def api_tagihan_upload_attachment(
    item_id: int,
    field: str,
    user=Depends(require_login),
    file: UploadFile = File(...),
):
    field = (field or "").lower()
    if field not in TAGIHAN_ATTACHMENT_FIELDS:
        return JSONResponse({"error": "Field lampiran tidak valid"}, status_code=400)

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALLOWED_NOTA_EXT:
        return JSONResponse({"error": "Format file tidak didukung"}, status_code=400)

    col = TAGIHAN_ATTACHMENT_COLS[field]
    conn = get_db()
    row = conn.execute(
        f"SELECT id, upload_id, kategori, {col} AS file_path FROM tagihan WHERE id = ?",
        (item_id,),
    ).fetchone()
    if not row:
        conn.close()
        return JSONResponse({"error": "Transaksi tidak ditemukan"}, status_code=404)
    if (row["kategori"] or "") not in TAGIHAN_ATTACHMENT_KATEGORI:
        conn.close()
        return JSONResponse({"error": "Lampiran tidak didukung untuk kategori ini"}, status_code=400)

    upload_key = row["upload_id"] or item_id
    attach_dir = os.path.join(UPLOAD_DIR, "lampiran", str(upload_key))
    os.makedirs(attach_dir, exist_ok=True)
    safe_ext = ".jpg" if ext in (".jpg", ".jpeg") else ext
    fname = f"{field}_item_{item_id}{safe_ext}"
    rel_path = f"lampiran/{upload_key}/{fname}"
    full_path = os.path.join(UPLOAD_DIR, rel_path)

    if row["file_path"] and row["file_path"] != rel_path:
        _delete_upload_file(row["file_path"])

    content = await file.read()
    with open(full_path, "wb") as f:
        f.write(content)

    conn.execute(f"UPDATE tagihan SET {col} = ? WHERE id = ?", (rel_path, item_id))
    conn.commit()
    conn.close()
    return JSONResponse({"success": True, "path": rel_path, "field": field})


@app.delete("/api/tagihan/{item_id}/{field}")
async def api_tagihan_delete_attachment(item_id: int, field: str, user=Depends(require_login)):
    field = (field or "").lower()
    if field not in TAGIHAN_ATTACHMENT_FIELDS:
        return JSONResponse({"error": "Field lampiran tidak valid"}, status_code=400)

    col = TAGIHAN_ATTACHMENT_COLS[field]
    conn = get_db()
    row = conn.execute(
        f"SELECT id, kategori, {col} AS file_path FROM tagihan WHERE id = ?",
        (item_id,),
    ).fetchone()
    if not row:
        conn.close()
        return JSONResponse({"error": "Transaksi tidak ditemukan"}, status_code=404)
    if (row["kategori"] or "") not in TAGIHAN_ATTACHMENT_KATEGORI:
        conn.close()
        return JSONResponse({"error": "Lampiran tidak didukung untuk kategori ini"}, status_code=400)

    if row["file_path"]:
        _delete_upload_file(row["file_path"])

    conn.execute(f"UPDATE tagihan SET {col} = NULL WHERE id = ?", (item_id,))
    conn.commit()
    conn.close()
    return JSONResponse({"success": True, "field": field})


@app.post("/api/petty-cash/bulk-delete")
async def api_petty_cash_bulk_delete(request: Request, user=Depends(require_login)):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        if not ids:
            return JSONResponse({"error": "no valid ids"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        upload_rows = conn.execute(
            f"SELECT DISTINCT upload_id FROM petty_cash_items WHERE id IN ({placeholders})",
            ids,
        ).fetchall()
        conn.execute(f"DELETE FROM petty_cash_items WHERE id IN ({placeholders})", ids)
        for row in upload_rows:
            recalc_petty_cash_laporan(conn, row[0])
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "deleted": len(ids)})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.post("/api/tagihan/bulk-delete")
async def api_bulk_delete(request: Request):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        if not ids:
            return JSONResponse({"error": "no valid ids"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        conn.execute(f"DELETE FROM tagihan WHERE id IN ({placeholders})", ids)
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "deleted": len(ids)})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.post("/api/tagihan/bulk-update")
async def api_bulk_update(request: Request):
    try:
        data = await request.json()
        raw_ids = data.get("ids", [])
        ids = [int(i) for i in raw_ids if str(i).isdigit()]
        pos = data.get("pos")
        status = data.get("status")
        if not ids:
            return JSONResponse({"error": "no valid ids"}, status_code=400)
        conn = get_db()
        placeholders = ",".join("?" * len(ids))
        if pos is not None:
            conn.execute(f"UPDATE tagihan SET pos = ?, updated_at = CURRENT_TIMESTAMP WHERE id IN ({placeholders})", [pos] + ids)
        if status is not None:
            print(f"[DEBUG BULK] Updating status to {status} for ids {ids}")
            conn.execute(f"UPDATE tagihan SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id IN ({placeholders})", [status] + ids)
            print(f"[DEBUG BULK] Rows affected: {conn.total_changes}")
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "updated": len(ids)})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

# ===================== API for dynamic UI =====================

@app.get("/api/tagihan")
async def api_tagihan(user=Depends(require_login)):
    data = get_all_tagihan()
    return [{"id": d["id"], **d, "jumlah_fmt": format_rupiah(d["jumlah"])} for d in data]

@app.get("/api/summary")
async def api_summary(user=Depends(require_login)):
    return get_summary()

# ===================== Export =====================

def _build_tagihan_pdf(rows: List[Dict]) -> bytes:
    from fpdf import FPDF

    data = [enrich_tagihan_item(r) for r in rows]
    total_jumlah = sum(r["jumlah"] for r in data)
    total_charges = sum(r["charges"] for r in data)
    total_grand = total_jumlah + total_charges

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_margins(10, 10, 10)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(7, 30, 73)
    pdf.cell(0, 8, "LAPORAN TAGIHAN - SPPG WISMA HAJI MADIUN", ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 6, f"Dicetak: {date.today().strftime('%d %B %Y')}  |  {len(data)} entri", ln=True, align="C")
    pdf.ln(4)

    col_widths = [20, 28, 52, 22, 18, 22, 16, 14, 22, 22, 18]
    headers = ["NO", "PEMASOK", "KETERANGAN", "JUMLAH", "CHARGES", "TOTAL", "STATUS", "BANK", "NO. REK", "ATAS NAMA", "TANGGAL"]
    row_h = 7

    pdf.set_font("Helvetica", "B", 7)
    pdf.set_fill_color(7, 30, 73)
    pdf.set_text_color(255, 255, 255)
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], row_h, header, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(30, 30, 30)
    fill = False
    for item in data:
        if pdf.get_y() > 185:
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 7)
            pdf.set_fill_color(7, 30, 73)
            pdf.set_text_color(255, 255, 255)
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], row_h, header, border=1, align="C", fill=True)
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)
            pdf.set_text_color(30, 30, 30)

        keterangan = (item.get("pengajuan") or "")[:45]
        cells = [
            str(item.get("no") or "")[:18],
            str(item.get("pos") or "")[:22],
            keterangan,
            format_rupiah(item.get("jumlah") or 0).replace("Rp", "Rp "),
            format_rupiah(item.get("charges") or 0).replace("Rp", "Rp ") if item.get("charges") else "-",
            format_rupiah(item.get("total") or 0).replace("Rp", "Rp "),
            str(item.get("status") or "")[:10],
            str(item.get("bank") or "")[:12],
            str(item.get("nomor_rekening") or "")[:18],
            str(item.get("atas_nama") or "")[:20],
            str(item.get("tanggal") or "")[:10],
        ]
        aligns = ["L", "L", "L", "R", "R", "R", "C", "L", "L", "L", "C"]
        if fill:
            pdf.set_fill_color(248, 250, 252)
        else:
            pdf.set_fill_color(255, 255, 255)
        for i, (text, align) in enumerate(zip(cells, aligns)):
            pdf.cell(col_widths[i], row_h, text, border=1, align=align, fill=True)
        pdf.ln()
        fill = not fill

    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(236, 253, 245)
    pdf.set_text_color(7, 30, 73)
    pdf.cell(sum(col_widths[:3]), row_h, "TOTAL", border=1, align="R", fill=True)
    pdf.cell(col_widths[3], row_h, format_rupiah(total_jumlah).replace("Rp", "Rp "), border=1, align="R", fill=True)
    pdf.cell(col_widths[4], row_h, format_rupiah(total_charges).replace("Rp", "Rp "), border=1, align="R", fill=True)
    pdf.cell(col_widths[5], row_h, format_rupiah(total_grand).replace("Rp", "Rp "), border=1, align="R", fill=True)
    pdf.cell(sum(col_widths[6:]), row_h, "", border=1, fill=True)
    pdf.ln()

    return bytes(pdf.output())


@app.get("/export/pdf")
async def export_pdf(user=Depends(require_login)):
    data = get_all_tagihan()
    content = _build_tagihan_pdf(data)
    filename = f"Laporan_Tagihan_SPPG_{date.today().isoformat()}.pdf"
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/export/csv")
async def export_csv(user=Depends(require_login)):
    import csv
    from io import StringIO

    data = get_all_tagihan()
    output = StringIO()
    writer = csv.writer(output)

    writer.writerow(["NO", "PEMASOK", "KETERANGAN", "JUMLAH", "CHARGES", "TOTAL", "STATUS", "BANK", "NOMOR REKENING", "ATAS NAMA REK.", "TANGGAL"])

    for raw in data:
        d = enrich_tagihan_item(raw)
        writer.writerow([
            d["no"] or "",
            d.get("pos") or "",
            d["pengajuan"],
            d["jumlah"],
            d["charges"],
            d["total"],
            d["status"] or "",
            d["bank"] or "",
            d["nomor_rekening"] or "",
            d["atas_nama"] or "",
            d["tanggal"] or "",
        ])

    output.seek(0)
    filename = f"tagihan_sppg_{date.today().isoformat()}.csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/export/xlsx")
async def export_xlsx(user=Depends(require_login)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    data = get_all_tagihan()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tagihan Mitra"

    # Header styling similar to original
    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Title row
    ws.merge_cells('B1:J1')
    ws['B1'] = "TAGIHAN MITRA - SPPG WISMA HAJI MADIUN"
    ws['B1'].font = Font(bold=True, size=14, color="0F766E")
    ws['B1'].alignment = Alignment(horizontal='center')

    # Headers (row 3 to match original structure)
    headers = ["NO", "PEMASOK", "KETERANGAN", "JUMLAH", "CHARGES", "TOTAL", "STATUS", "BANK", "NOMOR REKENING", "ATAS NAMA REK.", "TANGGAL"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data rows
    for idx, raw in enumerate(data, start=4):
        d = enrich_tagihan_item(raw)
        row_data = [
            d["no"] or "",
            d.get("pos") or "",
            d["pengajuan"],
            d["jumlah"],
            d["charges"],
            d["total"],
            d["status"] or "",
            d["bank"] or "",
            d["nomor_rekening"] or "",
            d["atas_nama"] or "",
            d["tanggal"] or "",
        ]
        for col, val in enumerate(row_data, start=2):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border
            if col in (5, 6, 7):  # JUMLAH, CHARGES, TOTAL
                cell.number_format = '#,##0'
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # Column widths (close to original)
    widths = [7, 22, 35, 15, 12, 15, 16, 14, 18, 22, 14]
    for i, w in enumerate(widths, start=2):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = 'A4'

    # Total row
    total_row = 4 + len(data)
    ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=4)
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal='right', vertical='center')
    for col in (5, 6, 7):
        total_cell = ws.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_row-1})")
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Tagihan_Mitra_SPPG_{date.today().isoformat()}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ===================== UPLOAD LAPORAN =====================

def _parse_ledger_amount(text: str) -> int:
    """Parse Accurate ledger amounts (e.g. 5.000.000 or -1.228.456)."""
    import re
    raw = text.strip().replace(".", "").replace(",", "")
    try:
        return int(raw)
    except ValueError:
        m = re.search(r"-?[\d.,]+", text)
        if m:
            try:
                return int(m.group(0).replace(".", "").replace(",", ""))
            except ValueError:
                return 0
        return 0


def _extract_buku_besar_lines(pdf_path: str) -> List[str]:
    """Extract clean text lines from Accurate Buku Besar PDF."""
    import pdfplumber
    skip_exact = {
        "SPPG WISMA HAJI", "Rincian Buku Besar", "Filter berdasarkan : Kode Perkiraan",
        "Tanggal Tipe Transaksi Keterangan Debit Kredit Saldo Akhir", "11010101",
    }
    lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for raw in (page.extract_text() or "").split("\n"):
                line = raw.strip()
                if not line:
                    continue
                if line.startswith("ACCURATE") or "Halaman" in line or "Tercetak" in line:
                    continue
                if line in skip_exact:
                    continue
                lines.append(line)
    return lines


def parse_buku_besar_petty_cash(pdf_path: str, upload_id: int, filename: str = "") -> Dict[str, Any]:
    """Parse Accurate 'Rincian Buku Besar' PDF untuk akun PETTY CASH (11010101)."""
    import re

    meta = {
        "report_type": "buku_besar",
        "nama_karyawan": "11010101 — PETTY CASH",
        "divisi": "SPPG Wisma Haji Madiun",
        "saldo_awal": 0,
        "sisa_dana": 0,
        "saldo_akhir": 0,
        "total_digantikan": 0,
        "total_debit": 0,
        "total_kredit": 0,
        "payment_info": "Sumber: Accurate Accounting — Rincian Buku Besar",
        "bank": None,
        "nomor_rekening": None,
        "atas_nama": None,
        "periode": None,
        "filename": filename,
    }

    all_lines = _extract_buku_besar_lines(pdf_path)
    for line in all_lines:
        if line.startswith("Dari ") and "s/d" in line:
            meta["periode"] = line.replace("Dari ", "").strip()
            break
    if not meta["periode"]:
        periode_m = re.search(r"(\d{1,2}\s+\w+\s+\d{4})", filename.replace("_", " "), re.I)
        if periode_m:
            meta["periode"] = periode_m.group(1)

    tx_re = re.compile(
        r"^(\d{1,2}\s+\w{3}\s+\d{4})\s+(.+?)\s+(-?[\d.,]+)\s+(-?[\d.,]+)\s+(-?[\d.,]+)$"
    )

    items = []
    pending = None

    def _commit_pending():
        nonlocal pending
        if not pending or "debit" not in pending:
            pending = None
            return
        pending["no"] = str(len(items) + 1)
        pending["status"] = "DIAJUKAN"
        pending["rekening"] = "PETTY CASH"
        pending["jumlah"] = pending["kredit"] if pending["kredit"] > 0 else pending["debit"]
        items.append(pending)
        pending = None

    for line in all_lines:
        if line.startswith("11010101 -"):
            continue
        if re.match(r"^Dari \d", line):
            continue

        if re.match(r"^\d{1,2}\s+\w{3}\s+\d{4}\s+Saldo per", line):
            nums = re.findall(r"-?[\d.,]+", line)
            if nums:
                meta["saldo_awal"] = _parse_ledger_amount(nums[-1])
            continue

        if re.match(r"^[\d.,]+\s+[\d.,]+$", line) and len(line.split()) == 2:
            parts = line.split()
            meta["total_debit"] = _parse_ledger_amount(parts[0])
            meta["total_kredit"] = _parse_ledger_amount(parts[1])
            continue

        m = tx_re.match(line)
        if m:
            _commit_pending()
            tanggal_raw = m.group(1)
            body = m.group(2).strip()
            debit = _parse_ledger_amount(m.group(3))
            kredit = _parse_ledger_amount(m.group(4))
            saldo = _parse_ledger_amount(m.group(5))

            tipe = "Jurnal Umum"
            keterangan = body
            if body.startswith("Transfer Bank"):
                tipe = "Transfer Bank"
                keterangan = re.sub(r"^Transfer Bank\s*", "", body).strip()
            elif body.startswith("Jurnal Umum"):
                keterangan = body.replace("Jurnal Umum", "", 1).strip()

            pending = {
                "tanggal": _parse_id_date(tanggal_raw),
                "tanggal_display": tanggal_raw,
                "tipe_transaksi": tipe,
                "pengajuan": keterangan,
                "debit": debit,
                "kredit": kredit,
                "saldo_akhir": saldo,
            }
            continue

        if pending and "pengajuan" in pending:
            pending["pengajuan"] = (pending["pengajuan"] + " " + line).strip()

    _commit_pending()

    if items:
        meta["saldo_akhir"] = items[-1].get("saldo_akhir") or 0
        meta["sisa_dana"] = meta["saldo_akhir"]
    if not meta["total_kredit"]:
        meta["total_kredit"] = sum(i["kredit"] for i in items)
    if not meta["total_debit"]:
        meta["total_debit"] = sum(i["debit"] for i in items)
    meta["total_digantikan"] = meta["total_kredit"]

    return {"meta": meta, "items": items}


def _extract_nota_images_from_pdf(pdf_path: str, start_page: int = 1) -> List[Any]:
    """Ekstrak foto nota dari halaman lampiran (portrait, urutan baca)."""
    from io import BytesIO
    from pypdf import PdfReader
    from PIL import Image

    images = []
    try:
        reader = PdfReader(pdf_path)
        for pno in range(start_page, len(reader.pages)):
            for img_obj in reader.pages[pno].images:
                try:
                    im = Image.open(BytesIO(img_obj.data))
                    w, h = im.size
                    if min(w, h) < 400:
                        continue
                    if h >= w * 0.95:
                        images.append(im)
                except Exception:
                    pass
    except Exception as e:
        print("Nota image extract error:", e)
    return images


def parse_reimbursement_petty_cash(pdf_path: str, upload_id: int, filename: str = "") -> Dict[str, Any]:
    """Parse FORM REIMBURSEMENT / FORM PETTY CASH + lampiran nota per transaksi."""
    import re
    import pdfplumber

    meta = {
        "report_type": "reimbursement",
        "nama_karyawan": None,
        "divisi": None,
        "saldo_awal": 0,
        "sisa_dana": 0,
        "saldo_akhir": 0,
        "total_digantikan": 0,
        "total_debit": 0,
        "total_kredit": 0,
        "payment_info": None,
        "bank": None,
        "nomor_rekening": None,
        "atas_nama": None,
        "yang_menyetujui": None,
        "tanggal_ttd_pemohon": None,
        "tanggal_ttd_menyetujui": None,
        "periode": None,
        "filename": filename,
    }

    import re
    fname_norm = filename.replace("_", " ")
    periode_m = re.search(
        r"(?:FORM\s+PETTY\s+CASH|REIMBURSEMENT|FORM)\s+(\d{1,2}\s+\w+\s+\d{4})",
        fname_norm, re.IGNORECASE,
    )
    if not periode_m:
        periode_m = re.search(r"(\d{1,2}\s+\w+\s+\d{4})", fname_norm, re.IGNORECASE)
    if periode_m:
        meta["periode"] = periode_m.group(1)

    skip_prefixes = (
        "FORM REIMBURSEMENT", "Tanggal Deskripsi", "PAYMENT",
    )
    awaiting_approver = False
    ttd_date_re = re.compile(r"(\d{1,2}/\d{1,2}/\d{4})")
    tx_re = re.compile(
        r"^(\d{1,2}\s+\w+\s+\d{4})\s+(.+?)\s+(Rp\s*[\d.,]+)$",
        re.IGNORECASE,
    )

    items = []
    with pdfplumber.open(pdf_path) as pdf:
        if not pdf.pages:
            return {"meta": meta, "items": items}
        lines = [l.strip() for l in (pdf.pages[0].extract_text() or "").split("\n") if l.strip()]

    for line in lines:
        if line.startswith("Nama Karyawan"):
            meta["nama_karyawan"] = line.replace("Nama Karyawan", "").strip()
        elif line.startswith("Divisi"):
            meta["divisi"] = line.replace("Divisi", "").strip()
        elif line.startswith("Saldo"):
            meta["saldo_awal"] = _parse_rp_amount(line)
        elif line.startswith("Sisa Dana"):
            amt = _parse_rp_amount(line.replace("-", ""))
            meta["sisa_dana"] = -amt if "-" in line else amt
        elif "Total yang Digantikan" in line:
            meta["total_digantikan"] = _parse_rp_amount(line)
        elif line.upper().startswith("PAYMENT"):
            meta["payment_info"] = line
            bm = re.search(r"(MANDIRI|BRI|BCA|BSI|BNI)\s+(\d+)", line, re.I)
            if bm:
                meta["bank"] = bm.group(1).upper()
                meta["nomor_rekening"] = bm.group(2)
            am = re.search(r"AN\s+(.+)$", line, re.I)
            if am:
                meta["atas_nama"] = am.group(1).strip()
            continue
        elif "Tanda Tangan Pemohon" in line:
            dm = ttd_date_re.search(line)
            if dm:
                meta["tanggal_ttd_pemohon"] = _parse_slash_date(dm.group(1))
            continue
        elif "yang menyetujui" in line.lower():
            dm = ttd_date_re.search(line)
            if dm:
                meta["tanggal_ttd_menyetujui"] = _parse_slash_date(dm.group(1))
            awaiting_approver = True
            continue
        elif awaiting_approver and not meta["yang_menyetujui"]:
            if not line.startswith("Tanda Tangan"):
                meta["yang_menyetujui"] = line.strip()
            awaiting_approver = False
            continue

        if any(line.startswith(p) for p in skip_prefixes):
            continue
        if line.startswith("Tanda Tangan"):
            continue
        if meta["nama_karyawan"] and line.upper() == meta["nama_karyawan"].upper():
            continue

        m = tx_re.match(line)
        if m:
            amt = _parse_rp_amount(m.group(3))
            if amt <= 0:
                continue
            desc = m.group(2).strip()
            items.append({
                "no": str(len(items) + 1),
                "tanggal": _parse_id_date(m.group(1)),
                "tanggal_display": m.group(1),
                "deskripsi": desc,
                "pengajuan": desc,
                "jumlah": amt,
                "kredit": amt,
                "debit": 0,
                "status": "DIAJUKAN",
                "rekening": "PETTY CASH",
                "tipe_transaksi": "Pengeluaran",
            })

    meta["total_kredit"] = sum(i["kredit"] for i in items)
    if not meta["total_digantikan"]:
        meta["total_digantikan"] = meta["total_kredit"]
    meta["saldo_akhir"] = meta["sisa_dana"]

    nota_dir = os.path.join(UPLOAD_DIR, "nota", str(upload_id))
    os.makedirs(nota_dir, exist_ok=True)
    nota_images = _extract_nota_images_from_pdf(pdf_path, start_page=1)

    for idx, item in enumerate(items):
        if idx < len(nota_images):
            fname = f"nota_{idx + 1:02d}.jpg"
            fpath = os.path.join(nota_dir, fname)
            nota_images[idx].convert("RGB").save(fpath, "JPEG", quality=88)
            item["nota_path"] = f"nota/{upload_id}/{fname}"

    return {"meta": meta, "items": items}


def parse_petty_cash_pdf(pdf_path: str, upload_id: int, filename: str = "") -> Dict[str, Any]:
    """Auto-detect format petty cash: Buku Besar Accurate atau Form Reimbursement."""
    import pdfplumber
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return {"meta": {}, "items": []}
            first_text = pdf.pages[0].extract_text() or ""
    except Exception as e:
        print("Petty cash PDF open error:", e)
        return {"meta": {}, "items": []}

    fname_up = filename.upper()
    if "FORM REIMBURSEMENT" in first_text or "FORM PETTY CASH" in fname_up or "REIMBURSEMENT" in fname_up:
        return parse_reimbursement_petty_cash(pdf_path, upload_id, filename)
    if "Rincian Buku Besar" in first_text or "11010101 - PETTY CASH" in first_text:
        return parse_buku_besar_petty_cash(pdf_path, upload_id, filename)
    return parse_reimbursement_petty_cash(pdf_path, upload_id, filename)

def parse_faktur_belum_lunas(pdf_path: str):
    """Parse PDF 'Faktur Belum Lunas' (Accurate) ke struktur kolom tagihan web.

    Mapping kolom PDF → database:
      - Baris sebelum PI.*           → pos (pemasok)
      - PI.2026.06.xxxxx             → no
      - Tanggal pertama pada baris PI → tanggal
      - PEMBELIAN ... / keterangan   → pengajuan
      - Jumlah (Total Utang)         → jumlah
      - MANDIRI/BRI/BCA/...          → bank
      - Nomor setelah bank           → nomor_rekening
      - Nama setelah nomor rekening  → atas_nama
    """
    import pdfplumber
    import re
    from dateutil import parser as date_parser

    BANK_RE = re.compile(r"\b(MANDIRI|BRI|BCA|BSI|VA|BNI)\s+(\d{5,})\s+(.+)$", re.IGNORECASE)
    PI_FULL_RE = re.compile(
        r"^(PI\.\d{4}\.\d{2}\.\d{5})\s+(\d{1,2}\s+\w+\s+\d{4})\s+(\d{1,2}\s+\w+\s+\d{4})\s+(.+)$"
    )
    PI_SHORT_RE = re.compile(r"^(PI\.\d{4}\.\d{2}\.\d{5})\s+(\d{1,2}\s+\w+\s+\d{4})\s+(.+)$")
    AMOUNT_RE = re.compile(r"(\d{1,3}(?:\.\d{3})+)")
    SUBTOTAL_RE = re.compile(r"^\d{1,3}(?:\.\d{3})+(?:\.\d{3})*\s+\d")
    SKIP_KEYWORDS = [
        "sppg wisma haji", "faktur belum lunas", "accurate", "tercetak", "halaman",
        "indonesian", "nomor #", "cabang :", "per tgl", "total utang", "rupiah",
        "daftar rekening pemasok", "jatuh tempo keterangan",
    ]

    def _parse_date(tgl_str: str):
        if not tgl_str:
            return None
        tgl = tgl_str.strip()
        norm = tgl.lower()
        for indo, eng in [("mei", "may"), ("agu", "aug"), ("agt", "aug"), ("okt", "oct"), ("des", "dec")]:
            norm = norm.replace(indo, eng)
        try:
            dt = date_parser.parse(norm, dayfirst=True, fuzzy=True)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
        try:
            p = tgl.lower().split()
            d = int(p[0])
            mon_str = p[1][:3]
            mon_map = {
                "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "mei": 5,
                "jun": 6, "jul": 7, "aug": 8, "agu": 8, "sep": 9,
                "oct": 10, "okt": 10, "nov": 11, "dec": 12, "des": 12,
            }
            mon = mon_map.get(mon_str, 1)
            y = int(p[2])
            return f"{y:04d}-{mon:02d}-{d:02d}"
        except Exception:
            return None

    def _should_skip(line: str) -> bool:
        low = line.lower()
        return any(kw in low for kw in SKIP_KEYWORDS)

    def _is_subtotal(line: str) -> bool:
        return bool(SUBTOTAL_RE.match(line))

    def _is_keterangan_continuation(line: str) -> bool:
        if line.startswith("PI.") or _is_subtotal(line) or _should_skip(line):
            return False
        if AMOUNT_RE.search(line) or BANK_RE.search(line):
            return False
        return len(line) <= 40

    def _is_supplier_line(line: str) -> bool:
        if line.startswith("PI.") or _is_subtotal(line) or _should_skip(line):
            return False
        if len(line) <= 2 or re.match(r"^\d", line):
            return False
        low = line.lower()
        if any(kw in low for kw in ["pembelian", "total", "peb", "report", "0 0", "rupiah", "wifi bulan"]):
            return False
        if re.search(r"\d", line) and len(line) < 28:
            return False
        if "," in line and len(line) < 30:
            return False
        return True

    def _is_supplier_continuation(line: str) -> bool:
        if line.startswith("PI.") or _is_subtotal(line) or _should_skip(line):
            return False
        if AMOUNT_RE.search(line) or BANK_RE.search(line):
            return False
        return len(line.split()) <= 3 and len(line) <= 24

    def _parse_pi_line(line: str, pemasok: str):
        m = PI_FULL_RE.match(line)
        if m:
            no, tgl_str, _, rest = m.group(1), m.group(2), m.group(3), m.group(4)
        else:
            m = PI_SHORT_RE.match(line)
            if not m:
                return None
            no, tgl_str, rest = m.group(1), m.group(2), m.group(3)

        tanggal = _parse_date(tgl_str)
        bank = ""
        rek = ""
        atas_nama = ""
        bank_m = BANK_RE.search(rest)
        if bank_m:
            bank = bank_m.group(1).upper()
            rek = bank_m.group(2)
            atas_nama = re.sub(r"\s+\d+\s+\d+\s*$", "", bank_m.group(3).strip()).strip()
            rest = rest[: bank_m.start()].strip()

        amount_m = AMOUNT_RE.search(rest)
        jumlah = 0
        keterangan = rest
        if amount_m:
            try:
                jumlah = int(amount_m.group(1).replace(".", ""))
            except Exception:
                jumlah = 0
            keterangan = rest[: amount_m.start()].strip()

        keterangan = re.sub(r"\s+", " ", keterangan).strip()
        if not keterangan or keterangan.isdigit():
            keterangan = f"Tagihan {pemasok}" if pemasok else f"Tagihan {no}"
        if jumlah <= 0:
            return None

        return {
            "no": no,
            "pos": pemasok or None,
            "pengajuan": keterangan,
            "jumlah": jumlah,
            "tanggal": tanggal,
            "bank": bank,
            "nomor_rekening": rek,
            "atas_nama": atas_nama or None,
            "status": "DIAJUKAN",
        }

    with pdfplumber.open(pdf_path) as pdf:
        all_lines = []
        for page in pdf.pages:
            text = page.extract_text() or ""
            all_lines.extend([l.strip() for l in text.split("\n") if l.strip()])

    items = []
    current_supplier = None
    pending_supplier_parts: List[str] = []
    awaiting_keterangan = False

    for line in all_lines:
        if not line or _should_skip(line) or _is_subtotal(line):
            awaiting_keterangan = False
            continue

        if line.startswith("PI."):
            pemasok = " ".join(pending_supplier_parts).strip() if pending_supplier_parts else current_supplier
            item = _parse_pi_line(line, pemasok)
            if item:
                items.append(item)
                if pemasok:
                    current_supplier = pemasok
                awaiting_keterangan = True
            pending_supplier_parts = []
            continue

        if awaiting_keterangan and items and _is_keterangan_continuation(line):
            items[-1]["pengajuan"] = f"{items[-1]['pengajuan']} {line}".strip()
            continue

        awaiting_keterangan = False

        if pending_supplier_parts and _is_supplier_continuation(line):
            pending_supplier_parts.append(line)
            continue

        if _is_supplier_line(line):
            pending_supplier_parts = [line]
            continue

    return items


def parse_pengajuan_dana_mitra_pdf(pdf_path: str, filename: str = "") -> Dict[str, Any]:
    """Parse formulir PDF 'Pengajuan Dana Mitra SPPG Wisma Haji'."""
    import pdfplumber
    import re
    from dateutil import parser as date_parser

    def _parse_date(tgl_str: str):
        if not tgl_str:
            return None
        tgl = tgl_str.strip()
        norm = tgl.lower()
        for indo, eng in [('mei', 'may'), ('agu', 'aug'), ('agt', 'aug'), ('okt', 'oct'), ('des', 'dec')]:
            norm = norm.replace(indo, eng)
        try:
            dt = date_parser.parse(norm, dayfirst=True, fuzzy=True)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
        try:
            p = tgl.lower().split()
            d = int(p[0])
            mon_str = p[1][:3]
            mon_map = {
                'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'mei': 5,
                'jun': 6, 'jul': 7, 'aug': 8, 'agu': 8, 'sep': 9,
                'oct': 10, 'okt': 10, 'nov': 11, 'dec': 12, 'des': 12,
            }
            mon = mon_map.get(mon_str, 1)
            y = int(p[2])
            return f"{y:04d}-{mon:02d}-{d:02d}"
        except Exception:
            return None

    lines: List[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines.extend([l.strip() for l in text.split("\n") if l.strip()])

    if not any("pengajuan dana mitra" in ln.lower() for ln in lines):
        return {"meta": {}, "items": []}

    meta: Dict[str, Any] = {"filename": filename}
    total_pengajuan = 0

    for line in lines:
        low = line.lower()
        if "no. form" in low or "no form" in low:
            m = re.search(r"No\.?\s*Form:?\s*(\d+)", line, re.IGNORECASE)
            if m:
                meta["no_form"] = m.group(1)
            m2 = re.search(r"Tanggal\s+Pengajuan:?\s*(.+?)(?:\s+Devisi|$)", line, re.IGNORECASE)
            if not m2:
                m2 = re.search(r"Tanggal\s+Pengajuan:?\s*(.+)$", line, re.IGNORECASE)
            if m2:
                meta["tanggal_pengajuan"] = _parse_date(m2.group(1).strip())
        if "pemohon" in low:
            m = re.search(r"Pemohon:?\s*(.+?)(?:\s+Devisi|$)", line, re.IGNORECASE)
            if m:
                meta["pemohon"] = m.group(1).strip()
            m2 = re.search(r"Devisi:?\s*(.+)$", line, re.IGNORECASE)
            if m2:
                meta["divisi"] = m2.group(1).strip()
        if "total pengajuan" in low:
            m = re.search(r"(\d{1,3}(?:\.\d{3})+)", line)
            if m:
                try:
                    total_pengajuan = int(m.group(1).replace(".", ""))
                except ValueError:
                    pass

    row_re = re.compile(
        r"^(\d+)\s+"
        r"(\d{1,2}\s+\w+\s+\d{4})\s+"
        r"(.+?)\s+"
        r"(\d{1,3}(?:\.\d{3})+)\s+"
        r"(\d{8,})\s+"
        r"(.+?)\s+"
        r"(Mandiri|BRI|BCA|BSI|BNI)\b"
        r"(?:\s+(.*))?$",
        re.IGNORECASE,
    )

    items: List[Dict[str, Any]] = []
    seen_keys = set()
    for line in lines:
        if line.lower().startswith("total pengajuan"):
            break
        m = row_re.match(line)
        if not m:
            continue
        jumlah = int(m.group(4).replace(".", ""))
        if jumlah <= 0:
            continue
        deskripsi = m.group(3).strip()
        key = (m.group(1), deskripsi, jumlah)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        items.append({
            "no": m.group(1),
            "tanggal": _parse_date(m.group(2)),
            "pengajuan": deskripsi,
            "jumlah": jumlah,
            "nomor_rekening": m.group(5),
            "atas_nama": m.group(6).strip(),
            "bank": m.group(7).upper(),
            "rekening": "PENGAJUAN DANA MITRA",
            "status": "DIAJUKAN",
            "ket": (m.group(8) or "").strip() or None,
        })

    if meta.get("no_form") and meta.get("tanggal_pengajuan"):
        tgl_label = meta["tanggal_pengajuan"]
        if meta.get("pemohon"):
            meta["periode"] = f"Form #{meta['no_form']} — {meta['pemohon']} ({tgl_label})"
        else:
            meta["periode"] = f"Form #{meta['no_form']} — {tgl_label}"
    elif meta.get("pemohon"):
        meta["periode"] = f"Pengajuan {meta['pemohon']}"
    else:
        meta["periode"] = filename or "Pengajuan Dana Mitra"

    if total_pengajuan:
        meta["total_pengajuan"] = total_pengajuan

    return {"meta": meta, "items": items}


def parse_upload_file(file_path: str, filename: str, kategori: str = "tagihan", upload_id: int = 0):
    """Parse Excel, CSV or PDF. Use special parser per kategori."""
    rows = []
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''

    if kategori == "gaji_relawan" and ext == "csv":
        try:
            return parse_gaji_relawan_csv(file_path, filename)
        except Exception as e:
            print("Gaji relawan CSV parse error:", e)
            return {"meta": {}, "items": []}

    if kategori == "gaji_relawan" and ext in ("xlsx", "xls"):
        try:
            return parse_gaji_relawan_xlsx(file_path, filename)
        except Exception as e:
            print("Gaji relawan XLSX parse error:", e)
            return {"meta": {}, "items": []}

    if kategori == "insentif_pic" and ext == "csv":
        try:
            return parse_insentif_pic_csv(file_path, filename)
        except Exception as e:
            print("Insentif PIC CSV parse error:", e)
            return {"meta": {}, "items": []}

    if kategori == "insentif_pic" and ext in ("xlsx", "xls"):
        try:
            return parse_insentif_pic_xlsx(file_path, filename)
        except Exception as e:
            print("Insentif PIC XLSX parse error:", e)
            return {"meta": {}, "items": []}
    if kategori == "insentif_mitra" and ext == "csv":
        try:
            return parse_insentif_mitra_csv(file_path, filename)
        except Exception as e:
            print("Insentif Mitra CSV parse error:", e)
            return {"meta": {}, "items": []}

    if kategori == "insentif_mitra" and ext in ("xlsx", "xls"):
        try:
            return parse_insentif_mitra_xlsx(file_path, filename)
        except Exception as e:
            print("Insentif Mitra XLSX parse error:", e)
            return {"meta": {}, "items": []}

    if ext == 'pdf' and kategori == "petty_cash":
        try:
            result = parse_petty_cash_pdf(file_path, upload_id, filename)
            return result
        except Exception as e:
            print("Petty cash PDF parse error:", e)
            return {"meta": {}, "items": []}

    if ext == 'pdf' and kategori == "pengajuan_dana_mitra":
        try:
            return parse_pengajuan_dana_mitra_pdf(file_path, filename)
        except Exception as e:
            print("Pengajuan dana mitra PDF parse error:", e)
            return {"meta": {}, "items": []}

    if ext == 'pdf':
        # Always try the special parser for known template PDFs first
        try:
            special = parse_faktur_belum_lunas(file_path)
            if special and any(item.get("jumlah", 0) > 0 for item in special):
                rows = special
            else:
                # General fallback for other PDFs: try to pull amounts from text
                import pdfplumber
                import re
                with pdfplumber.open(file_path) as pdf:
                    all_text = ""
                    for page in pdf.pages:
                        t = page.extract_text()
                        if t:
                            all_text += t + "\n"
                    lines = [l.strip() for l in all_text.split("\n") if l.strip()]
                    for line in lines[:20]:
                        m = re.search(r"([\d\.,]{4,})", line)  # look for reasonably sized numbers
                        if m:
                            try:
                                amt = int(m.group(1).replace(".", "").replace(",", ""))
                            except:
                                amt = None
                            if amt and amt > 100:
                                rows.append({
                                    "pengajuan": line[:200],
                                    "jumlah": amt,
                                    "tanggal": None,
                                    "status": "DIAJUKAN"
                                })
                    if not rows:
                        pengajuan = "Laporan dari PDF: " + " | ".join(lines[:3])
                        rows.append({
                            "pengajuan": pengajuan[:200],
                            "jumlah": 0,
                            "tanggal": None,
                            "status": "DIAJUKAN"
                        })
        except Exception as e:
            print("PDF parse error:", e)
            rows = []
    elif ext in ['xlsx', 'xls']:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                headers = [str(h).strip().lower() if h else '' for h in row]
                continue
            if not any(row):
                continue
            item = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                item[h] = val
            rows.append(item)
    else:
        # CSV
        import csv
        with open(file_path, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append({k.strip().lower(): v for k, v in row.items() if k})

    return rows

@app.get("/upload", response_class=HTMLResponse)
async def upload_page(request: Request, user=Depends(require_login), message: str = "", success: bool = False, pos: str = "", upload_id: int = 0):
    return render_template(request, "upload.html", {
        "user": user,
        "message": message,
        "success": success,
        "pos": pos,
        "upload_id": upload_id
    })

@app.post("/upload", response_class=HTMLResponse)
async def upload_laporan(
    request: Request,
    user=Depends(require_login),
    file: UploadFile = File(...),
    pos: str = Form(""),
    pos_other: str = Form(""),
    periode: str = Form(""),
    kategori: str = Form("tagihan")
):
    if is_viewer(user):
        return redirect_with_flash(
            request,
            "/dashboard",
            "Akses ditolak. Akun Anda hanya dapat melihat data.",
        )

    if not is_admin(user) and (kategori or "tagihan") not in ("tagihan", "", "tagihan_bulanan"):
        return redirect_with_flash(
            request,
            "/tagihan",
            "Akses ditolak. Member hanya dapat upload PDF Tagihan.",
        )

    if kategori == "petty_cash":
        return redirect_with_flash(
            request,
            "/petty-cash",
            "Upload Petty Cash hanya melalui halaman Petty Cash",
        )

    final_pos = (pos_other.strip() if pos_other.strip() else pos.strip()) or None

    # Save file
    safe_name = f"{date.today().isoformat()}_{file.filename.replace(' ', '_')}"
    file_path = os.path.join(UPLOAD_DIR, safe_name)
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    conn = get_db()

    # Create upload record first (needed for petty cash nota folder)
    conn.execute("""
        INSERT INTO uploads (filename, pos, periode, created_by)
        VALUES (?, ?, ?, ?)
    """, (file.filename, final_pos, periode or None, user["id"]))
    upload_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    # Parse
    try:
        parsed = parse_upload_file(file_path, file.filename, kategori or "tagihan", upload_id)
    except Exception as e:
        conn.close()
        if kategori == "gaji_relawan":
            return redirect_with_flash(request, "/gaji-relawan", f"Error parsing file: {str(e)}")
        return redirect_with_flash(request, f"/tagihan?kategori={kategori or 'tagihan'}", f"Error parsing file: {str(e)}")

    item_list = parsed if parsed else []

    if not item_list:
        conn.close()
        if kategori == "gaji_relawan":
            return redirect_with_flash(request, "/gaji-relawan", "Tidak ada data yang bisa dibaca dari file")
        return redirect_with_flash(request, f"/tagihan?kategori={kategori or 'tagihan'}", "Tidak ada data yang bisa dibaca dari file")

    inserted = 0
    for item in item_list:
        pengajuan = item.get('pengajuan') or item.get('deskripsi') or item.get('penajuan') or item.get('keterangan') or ''
        if not pengajuan:
            continue

        try:
            debit_val = int(item.get('debit') or 0)
            kredit_val = int(item.get('kredit') or 0)
            jumlah = int(float(item.get('jumlah') or item.get('total') or 0))
        except Exception:
            debit_val = kredit_val = jumlah = 0
        if not jumlah:
            jumlah = kredit_val if kredit_val > 0 else debit_val
        if jumlah <= 0 and debit_val <= 0 and kredit_val <= 0:
            continue

        no = str(item.get('no') or '').strip() or None
        status = str(item.get('status') or 'DIAJUKAN').strip() or 'DIAJUKAN'
        kategori_val = kategori or "tagihan"
        if kategori_val in ("tagihan", "tagihan_bulanan", ""):
            rekening = str(item.get('rekening') or '').strip() or None
        else:
            rekening = str(item.get('rekening') or 'PETTY CASH').strip() or 'PETTY CASH'
        pos_val = str(item.get('pos') or item.get('pemasok') or final_pos or '').strip() or None
        tanggal = str(item.get('tanggal') or '').strip() or None
        atas_nama = str(item.get('atas nama rek') or item.get('atas_nama') or item.get('atas nama') or '').strip() or None
        nomor_rek = str(item.get('nomor rekening') or item.get('nomor_rekening') or item.get('no rekening') or '').strip() or None
        bank = str(item.get('bank') or '').strip() or None
        nota_path = item.get('nota_path')
        saldo_akhir = item.get('saldo_akhir')
        tipe_transaksi = item.get('tipe_transaksi') or item.get('tipe')

        dup_check = conn.execute("""
            SELECT 1 FROM tagihan 
            WHERE COALESCE(no,'') = COALESCE(?, '') 
              AND pengajuan = ? 
              AND jumlah = ? 
              AND COALESCE(tanggal,'') = COALESCE(?, '') 
              AND kategori = ?
            LIMIT 1
        """, (no, pengajuan, jumlah, tanggal, kategori or "tagihan")).fetchone()

        if dup_check:
            continue

        conn.execute("""
            INSERT INTO tagihan 
            (no, pengajuan, jumlah, status, rekening, tanggal, atas_nama, nomor_rekening, bank, pos, kategori, upload_id, nota_path, debit, kredit, saldo_akhir, tipe_transaksi, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            no, pengajuan, jumlah, status, rekening, tanggal,
            atas_nama, nomor_rek, bank, pos_val,
            kategori or "tagihan", upload_id, nota_path,
            debit_val, kredit_val, saldo_akhir, tipe_transaksi,
            user["id"]
        ))
        inserted += 1

    conn.execute("UPDATE uploads SET record_count = ? WHERE id = ?", (inserted, upload_id))
    conn.commit()
    conn.close()

    msg = f"Berhasil mengunggah {inserted} baris data"
    if final_pos:
        msg += f" untuk Pos: {final_pos}"

    if kategori == "gaji_relawan":
        return redirect_with_flash(request, "/gaji-relawan", msg, success=True)
    if kategori == "insentif_mitra":
        return redirect_with_flash(request, "/insentif-mitra", msg, success=True)

    return redirect_with_flash(request, f"/tagihan?kategori={kategori or 'tagihan'}", msg, success=True)


# Need to import Response
from fastapi.responses import Response

# ===================== RUN =====================
if __name__ == "__main__":
    import uvicorn
    print("🚀 Menjalankan Sistem Pelaporan Keuangan SPPG Wisma Haji Madiun")
    print("   Akses: http://localhost:8001")
    print("   Login: swhm / A0312   (atau admin / sppg123 backup)")
    uvicorn.run(app, host="0.0.0.0", port=8001, reload=True)
